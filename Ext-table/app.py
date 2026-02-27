import os
import sys
import pandas as pd
import re
import json
import random
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, colors
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, jsonify, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename
import threading
import time
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create required directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Global variables for progress tracking
processing_progress = {'status': 'idle', 'progress': 0, 'message': '', 'error': ''}

class ExtractionProcessor:
    def __init__(self):
        print("="*50)
        print("EXTRACTION PROCESSOR INITIALIZED - LOADING SETTINGS")
        print("="*50)
        self.load_settings()
        print("="*50)
        print("SETTINGS LOADED COMPLETELY")
        print("="*50)
        self.swab_check_list = None
        self.tis_int_check_list = None
    
    def load_settings(self):
        """Load settings from JSON file if exists, otherwise start with empty settings"""
        try:
            settings_file = 'sample_settings.json'
            if os.path.exists(settings_file):
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
                    self.prefix_mapping = settings.get('prefix_mapping', {})
                    self.pool_samples = settings.get('pool_samples', [])
                    self.no_pool_samples = settings.get('no_pool_samples', [])
                    self.samples_per_pool = settings.get('samples_per_pool', None)
                    self.sample_id_pattern = settings.get('sample_id_pattern', None)
                    # Custom sample type classification
                    self.ana_samples = settings.get('ana_samples', [])
                    self.sal_samples = settings.get('sal_samples', [])
                    self.pt_samples = settings.get('pt_samples', [])
                    self.tis_samples = settings.get('tis_samples', [])
                    self.int_samples = settings.get('int_samples', [])
                    # H2O settings
                    self.enable_h2o_random = settings.get('enable_h2o_random', False)
                    self.h2o_count = settings.get('h2o_count', 2)
                    self.h2o_position_preference = settings.get('h2o_position_preference', 'random')
                    
                # Debug: Print loaded settings
                print(f"DEBUG: Loaded settings:")
                print(f"  prefix_mapping: {self.prefix_mapping}")
                print(f"  pool_samples: {self.pool_samples}")
                print(f"  no_pool_samples: {self.no_pool_samples}")
                print(f"  samples_per_pool: {self.samples_per_pool}")
                print(f"  sample_id_pattern: {self.sample_id_pattern}")
                print(f"  ana_samples: {self.ana_samples}")
                print(f"  sal_samples: {self.sal_samples}")
                print(f"  pt_samples: {self.pt_samples}")
                print(f"  tis_samples: {self.tis_samples}")
                print(f"  int_samples: {self.int_samples}")
                print(f"  enable_h2o_random: {self.enable_h2o_random}")
                print(f"  h2o_count: {self.h2o_count}")
                print(f"  h2o_position_preference: {self.h2o_position_preference}")
                    
                # Validate that required settings are present if file exists
                if not self.prefix_mapping:
                    raise ValueError("Prefix mapping is required")
                if not self.pool_samples and not self.no_pool_samples:
                    raise ValueError("Pool assignments are required")
                if not self.samples_per_pool:
                    raise ValueError("Samples per pool setting is required")
                if not self.sample_id_pattern:
                    raise ValueError("Sample ID pattern setting is required")
            else:
                # Settings file doesn't exist - start with empty settings
                print(f"DEBUG: Settings file not found, using empty settings")
                self.prefix_mapping = {}
                self.pool_samples = []
                self.no_pool_samples = []
                self.samples_per_pool = None
                self.sample_id_pattern = None
                self.ana_samples = []
                self.sal_samples = []
                self.pt_samples = []
                self.tis_samples = []
                self.int_samples = []
                # Default H2O settings
                self.enable_h2o_random = False
                self.h2o_count = 2
                self.h2o_position_preference = 'random'
                
        except Exception as e:
            print(f"Error loading settings: {e}")
            # Start with empty settings on error
            self.prefix_mapping = {}
            self.pool_samples = []
            self.no_pool_samples = []
            self.samples_per_pool = None
            self.sample_id_pattern = None
            self.ana_samples = []
            self.sal_samples = []
            self.pt_samples = []
            self.tis_samples = []
            self.int_samples = []
            # Default H2O settings
            self.enable_h2o_random = False
            self.h2o_count = 2
            self.h2o_position_preference = 'random'

    def transform_sample_id(self, sample_id):
        if not isinstance(sample_id, str):
            return sample_id
        for old_prefix, new_prefix in self.prefix_mapping.items():
            if old_prefix in sample_id:
                return sample_id.replace(old_prefix, new_prefix)
        return sample_id

    def reverse_transform_sample_id(self, sample_id):
        """Transform sample ID back to original form"""
        if not isinstance(sample_id, str):
            return str(sample_id)
        
        # Reverse the prefix mapping
        for new_prefix, old_prefix in [(v, k) for k, v in self.prefix_mapping.items()]:
            if sample_id.startswith(new_prefix):
                return sample_id.replace(new_prefix, old_prefix)
        return sample_id

    def split_sample_id(self, sample_id):
        if not isinstance(sample_id, str):
            return str(sample_id)
        
        # Remove underscore and split into prefix and number
        if '_' in sample_id:
            parts = sample_id.split('_')
            if len(parts) == 2:
                prefix = parts[0]
                number = parts[1]
                return f"{prefix}\n{number}"  # Two lines without underscore
        
        # Fallback for other formats
        sample_id = ''.join(c for c in sample_id if c.isalnum() or c in ['_', '-'])
        try:
            match = re.match(self.sample_id_pattern, sample_id)
            if match:
                prefix = match.group(1) + match.group(2)
                number = match.group(3)
                return f"{prefix}\n{number}"
            else:
                return sample_id
        except:
            return sample_id

    def format_cell_for_sample_id(self, cell, worksheet):
        worksheet.row_dimensions[cell.row].height = 35  # Increase height for two lines
        column_letter = get_column_letter(cell.column)
        worksheet.column_dimensions[column_letter].width = 12  # Adjust width for better fit
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable wrap_text for two lines
        cell.number_format = '@'

    def get_sample_extraction_id(self, sample_ids):
        if not sample_ids:
            return "No samples"
        sample_groups = {}
        for sid in sample_ids:
            if not isinstance(sid, str):
                continue
            try:
                match = re.match(r'([A-Z]+)(\d{2})_(\d+)', sid)
                if match:
                    prefix = match.group(1) + match.group(2)
                    if prefix not in sample_groups:
                        sample_groups[prefix] = []
                    number = ''.join(c for c in match.group(3) if c.isdigit())
                    if number:
                        sample_groups[prefix].append(sid)
            except Exception:
                continue
        if not sample_groups:
            return "No valid samples"
        result = []
        reverse_mapping = {v: k for k, v in self.prefix_mapping.items()}
        for prefix, sids in sample_groups.items():
            if prefix not in reverse_mapping:
                continue
            original_prefix = reverse_mapping[prefix]
            try:
                numbers = [int(''.join(c for c in sid.split('_')[-1] if c.isdigit())) for sid in sids if ''.join(c for c in sid.split('_')[-1] if c.isdigit())]
                if numbers:
                    min_num = min(numbers)
                    max_num = max(numbers)
                    range_str = f"{original_prefix}_{min_num:03d} to {max_num:03d}"
                    result.append(range_str)
            except Exception:
                continue
        return " and ".join(result) if result else "No samples in range"

    def generate_h2o_positions(self, max_rows=8, num_columns=12, plate_key=None):
        """Generate random H2O positions for a plate based on settings"""
        print(f"DEBUG: generate_h2o_positions called with enable_h2o_random={self.enable_h2o_random}, plate_key={plate_key}")
        
        if not self.enable_h2o_random:
            print("DEBUG: H2O random is disabled, returning empty positions")
            return []
        
        # Use plate_key as seed for reproducible randomness across plates
        if plate_key is None:
            plate_key = str(time.time())
        
        random.seed(hash(plate_key))
        
        # Generate all possible positions (row, col) 0-indexed
        all_positions = []
        for row in range(max_rows):
            for col in range(num_columns):
                all_positions.append((row, col))
        
        # Filter positions based on preference
        if self.h2o_position_preference == 'edges':
            # Edge positions: first/last row or first/last column
            edge_positions = []
            for row in range(max_rows):
                for col in range(num_columns):
                    if row == 0 or row == max_rows - 1 or col == 0 or col == num_columns - 1:
                        edge_positions.append((row, col))
            available_positions = edge_positions
        elif self.h2o_position_preference == 'center':
            # Center positions: not on edges
            center_positions = []
            for row in range(1, max_rows - 1):
                for col in range(1, num_columns - 1):
                    center_positions.append((row, col))
            available_positions = center_positions if center_positions else all_positions
        else:  # random
            available_positions = all_positions
        
        # Ensure we don't try to place more H2O than available positions
        h2o_count = min(self.h2o_count, len(available_positions))
        
        # Shuffle positions for true randomness
        random.shuffle(available_positions)
        
        # Apply column spacing constraint while maintaining randomness
        filtered_positions = []
        used_columns = set()
        
        for pos in available_positions:
            row, col = pos
            # Check if this column is at least 2 columns away from existing H2O columns
            column_valid = True
            for used_col in used_columns:
                if abs(col - used_col) < 2:  # Less than 2 columns apart
                    column_valid = False
                    break
            
            if column_valid:
                filtered_positions.append(pos)
                used_columns.add(col)
                if len(filtered_positions) >= h2o_count:
                    break
        
        # If we couldn't place all H2O with spacing, try without spacing constraint
        if len(filtered_positions) < h2o_count:
            remaining_needed = h2o_count - len(filtered_positions)
            # Shuffle again and pick remaining positions without spacing constraint
            remaining_positions = [pos for pos in available_positions if pos not in filtered_positions]
            random.shuffle(remaining_positions)
            filtered_positions.extend(remaining_positions[:remaining_needed])
        
        # Take only the required number and DON'T sort - keep random order
        h2o_positions = filtered_positions[:h2o_count]
        
        print(f"DEBUG: Generated H2O positions: {h2o_positions}")
        return h2o_positions

    def create_table_with_h2o(self, samples, max_rows=8, num_columns=12, plate_key=None, format_for_pcr_cdna=False):
        """Create a table layout with H2O positions properly integrated"""
        if not self.enable_h2o_random:
            # Create standard table without H2O
            return self.create_standard_table(samples, max_rows, num_columns, format_for_pcr_cdna)
        
        # Generate H2O positions first
        h2o_positions = self.generate_h2o_positions(max_rows, num_columns, plate_key)
        
        # Create empty table
        table = [['' for _ in range(num_columns)] for _ in range(max_rows)]
        
        # Mark H2O positions
        h2o_set = set(h2o_positions)
        for row, col in h2o_positions:
            table[row][col] = 'H2O'
        
        # Create pools column by column based on actual H2O positions
        pool_assignments = {}
        
        # Use a global pool counter that persists across plates
        if not hasattr(self, 'global_pool_count'):
            self.global_pool_count = 1
        
        sample_index = 0
        
        print(f"DEBUG: Creating pools column by column for {len(samples)} samples, starting from Pool_{self.global_pool_count:03}")
        
        for col in range(num_columns):
            # Check if this column has H2O
            has_h2o = any(pos[1] == col for pos in h2o_positions)
            column_capacity = max_rows - 1 if has_h2o else max_rows
            
            print(f"DEBUG: Column {col+1} - Has H2O: {has_h2o}, Capacity: {column_capacity}")
            
            # Create pool for this column if we have samples
            if sample_index < len(samples) and self.samples_per_pool:
                pool_name = f"Pool_{self.global_pool_count:03}"
                samples_in_this_column = min(column_capacity, len(samples) - sample_index)
                
                print(f"DEBUG: Creating {pool_name} with {samples_in_this_column} samples in column {col+1}")
                
                # Place samples in this column
                placed_in_column = 0
                for row in range(max_rows):
                    if table[row][col] == '' and sample_index < len(samples):
                        sample_name = samples[sample_index]
                        if format_for_pcr_cdna and '_' in sample_name:
                            sample_name = sample_name.replace('_', '\n')
                        
                        table[row][col] = sample_name
                        pool_assignments[samples[sample_index]] = pool_name
                        
                        sample_index += 1
                        placed_in_column += 1
                        
                        if placed_in_column >= column_capacity:
                            break
                
                self.global_pool_count += 1
        
        print(f"DEBUG: Created {self.global_pool_count - 1} pools, placed {sample_index} samples")
        
        # Store pool assignments for later use
        if not hasattr(self, 'column_pool_assignments'):
            self.column_pool_assignments = {}
        self.column_pool_assignments[plate_key] = pool_assignments
        
        return table

    def create_standard_table(self, samples, max_rows=8, num_columns=12, format_for_pcr_cdna=False):
        """Create standard table without H2O"""
        table = []
        for col in range(num_columns):
            column = []
            for row in range(max_rows):
                index = col * max_rows + row
                if index < len(samples):
                    sample_name = samples[index]
                    if format_for_pcr_cdna and '_' in sample_name:
                        sample_name = sample_name.replace('_', '\n')
                    column.append(sample_name)
                else:
                    column.append('')
            table.append(column)
        
        # Transpose to get rows
        table = list(zip(*table))
        table = [list(row) for row in table]
        return table

    def get_existing_done_samples(self, destination_file):
        """Get list of sample IDs that are already marked as Done in the output file"""
        try:
            if not os.path.exists(destination_file):
                return set()
            
            # Try to read the Excel file
            try:
                df = pd.read_excel(destination_file, sheet_name='Extraction', engine='openpyxl')
            except:
                # Try without sheet specification
                df = pd.read_excel(destination_file)
            
            if df.empty or 'Sample_Id' not in df.columns or 'Status' not in df.columns:
                return set()
            
            # Get samples with Done status
            done_samples = df[df['Status'] == 'Done']['Sample_Id'].tolist()
            print(f"DEBUG: Found {len(done_samples)} samples already marked as Done")
            return set(done_samples)
            
        except Exception as e:
            print(f"DEBUG: Error reading existing Done samples: {e}")
            return set()

    def pre_save_formatting(self, worksheet):
        for row in worksheet.iter_rows(min_row=8, max_row=15, min_col=2, max_col=13):
            for cell in row:
                if cell.value and '\n' in str(cell.value):
                    parts = str(cell.value).split('\n')
                    cell.value = parts[0] + chr(10) + parts[1]
                    self.format_cell_for_sample_id(cell, worksheet)
        for row in worksheet.iter_rows(min_row=26, max_row=33, min_col=2, max_col=13):
            for cell in row:
                if cell.value and '\n' in str(cell.value):
                    parts = str(cell.value).split('\n')
                    cell.value = parts[0] + chr(10) + parts[1]
                    self.format_cell_for_sample_id(cell, worksheet)
        for row in worksheet.iter_rows(min_row=44, max_row=51, min_col=2, max_col=13):
            for cell in row:
                if cell.value and '\n' in str(cell.value):
                    parts = str(cell.value).split('\n')
                    cell.value = parts[0] + chr(10) + parts[1]
                    self.format_cell_for_sample_id(cell, worksheet)

    def process_ana_sal_samples(self, destination_file, current_step, total_steps):
        try:
            processing_progress['message'] = f"Loading settings: pool_samples={len(self.pool_samples)}, no_pool_samples={len(self.no_pool_samples)}"
            
            # Get existing Done samples to skip them
            done_samples = self.get_existing_done_samples(destination_file)
            processing_progress['message'] = f"Found {len(done_samples)} samples already completed, will skip them"
            
            # Try to read Excel file with comprehensive error handling
            try:
                # First, check if file exists and is readable
                if not os.path.exists(destination_file):
                    raise FileNotFoundError(f"Destination file not found: {destination_file}")
                
                # Check file extension and try to handle different formats
                file_ext = os.path.splitext(destination_file)[1].lower()
                processing_progress['message'] = f"Reading Excel file (format: {file_ext})..."
                
                # Try different approaches to read the Excel file
                df = None
                error_messages = []
                
                # Method 1: Try with openpyxl (most reliable for .xlsx)
                try:
                    df = pd.read_excel(destination_file, sheet_name='Extraction', engine='openpyxl')
                    processing_progress['message'] = f"Successfully read Excel with openpyxl engine"
                except Exception as e:
                    error_messages.append(f"openpyxl failed: {str(e)}")
                
                # Method 2: Try with default engine
                if df is None:
                    try:
                        df = pd.read_excel(destination_file, sheet_name='Extraction')
                        processing_progress['message'] = f"Successfully read Excel with default engine"
                    except Exception as e:
                        error_messages.append(f"Default engine failed: {str(e)}")
                
                # Method 3: Try reading without specifying sheet name first
                if df is None:
                    try:
                        # Try a more direct approach without ExcelFile
                        try:
                            df = pd.read_excel(destination_file, sheet_name=0)  # Try first sheet by index
                            processing_progress['message'] = f"Successfully read first sheet by index"
                        except Exception as index_error:
                            error_messages.append(f"Index-based reading failed: {str(index_error)}")
                            
                            # Try with different parameters
                            try:
                                df = pd.read_excel(destination_file, header=None)  # Read without header assumptions
                                processing_progress['message'] = f"Successfully read Excel without header processing"
                            except Exception as header_error:
                                error_messages.append(f"No-header reading failed: {str(header_error)}")
                                
                                # Last resort - try with all available engines
                                engines_to_try = ['openpyxl', 'xlrd', None]
                                for engine in engines_to_try:
                                    try:
                                        if engine:
                                            df = pd.read_excel(destination_file, engine=engine)
                                        else:
                                            df = pd.read_excel(destination_file)
                                        processing_progress['message'] = f"Successfully read with engine: {engine or 'default'}"
                                        break
                                    except Exception as engine_error:
                                        error_messages.append(f"Engine {engine or 'default'} failed: {str(engine_error)}")
                                        continue
                    
                    except Exception as e:
                        error_messages.append(f"Direct reading method failed: {str(e)}")
                
                # Method 4: Last resort - try to read any Excel data
                if df is None:
                    try:
                        # Try reading without sheet specification
                        df = pd.read_excel(destination_file)
                        processing_progress['message'] = f"Read Excel without sheet specification"
                    except Exception as e:
                        error_messages.append(f"Last resort method failed: {str(e)}")
                
                # If all methods failed, provide detailed error and skip processing
                if df is None:
                    processing_progress['error'] = f"All Excel reading methods failed. File: {destination_file}. Errors: {'; '.join(error_messages)}"
                    processing_progress['message'] = "Skipping ANA/SAL processing due to Excel read error"
                    return  # Skip ANA/SAL processing but don't crash
                
                # Validate dataframe
                if df.empty:
                    processing_progress['message'] = "Excel file is empty - skipping ANA/SAL processing"
                    return
                    
                processing_progress['message'] = f"Excel loaded successfully. Shape: {df.shape}"
                    
            except Exception as e:
                processing_progress['error'] = f'Excel reading error: {str(e)}'
                processing_progress['message'] = "Skipping ANA/SAL processing due to Excel read error"
                return  # Skip processing but continue with other steps
                    
            processing_progress['message'] = f"Processing ANA/SAL samples..."
            
            # Use custom sample type classification instead of automatic detection
            if hasattr(self, 'ana_samples') and self.ana_samples:
                ana_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.ana_samples]
            else:
                ana_prefixes = []
            
            if hasattr(self, 'sal_samples') and self.sal_samples:
                sal_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.sal_samples]
            else:
                sal_prefixes = []
                
            if hasattr(self, 'pt_samples') and self.pt_samples:
                pt_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.pt_samples]
            else:
                pt_prefixes = []
            print(f"DEBUG: Prefix breakdown:")
            print(f"  ANA prefixes: {ana_prefixes}")
            print(f"  SAL prefixes: {sal_prefixes}")
            print(f"  PT prefixes: {pt_prefixes}")
            swab_prefixes = ana_prefixes + sal_prefixes + pt_prefixes
            swab_pattern = '|'.join(swab_prefixes)
            # Filter out samples that are already Done
            available_samples = df[~df['Sample_Id'].isin(done_samples)]
            ana_sal_pt_samples = available_samples[available_samples['Sample_Id'].str.contains(swab_pattern, na=False) & (available_samples['Status'] == 'in process')]
            
            # Filter ANA samples (transformed prefixes from original ANA prefixes)
            if ana_prefixes:
                ana_sample_ids = ana_sal_pt_samples[ana_sal_pt_samples['Sample_Id'].str.contains('|'.join(ana_prefixes), na=False)]['Sample_Id'].tolist()
            else:
                ana_sample_ids = []
            ana_sample_ids.sort()
            processing_progress['message'] = f"ANA samples found: {len(ana_sample_ids)}"
            
            # Filter SAL samples (transformed prefixes from original SAL prefixes)
            if sal_prefixes:
                sal_sample_ids = ana_sal_pt_samples[ana_sal_pt_samples['Sample_Id'].str.contains('|'.join(sal_prefixes), na=False)]['Sample_Id'].tolist()
            else:
                sal_sample_ids = []
            sal_sample_ids.sort()
            processing_progress['message'] = f"SAL samples found: {len(sal_sample_ids)}"
            
            # Filter PT samples (transformed prefixes from original CANA_PT prefixes)
            if pt_prefixes:
                pt_sample_ids = ana_sal_pt_samples[ana_sal_pt_samples['Sample_Id'].str.contains('|'.join(pt_prefixes), na=False)]['Sample_Id'].tolist()
            else:
                pt_sample_ids = []  # No PT prefixes, so no PT samples
            pt_sample_ids.sort()
            print(f"DEBUG: PT prefixes: {pt_prefixes}")
            print(f"DEBUG: PT pattern: {'|'.join(pt_prefixes) if pt_prefixes else 'EMPTY'}")
            print(f"DEBUG: PT sample_ids: {len(pt_sample_ids)}")
            if len(pt_sample_ids) > 0:
                print(f"DEBUG: PT sample examples: {pt_sample_ids[:5]}")
            processing_progress['message'] = f"PT samples found: {len(pt_sample_ids)}"
            
            ordered_sample_ids = []
            # Order by original source based on transformed prefixes
            canb_prefixes = [new_prefix for old_prefix, new_prefix in self.prefix_mapping.items() if 'CANB_' in old_prefix]
            canb_samples = [sid for sid in ana_sample_ids + sal_sample_ids if any(p in sid for p in canb_prefixes)]
            ordered_sample_ids.extend(sorted(canb_samples))
            canr_prefixes = [new_prefix for old_prefix, new_prefix in self.prefix_mapping.items() if 'CANR_' in old_prefix]
            canr_samples = [sid for sid in ana_sample_ids + sal_sample_ids if any(p in sid for p in canr_prefixes)]
            ordered_sample_ids.extend(sorted(canr_samples))
            iplnahl_prefixes = [new_prefix for old_prefix, new_prefix in self.prefix_mapping.items() if 'IPLNAHL_' in old_prefix]
            iplnahl_samples = [sid for sid in ana_sample_ids + sal_sample_ids if any(p in sid for p in iplnahl_prefixes)]
            ordered_sample_ids.extend(sorted(iplnahl_samples))
            ordered_sample_ids.extend(sorted(pt_sample_ids))
            # Remove duplicates from ordered_sample_ids to prevent multiple assignments
            print(f"DEBUG: Before deduplication: {len(ordered_sample_ids)} samples")
            ordered_sample_ids = list(dict.fromkeys(ordered_sample_ids))
            print(f"DEBUG: After deduplication: {len(ordered_sample_ids)} samples")
            print(f"DEBUG: Sample IDs: {ordered_sample_ids[:10]}...")  # Show first 10
            transformed_sample_ids = [self.transform_sample_id(sid) for sid in ordered_sample_ids]
            pool_assignments = {}
            
            # Apply custom pool assignments from environment variables
            # First, handle explicitly pooled samples
            pooled_samples = []
            not_pooled_samples = []
            auto_pooled_samples = []
            
            for sample_id in ordered_sample_ids:
                # Transform pool prefixes to match transformed sample IDs
                transformed_pool_prefixes = [self.transform_sample_id(prefix) for prefix in self.pool_samples]
                transformed_no_pool_prefixes = [self.transform_sample_id(prefix) for prefix in self.no_pool_samples]
                
                # Debug: Print what we're checking
                print(f"DEBUG: Checking sample {sample_id} against pool prefixes {transformed_pool_prefixes}")
                
                # Check if sample ID starts with any transformed pooled prefix
                is_pooled = any(sample_id.startswith(pool_prefix) for pool_prefix in transformed_pool_prefixes)
                is_not_pooled = any(sample_id.startswith(no_pool_prefix) for no_pool_prefix in transformed_no_pool_prefixes)
                
                print(f"DEBUG: Sample {sample_id} - is_pooled: {is_pooled}, is_not_pooled: {is_not_pooled}")
                
                if is_pooled:
                    pooled_samples.append(sample_id)
                elif is_not_pooled:
                    not_pooled_samples.append(sample_id)
                else:
                    auto_pooled_samples.append(sample_id)
            
            # Assign pools to explicitly pooled samples (8 samples per pool)
            print(f"DEBUG: Pool assignment summary:")
            print(f"  Pooled samples: {len(pooled_samples)}")
            print(f"  Not pooled samples: {len(not_pooled_samples)}")
            print(f"  Auto pooled samples: {len(auto_pooled_samples)}")
            
            pool_count = 1
            # Group pooled samples column by column based on H2O positions
            if self.enable_h2o_random and self.h2o_count > 0:
                print(f"DEBUG: Column-by-column pooling - H2O count: {self.h2o_count}")
                # When H2O is enabled, defer pool creation to table creation phase
                # This allows pools to be created based on actual column capacity
                print(f"DEBUG: Deferring pool creation for {len(pooled_samples)} samples to table creation")
                # Create a placeholder pool assignment that will be updated during table creation
                for sample_id in pooled_samples:
                    df.loc[df['Sample_Id'] == sample_id, 'Pool'] = "TO_BE_ASSIGNED"
                    transformed_sample = self.transform_sample_id(sample_id)
                    pool_assignments[transformed_sample] = "TO_BE_ASSIGNED"
            else:
                # Standard pooling when H2O is disabled
                pool_count = 1
                for i in range(0, len(pooled_samples), self.samples_per_pool):
                    pool_name = f"Pool_{pool_count:03}"
                    pool_sample_ids = pooled_samples[i:i + self.samples_per_pool]
                    for sample_id in pool_sample_ids:
                        print(f"DEBUG: Assigning {pool_name} to sample {sample_id}")
                        df.loc[df['Sample_Id'] == sample_id, 'Pool'] = pool_name
                        transformed_sample = self.transform_sample_id(sample_id)
                        pool_assignments[transformed_sample] = pool_name
                    pool_count += 1
            
            # Assign "Not Pool" to samples that should not be pooled
            for sample_id in not_pooled_samples:
                df.loc[df['Sample_Id'] == sample_id, 'Pool'] = "Not Pool"
                transformed_sample = self.transform_sample_id(sample_id)
                pool_assignments[transformed_sample] = "Not Pool"
            
            # No automatic pooling - only samples explicitly in pool_samples get pooled
            # All other samples that weren't explicitly assigned get "Not Pool"
            remaining_samples = set(ordered_sample_ids) - set(pooled_samples) - set(not_pooled_samples)
            for sample_id in remaining_samples:
                df.loc[df['Sample_Id'] == sample_id, 'Pool'] = "Not Pool"
                transformed_sample = self.transform_sample_id(sample_id)
                pool_assignments[transformed_sample] = "Not Pool"
            
            # Create a mapping from original to transformed sample IDs
            sample_id_mapping = {original: transformed for original, transformed in zip(ordered_sample_ids, transformed_sample_ids)}
            # Update the dataframe with transformed sample IDs
            print(f"DEBUG: Before transformation - DataFrame has {len(df)} rows")
            print(f"DEBUG: Sample IDs to transform: {len(ordered_sample_ids)}")
            for original, transformed in zip(ordered_sample_ids, transformed_sample_ids):
                df.loc[df['Sample_Id'] == original, 'Sample_Id'] = transformed
            print(f"DEBUG: After transformation - DataFrame has {len(df)} rows")
            
            os.makedirs('Extraction', exist_ok=True)
            ana_sal_pt_samples_copy = ana_sal_pt_samples.copy()
            original_pool_mapping = {}
            
            # Rebuild pool mapping using column-by-column logic
            pool_count = 1
            original_pool_mapping = {}
            
            # Handle pooled samples
            if self.enable_h2o_random and self.h2o_count > 0:
                print(f"DEBUG: Deferring pool creation for check list - will be handled in table creation")
                # For pooled samples, create placeholder assignments
                for sample_id in pooled_samples:
                    original_pool_mapping[sample_id] = "TO_BE_ASSIGNED"
            else:
                # Standard pooling when H2O is disabled
                for sample_id in pooled_samples:
                    pool_name = f"Pool_{pool_count:03}"
                    original_pool_mapping[sample_id] = pool_name
                    pool_count += 1
            
            # Not pooled samples
            for sample_id in not_pooled_samples:
                original_pool_mapping[sample_id] = "Not Pool"
            
            # Auto-pooled samples
            if self.enable_h2o_random and self.h2o_count > 0:
                print(f"DEBUG: Deferring auto-pool creation for check list - will be handled in table creation")
                # For auto-pooled samples, create placeholder assignments
                for sample_id in auto_pooled_samples:
                    original_pool_mapping[sample_id] = "TO_BE_ASSIGNED"
            else:
                # Standard auto-pooling when H2O is disabled
                for i in range(0, len(auto_pooled_samples), self.samples_per_pool):
                    pool_name = f"Pool_{pool_count:03}"
                    pool_sample_ids = auto_pooled_samples[i:i + self.samples_per_pool]
                    for sid in pool_sample_ids:
                        original_pool_mapping[sid] = pool_name
                    pool_count += 1
            
            # Map transformed IDs to pools directly
            transformed_pool_mapping = {sample_id_mapping.get(sid, sid): pool for sid, pool in original_pool_mapping.items()}
            # Update sample IDs in the check list dataframe
            ana_sal_pt_samples_copy['Sample_Id'] = ana_sal_pt_samples_copy['Sample_Id'].map(lambda x: sample_id_mapping.get(x, x))
            # Assign pool values using the transformed IDs
            ana_sal_pt_samples_copy['Pool'] = ana_sal_pt_samples_copy['Sample_Id'].map(transformed_pool_mapping)
            
            def get_sort_key(sample_id):
                if 'CANB_' in sample_id:
                    return (0, sample_id)
                elif 'CANR_' in sample_id:
                    return (1, sample_id)
                elif 'IPLNAHL_' in sample_id:
                    return (2, sample_id)
                else:
                    return (3, sample_id)
            
            ana_sal_pt_samples_copy['sort_key'] = ana_sal_pt_samples_copy['Sample_Id'].apply(get_sort_key)
            ana_sal_pt_samples_copy = ana_sal_pt_samples_copy.sort_values('sort_key')
            ana_sal_pt_samples_copy = ana_sal_pt_samples_copy.drop('sort_key', axis=1)
            
            # Track plate assignments for each sample
            sample_plate_assignments = {}
            
            extraction_tables = []
            num_columns = 12
            max_rows_per_column = 8
            items_per_plate = num_columns * max_rows_per_column
            
            # Get the starting plate number from existing Extraction_Tables.xlsx
            file_path = 'Extraction/Extraction_Tables.xlsx'
            if not os.path.exists(file_path):
                next_index = 1  # Start with Plate_001 when file doesn't exist
            else:
                try:
                    existing_sheets = pd.ExcelFile(file_path).sheet_names
                    # Extract existing plate numbers and find the highest
                    existing_numbers = []
                    for sheet in existing_sheets:
                        if sheet.startswith('Plate_'):
                            try:
                                num = int(sheet.replace('Plate_', ''))
                                existing_numbers.append(num)
                            except ValueError:
                                continue
                    next_index = max(existing_numbers) + 1 if existing_numbers else 1
                except Exception as e:
                    print(f"DEBUG: Error reading existing sheets: {e}")
                    next_index = 1
            
            # Create tables with H2O integration
            print(f"DEBUG: Creating {len(transformed_sample_ids)} extraction samples with H2O enabled={self.enable_h2o_random}")
            
            # Calculate how many samples can fit per plate accounting for H2O
            total_positions = num_columns * max_rows_per_column  # 96 positions
            samples_per_plate = total_positions - self.h2o_count if self.enable_h2o_random else total_positions
            print(f"DEBUG: Total positions per plate: {total_positions}, H2O count: {self.h2o_count}, Samples per plate: {samples_per_plate}")
            
            plate_index = 0
            for plate_start in range(0, len(transformed_sample_ids), samples_per_plate):
                plate_samples = transformed_sample_ids[plate_start:plate_start + samples_per_plate]
                plate_key = f"Extraction_Plate_{next_index + plate_index}"
                print(f"DEBUG: Creating Extraction table {plate_key} with {len(plate_samples)} samples (should be {samples_per_plate})")
                
                # Create table with H2O positions properly integrated
                current_table = self.create_table_with_h2o(plate_samples, max_rows_per_column, num_columns, plate_key)
                print(f"DEBUG: Created table with {sum(1 for row in current_table for cell in row if cell == 'H2O')} H2O positions")
                
                # Update pool assignments from table creation if H2O is enabled
                if self.enable_h2o_random and hasattr(self, 'column_pool_assignments') and plate_key in self.column_pool_assignments:
                    pool_assignments_from_table = self.column_pool_assignments[plate_key]
                    print(f"DEBUG: Updating {len(pool_assignments_from_table)} pool assignments from table creation")
                    
                    # Update the main pool_assignments dictionary
                    for sample_id, pool_name in pool_assignments_from_table.items():
                        pool_assignments[sample_id] = pool_name
                    
                    # Update the dataframe as well
                    for sample_id, pool_name in pool_assignments_from_table.items():
                        df.loc[df['Sample_Id'] == sample_id, 'Pool'] = pool_name
                
                # Assign plate numbers to samples (skip H2O positions)
                for row_idx, row in enumerate(current_table):
                    for col_idx, cell in enumerate(row):
                        if cell and cell != 'H2O':
                            sample_plate_assignments[cell] = f'Plate_{next_index + plate_index:03d}'
                
                extraction_tables.append(current_table)
                plate_index += 1
            
            # After all plates are created, update the check list with final pool assignments
            print(f"DEBUG: Final pool assignments: {len(pool_assignments)} entries")
            
            # Add plate information to check list
            ana_sal_pt_samples_copy['Plate'] = ana_sal_pt_samples_copy['Sample_Id'].map(sample_plate_assignments)
            
            # Update pool assignments in check list dataframe
            for transformed_id, pool_name in pool_assignments.items():
                ana_sal_pt_samples_copy.loc[ana_sal_pt_samples_copy['Sample_Id'] == transformed_id, 'Pool'] = pool_name
            
            # Reorder columns to include Plate
            columns = ana_sal_pt_samples_copy.columns.tolist()
            print(f"DEBUG: Available columns: {columns}")
            
            # Check if columns exist before trying to remove them
            for col in ['Pool', 'Status', 'Plate']:
                if col in columns:
                    columns.remove(col)
                    print(f"DEBUG: Removed column {col}")
                else:
                    print(f"DEBUG: Column {col} not found in dataframe")
            
            columns.extend(['Pool', 'Plate', 'Status'])
            ana_sal_pt_samples_copy = ana_sal_pt_samples_copy[columns]
            
            # Add logging to indicate what's being saved
            processing_progress['message'] = f"Writing {len(ana_sal_pt_samples_copy)} samples to Check_list_Swab-sample-table.xlsx"
            
            # Debug output first few rows of data to verify sample IDs
            if not ana_sal_pt_samples_copy.empty:
                sample_examples = ana_sal_pt_samples_copy['Sample_Id'].head(5).tolist()
                processing_progress['message'] = f"Sample ID examples: {sample_examples}"
            
            # Create check list with INDIVIDUAL sample entries (matching extraction table)
            original_check_list_data = []
            
            # Create reverse mapping from transformed back to original
            reverse_sample_id_mapping = {transformed: original for original, transformed in zip(ordered_sample_ids, transformed_sample_ids)}
            
            # Map plate assignments from transformed IDs back to original IDs
            original_plate_assignments = {}
            for transformed_id, plate in sample_plate_assignments.items():
                original_id = reverse_sample_id_mapping.get(transformed_id, transformed_id)
                original_plate_assignments[original_id] = plate
            
            # Map pool assignments from transformed IDs back to original IDs
            original_pool_assignments = {}
            for transformed_id, pool in pool_assignments.items():
                original_id = reverse_sample_id_mapping.get(transformed_id, transformed_id)
                original_pool_assignments[original_id] = pool
            
            # Create individual entries for each sample
            for original_id in ordered_sample_ids:
                check_list_entry = {
                    'Sample_Id': original_id,
                    'Plate': original_plate_assignments.get(original_id, ''),
                    'Pool': original_pool_assignments.get(original_id, 'Not Pool'),
                    'Status': 'in process'
                }
                original_check_list_data.append(check_list_entry)
            
            original_check_list = pd.DataFrame(original_check_list_data)
            
            # Reorder columns for original check list
            orig_columns = original_check_list.columns.tolist()
            for col in ['Pool', 'Plate', 'Status']:
                if col in orig_columns:
                    orig_columns.remove(col)
            orig_columns.extend(['Pool', 'Plate', 'Status'])
            original_check_list = original_check_list[orig_columns]
            
            # Store original check list data for later combination
            self.swab_check_list = original_check_list
            
            # Use the existing file_path and next_index from above
            if os.path.exists(file_path):
                mode = 'a'
                if_sheet_exists = 'overlay'
            else:
                mode = 'w'
                if_sheet_exists = None
                
            with pd.ExcelWriter(file_path, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
                for table_index, table in enumerate(extraction_tables):
                    # Table already has H2O positions integrated
                    modified_table = table
                    
                    extraction_df = pd.DataFrame(modified_table)
                    # Count only actual samples (not H2O) for sample range calculation
                    table_sample_ids = []
                    for row in modified_table:
                        for cell in row:
                            if cell and cell != 'H2O':
                                table_sample_ids.append(cell)
                    sample_extraction_id = self.get_sample_extraction_id(table_sample_ids)
                    column_pools = [None] * num_columns
                    for col in range(num_columns):
                        for row in range(len(modified_table)):
                            sample_id = modified_table[row][col]
                            if sample_id and sample_id != 'H2O' and sample_id in pool_assignments:
                                column_pools[col] = pool_assignments[sample_id]
                                break
                    sheet_name = f'Plate_{next_index:03d}'
                    workbook = writer.book
                    worksheet = workbook.create_sheet(sheet_name) if sheet_name not in workbook.sheetnames else workbook[sheet_name]
                    
                    # Define formatting styles
                    sample_id_font = Font(name='Arial', size=14, bold=False)
                    pool_font = Font(name='Arial', size=14, bold=True)
                    col_row_font = Font(name='Arial', size=22, bold=True)
                    header_font = Font(name='Arial', size=14, bold=True)
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    fills = [
                        PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
                        PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),  # Sky Blue
                        PatternFill(start_color="808080", end_color="808080", fill_type="solid")   # Gray
                    ]
                    
                    # Write three tables vertically stacked
                    for variant in range(3):
                        row_offset = variant * 18  # 18 rows between each table (including spacing)
                        
                        # Headers and metadata with font size 14
                        worksheet[f'I{3 + row_offset}'] = "Date:"
                        worksheet[f'I{3 + row_offset}'].font = header_font
                        worksheet[f'D{3 + row_offset}'] = "Inactivated by:"
                        worksheet[f'D{3 + row_offset}'].font = header_font
                        worksheet[f'C{1 + row_offset}'] = "Protocol: "
                        worksheet[f'C{1 + row_offset}'].font = header_font
                        worksheet[f'F{1 + row_offset}'] = "Kitt#: "
                        worksheet[f'F{1 + row_offset}'].font = header_font
                        worksheet[f'I{1 + row_offset}'] = "Lot#:"
                        worksheet[f'I{1 + row_offset}'].font = header_font
                        worksheet[f'L{1 + row_offset}'] = "Expiry:"
                        worksheet[f'L{1 + row_offset}'].font = header_font
                        worksheet[f'C{6 + row_offset}'] = "Sample Id"
                        worksheet[f'C{6 + row_offset}'].font = header_font
                        worksheet[f'E{6 + row_offset}'] = sample_extraction_id
                        worksheet[f'E{6 + row_offset}'].font = header_font
                        worksheet[f'I{5 + row_offset}'] = "Date:"
                        worksheet[f'I{5 + row_offset}'].font = header_font
                        worksheet[f'D{5 + row_offset}'] = "Extracted by:"
                        worksheet[f'D{5 + row_offset}'].font = header_font
                        
                        # Column headers
                        for col in range(num_columns):
                            cell = worksheet.cell(row=7 + row_offset, column=col + 2)
                            cell.value = str(col + 1)
                            cell.font = col_row_font
                            if fills[variant]:
                                cell.fill = fills[variant]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Row headers
                        for row in range(max_rows_per_column):
                            cell = worksheet.cell(row=row + 8 + row_offset, column=1)
                            cell.value = chr(65 + row)
                            cell.font = col_row_font
                            if fills[variant]:
                                cell.fill = fills[variant]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Sample IDs
                        for row in range(len(extraction_df)):
                            for col in range(len(extraction_df.columns)):
                                sample_id = extraction_df.iloc[row, col]
                                if sample_id:
                                    cell = worksheet.cell(row=row + 8 + row_offset, column=col + 2)
                                    cell.value = self.split_sample_id(sample_id) if sample_id != 'H2O' else 'H2O'
                                    if sample_id == 'H2O':
                                        cell.font = Font(name='Arial', size=14, bold=True, color="FF0000")  # Red bold for H2O
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red background
                                    else:
                                        cell.font = sample_id_font
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable wrap_text for two lines
                                        self.format_cell_for_sample_id(cell, worksheet)
                                    cell.border = thin_border
                        
                        # Pool row (no fill applied)
                        for col in range(num_columns):
                            if column_pools[col]:
                                cell = worksheet.cell(row=16 + row_offset, column=col + 2)
                                cell.value = column_pools[col]
                                cell.font = pool_font
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Plate number
                        worksheet[f'N{8 + row_offset}'] = str(next_index)
                        worksheet.merge_cells(f'N{8 + row_offset}:N{15 + row_offset}')
                        merged_cell = worksheet[f'N{8 + row_offset}']
                        merged_cell.font = Font(name='Arial', size=22, bold=True)
                        if fills[variant]:
                            merged_cell.fill = fills[variant]
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        merged_cell.border = Border(left=Side(style='thin'))
                        
                        # Set dimensions
                        for row in range(8 + row_offset, 16 + row_offset):
                            worksheet.row_dimensions[row].height = 35
                        for col in range(2, num_columns + 2):
                            worksheet.column_dimensions[get_column_letter(col)].width = 10
                    
                    next_index += 1
                    self.pre_save_formatting(worksheet)
            
            processing_progress['message'] = f"Created {len(extraction_tables)} Swab sample tables (3 tables per sheet)"
            
        except Exception as e:
            processing_progress['error'] = f'Error processing ANA/SAL samples: {str(e)}'

    def process_tis_int_samples(self, destination_file, current_step, total_steps):
        try:
            # Get existing Done samples to skip them
            done_samples = self.get_existing_done_samples(destination_file)
            processing_progress['message'] = f"Found {len(done_samples)} samples already completed, will skip them"
            
            # Try to read Excel file with better error handling
            try:
                df = pd.read_excel(destination_file, sheet_name='Extraction', engine='openpyxl')
            except Exception as e:
                # If openpyxl fails, try with different engine
                try:
                    df = pd.read_excel(destination_file, sheet_name='Extraction', engine='xlrd')
                except Exception as e2:
                    # If all engines fail, try basic read
                    df = pd.read_excel(destination_file, sheet_name='Extraction')
                    
            processing_progress['message'] = f"Processing TIS/INT samples..."
            
            # Custom TIS sample type classification
            if hasattr(self, 'tis_samples') and self.tis_samples:
                tis_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.tis_samples]
            else:
                tis_prefixes = []
            
            # Custom INT sample type classification
            if hasattr(self, 'int_samples') and self.int_samples:
                int_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.int_samples]
            else:
                int_prefixes = []
            tis_int_prefixes = tis_prefixes + int_prefixes
            if tis_int_prefixes:
                tis_int_pattern = '|'.join(tis_int_prefixes)
                # Filter out samples that are already Done
                available_samples = df[~df['Sample_Id'].isin(done_samples)]
                tis_int_samples = available_samples[available_samples['Sample_Id'].str.contains(tis_int_pattern, na=False) & (available_samples['Status'] == 'in process')]
            else:
                tis_int_samples = pd.DataFrame()  # Empty dataframe if no TIS/INT samples
                processing_progress['message'] = "No TIS/INT sample types configured - skipping TIS/INT processing"
                return  # Skip TIS/INT processing entirely
            
            # Filter TIS samples (original prefixes ending with TIS)
            if tis_prefixes:
                tis_sample_ids = tis_int_samples[tis_int_samples['Sample_Id'].str.contains('|'.join(tis_prefixes), na=False)]['Sample_Id'].tolist()
            else:
                tis_sample_ids = []
            tis_sample_ids.sort()
            processing_progress['message'] = f"TIS samples found: {len(tis_sample_ids)}"
            
            # Filter INT samples (original prefixes ending with INT)
            if int_prefixes:
                int_sample_ids = tis_int_samples[tis_int_samples['Sample_Id'].str.contains('|'.join(int_prefixes), na=False)]['Sample_Id'].tolist()
            else:
                int_sample_ids = []
            int_sample_ids.sort()
            processing_progress['message'] = f"INT samples found: {len(int_sample_ids)}"
            
            all_sample_ids = tis_sample_ids + int_sample_ids
            # Remove duplicates from all_sample_ids to prevent multiple assignments
            all_sample_ids = list(dict.fromkeys(all_sample_ids))
            transformed_sample_ids = [self.transform_sample_id(sid) for sid in all_sample_ids]
            
            # Create mapping from original to transformed sample IDs
            sample_id_mapping = {original: transformed for original, transformed in zip(all_sample_ids, transformed_sample_ids)}
            
            # Update the main dataframe
            for original, transformed in zip(all_sample_ids, transformed_sample_ids):
                df.loc[df['Sample_Id'] == original, 'Sample_Id'] = transformed
                
            os.makedirs('Extraction', exist_ok=True)
            
            # Create a copy for the check list
            tis_int_samples_copy = tis_int_samples.copy()
            
            # Update the Sample_Id column in the check list dataframe with transformed IDs
            tis_int_samples_copy['Sample_Id'] = tis_int_samples_copy['Sample_Id'].map(lambda x: sample_id_mapping.get(x, x))
            
            def get_sort_key(sample_id):
                if 'CANB_' in sample_id:
                    return (0, sample_id)
                else:
                    return (1, sample_id)
            
            tis_int_samples_copy['sort_key'] = tis_int_samples_copy['Sample_Id'].apply(get_sort_key)
            tis_int_samples_copy = tis_int_samples_copy.sort_values('sort_key')
            tis_int_samples_copy = tis_int_samples_copy.drop('sort_key', axis=1)
            
            # Track plate assignments for each sample
            tis_int_plate_assignments = {}
            
            extraction_tables = []
            num_columns = 11  # TIS/INT uses 11 columns in Extraction tables
            max_rows_per_column = 8
            items_per_plate = num_columns * max_rows_per_column
            
            # Get the starting plate number from existing Extraction_Tables.xlsx
            file_path = 'Extraction/Extraction_Tables.xlsx'
            if not os.path.exists(file_path):
                next_index = 1  # Start with Plate_001 when file doesn't exist
            else:
                try:
                    existing_sheets = pd.ExcelFile(file_path).sheet_names
                    # Extract existing plate numbers and find the highest
                    existing_numbers = []
                    for sheet in existing_sheets:
                        if sheet.startswith('Plate_'):
                            try:
                                num = int(sheet.replace('Plate_', ''))
                                existing_numbers.append(num)
                            except ValueError:
                                continue
                    next_index = max(existing_numbers) + 1 if existing_numbers else 1
                except Exception as e:
                    print(f"DEBUG: Error reading existing sheets: {e}")
                    next_index = 1
            
            # Create TIS/INT tables with H2O integration
            print(f"DEBUG: Creating {len(transformed_sample_ids)} TIS/INT extraction samples with H2O enabled={self.enable_h2o_random}")
            
            # Calculate how many samples can fit per plate accounting for H2O (11 columns for TIS/INT)
            total_positions = num_columns * max_rows_per_column  # 88 positions (11×8)
            samples_per_plate = total_positions - self.h2o_count if self.enable_h2o_random else total_positions
            print(f"DEBUG: TIS/INT - Total positions per plate: {total_positions}, H2O count: {self.h2o_count}, Samples per plate: {samples_per_plate}")
            
            plate_index = 0
            for plate_start in range(0, len(transformed_sample_ids), samples_per_plate):
                plate_samples = transformed_sample_ids[plate_start:plate_start + samples_per_plate]
                plate_key = f"Extraction_Plate_{next_index + plate_index}_TISINT"
                print(f"DEBUG: Creating TIS/INT Extraction table {plate_key} with {len(plate_samples)} samples")
                
                # Create table with H2O positions properly integrated (11 columns for TIS/INT)
                current_table = self.create_table_with_h2o(plate_samples, max_rows_per_column, num_columns, plate_key)
                print(f"DEBUG: Created TIS/INT table with {sum(1 for row in current_table for cell in row if cell == 'H2O')} H2O positions")
                
                # Assign plate numbers to samples (skip H2O positions)
                for row_idx, row in enumerate(current_table):
                    for col_idx, cell in enumerate(row):
                        if cell and cell != 'H2O':
                            tis_int_plate_assignments[cell] = f'Plate_{next_index + plate_index:03d}'
                
                extraction_tables.append(current_table)
                plate_index += 1
            
            # Add plate information to check list
            tis_int_samples_copy['Plate'] = tis_int_samples_copy['Sample_Id'].map(tis_int_plate_assignments)
            
            # Reorder columns to include Plate
            columns = tis_int_samples_copy.columns.tolist()
            columns.remove('Status')
            columns.remove('Plate')
            columns.extend(['Plate', 'Status'])
            tis_int_samples_copy = tis_int_samples_copy[columns]
            
            # Add logging to indicate what's being saved
            processing_progress['message'] = f"Writing {len(tis_int_samples_copy)} samples to Check_list_Tissue-Intestine.xlsx"
            
            # Debug output first few rows of data to verify sample IDs
            if not tis_int_samples_copy.empty:
                sample_examples = tis_int_samples_copy['Sample_Id'].head(5).tolist()
                processing_progress['message'] = f"Sample ID examples: {sample_examples}"
                
            # Create check list with ORIGINAL sample IDs (not transformed)
            original_tis_int_check_list = tis_int_samples.copy()
            # Create reverse mapping from transformed back to original
            reverse_sample_id_mapping = {transformed: original for original, transformed in zip(all_sample_ids, transformed_sample_ids)}
            
            # Map plate assignments from transformed IDs back to original IDs
            original_plate_assignments = {}
            for transformed_id, plate in tis_int_plate_assignments.items():
                original_id = reverse_sample_id_mapping.get(transformed_id, transformed_id)
                original_plate_assignments[original_id] = plate
            
            # Assign plate to original check list using original sample IDs
            original_tis_int_check_list['Plate'] = original_tis_int_check_list['Sample_Id'].map(original_plate_assignments)
            original_tis_int_check_list['Status'] = 'in process'
            
            # Reorder columns for original check list
            orig_columns = original_tis_int_check_list.columns.tolist()
            for col in ['Plate', 'Status']:
                if col in orig_columns:
                    orig_columns.remove(col)
            orig_columns.extend(['Plate', 'Status'])
            original_tis_int_check_list = original_tis_int_check_list[orig_columns]
            
            # Store original check list data for later combination
            self.tis_int_check_list = original_tis_int_check_list
            
            # Use the existing file_path and next_index from above
            if os.path.exists(file_path):
                mode = 'a'
                if_sheet_exists = 'overlay'
            else:
                mode = 'w'
                if_sheet_exists = None
                
            with pd.ExcelWriter(file_path, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
                for table_index, table in enumerate(extraction_tables):
                    extraction_df = pd.DataFrame(table)
                    table_sample_ids = [sid for row in table for sid in row if sid is not None]
                    sample_extraction_id = self.get_sample_extraction_id(table_sample_ids)
                    sheet_name = f'Plate_{next_index:03d}'
                    workbook = writer.book
                    worksheet = workbook.create_sheet(sheet_name) if sheet_name not in workbook.sheetnames else workbook[sheet_name]
                    
                    # Define formatting styles
                    sample_id_font = Font(name='Arial', size=14, bold=False)
                    col_row_font = Font(name='Arial', size=22, bold=True)
                    header_font = Font(name='Arial', size=14, bold=True)
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    fills = [
                        PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
                        PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),  # Sky Blue
                        PatternFill(start_color="808080", end_color="808080", fill_type="solid")   # Gray
                    ]
                    
                    # Write three tables vertically stacked
                    for variant in range(3):
                        row_offset = variant * 18  # 18 rows between each table (including spacing)
                        
                        # Headers and metadata with font size 14
                        worksheet[f'I{3 + row_offset}'] = "Date:"
                        worksheet[f'I{3 + row_offset}'].font = header_font
                        worksheet[f'D{3 + row_offset}'] = "Inactivated by:"
                        worksheet[f'D{3 + row_offset}'].font = header_font
                        worksheet[f'C{1 + row_offset}'] = "Protocol: "
                        worksheet[f'C{1 + row_offset}'].font = header_font
                        worksheet[f'F{1 + row_offset}'] = "Kitt#: "
                        worksheet[f'F{1 + row_offset}'].font = header_font
                        worksheet[f'I{1 + row_offset}'] = "Lot#:"
                        worksheet[f'I{1 + row_offset}'].font = header_font
                        worksheet[f'L{1 + row_offset}'] = "Expiry:"
                        worksheet[f'L{1 + row_offset}'].font = header_font
                        worksheet[f'C{6 + row_offset}'] = "Sample Id"
                        worksheet[f'C{6 + row_offset}'].font = header_font
                        worksheet[f'E{6 + row_offset}'] = sample_extraction_id
                        worksheet[f'E{6 + row_offset}'].font = header_font
                        worksheet[f'I{5 + row_offset}'] = "Date:"
                        worksheet[f'I{5 + row_offset}'].font = header_font
                        worksheet[f'D{5 + row_offset}'] = "Extracted by:"
                        worksheet[f'D{5 + row_offset}'].font = header_font
                        
                        # Column headers
                        for col in range(num_columns):
                            cell = worksheet.cell(row=7 + row_offset, column=col + 2)
                            cell.value = str(col + 1)
                            cell.font = col_row_font
                            if fills[variant]:
                                cell.fill = fills[variant]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Row headers
                        for row in range(max_rows_per_column):
                            cell = worksheet.cell(row=row + 8 + row_offset, column=1)
                            cell.value = chr(65 + row)
                            cell.font = col_row_font
                            if fills[variant]:
                                cell.fill = fills[variant]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Sample IDs
                        for row in range(len(extraction_df)):
                            for col in range(len(extraction_df.columns)):
                                sample_id = extraction_df.iloc[row, col]
                                if sample_id:
                                    cell = worksheet.cell(row=row + 8 + row_offset, column=col + 2)
                                    cell.value = self.split_sample_id(sample_id) if sample_id != 'H2O' else 'H2O'
                                    if sample_id == 'H2O':
                                        cell.font = Font(name='Arial', size=14, bold=True, color="FF0000")  # Red bold for H2O
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red background
                                    else:
                                        cell.font = sample_id_font
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable wrap_text for two lines
                                        self.format_cell_for_sample_id(cell, worksheet)
                                    cell.border = thin_border
                        
                        # Plate number
                        worksheet[f'M{8 + row_offset}'] = str(next_index)
                        worksheet.merge_cells(f'M{8 + row_offset}:M{15 + row_offset}')
                        merged_cell = worksheet[f'M{8 + row_offset}']
                        merged_cell.font = Font(name='Arial', size=22, bold=True)
                        if fills[variant]:
                            merged_cell.fill = fills[variant]
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        merged_cell.border = Border(left=Side(style='thin'))
                        
                        # Set dimensions
                        for row in range(8 + row_offset, 16 + row_offset):
                            worksheet.row_dimensions[row].height = 35
                        for col in range(2, num_columns + 2):
                            worksheet.column_dimensions[get_column_letter(col)].width = 10
                    
                    next_index += 1
                    self.pre_save_formatting(worksheet)
            
            processing_progress['message'] = f"Created {len(extraction_tables)} Tissue sample tables (3 tables per sheet)"
            
        except Exception as e:
            processing_progress['error'] = f'Error processing TIS/INT samples: {str(e)}'

    def create_combined_check_list(self, swab_check_list, tis_int_check_list):
        """Create a single combined check list with all samples"""
        try:
            if swab_check_list is None and tis_int_check_list is None:
                processing_progress['message'] = 'No samples to create check list'
                return
            
            # Combine both check lists
            combined_check_list = pd.DataFrame()
            
            if swab_check_list is not None and not swab_check_list.empty:
                # Ensure Pool column exists for swab samples
                if 'Pool' not in swab_check_list.columns:
                    swab_check_list['Pool'] = 'Not Pool'
                combined_check_list = pd.concat([combined_check_list, swab_check_list], ignore_index=True)
            
            if tis_int_check_list is not None and not tis_int_check_list.empty:
                # Add Pool column for tissue samples (they don't have pools)
                if 'Pool' not in tis_int_check_list.columns:
                    tis_int_check_list['Pool'] = 'Not Pool'
                combined_check_list = pd.concat([combined_check_list, tis_int_check_list], ignore_index=True)
            
            if combined_check_list.empty:
                processing_progress['message'] = 'No valid samples for check list'
                return
            
            # Sort by sample type and ID for better organization
            def get_sort_priority(sample_id):
                if isinstance(sample_id, str):
                    if 'CANB_' in sample_id:
                        return (0, sample_id)
                    elif 'CANR_' in sample_id:
                        return (1, sample_id)
                    elif 'IPLNAHL_' in sample_id:
                        return (2, sample_id)
                    elif 'CANA_' in sample_id:
                        return (3, sample_id)
                    else:
                        return (4, sample_id)
                return (5, str(sample_id))
            
            combined_check_list['sort_priority'] = combined_check_list['Sample_Id'].apply(get_sort_priority)
            combined_check_list = combined_check_list.sort_values('sort_priority')
            combined_check_list = combined_check_list.drop('sort_priority', axis=1)
            
            # Ensure consistent column order
            columns = combined_check_list.columns.tolist()
            print(f"DEBUG: Combined check list columns: {columns}")
            
            # Check if columns exist before trying to remove them
            for col in ['Pool', 'Plate', 'Status']:
                if col in columns:
                    columns.remove(col)
                    print(f"DEBUG: Removed column {col} from combined check list")
                else:
                    print(f"DEBUG: Column {col} not found in combined check list")
            
            columns.extend(['Pool', 'Plate', 'Status'])
            combined_check_list = combined_check_list[columns]
            
            # Save the combined check list
            combined_check_list.to_excel('Extraction/Check_list_All_Samples.xlsx', sheet_name='Check_List', index=False)
            processing_progress['message'] = f"Created combined check list with {len(combined_check_list)} samples"
            
        except Exception as e:
            processing_progress['error'] = f'Error creating combined check list: {str(e)}'

    def create_pcr_cdna_plates(self, destination_file, new_sample_ids):
        """Create PCR and cDNA plates based on pools and individual samples"""
        try:
            # Read pool information from combined check list
            pool_file = 'Extraction/Check_list_All_Samples.xlsx'
            if os.path.exists(pool_file):
                pool_df = pd.read_excel(pool_file, sheet_name='Check_List')
                pools = pool_df['Pool'].dropna().unique().tolist()
                # Filter out 'Not Pool' entries
                pools = [pool for pool in pools if pool != 'Not Pool']
                pools.sort(key=lambda x: int(x.split('_')[1]))
                
                # Create pool to samples mapping from check list
                pool_to_samples = {}
                for _, row in pool_df.iterrows():
                    pool_name = row['Pool']
                    sample_id = row['Sample_Id']
                    if pool_name != 'Not Pool' and pd.notna(pool_name):
                        if pool_name not in pool_to_samples:
                            pool_to_samples[pool_name] = []
                        # Transform back to original sample ID for PCR/cDNA plates
                        original_sample_id = self.reverse_transform_sample_id(sample_id)
                        pool_to_samples[pool_name].append(original_sample_id)
                
                processing_progress['message'] = f'Found {len(pools)} unique pools with sample mapping'
            else:
                pools = []
                pool_to_samples = {}
                processing_progress['message'] = f'Warning: Combined check list file not found'
            
            # Read extraction data
            df = pd.read_excel(destination_file, sheet_name='Extraction')
            new_samples_df = df[df['Sample_Id'].isin(new_sample_ids)]
            
            # Get Tissue/Intestine samples
            tis_int_prefixes = [new_prefix for old_prefix, new_prefix in self.prefix_mapping.items() 
                              if 'TIS' in old_prefix or 'INT' in old_prefix]
            tis_int_pattern = '|'.join(tis_int_prefixes)
            print(f"DEBUG: TIS/INT prefixes found: {tis_int_prefixes}")
            print(f"DEBUG: TIS/INT pattern: {tis_int_pattern}")
            print(f"DEBUG: New samples dataframe shape: {new_samples_df.shape}")
            print(f"DEBUG: New samples Sample_Ids: {new_samples_df['Sample_Id'].tolist()[:10]}")
            
            tis_int_df = new_samples_df[new_samples_df['Sample_Id'].str.contains(tis_int_pattern, na=False)]
            tis_int_samples = tis_int_df['Sample_Id'].tolist()
            print(f"DEBUG: Found TIS/INT samples: {tis_int_samples}")
            
            # Sort samples
            sorted_tis_int_samples = []
            for prefix in tis_int_prefixes:
                prefix_samples = [s for s in tis_int_samples if s.startswith(prefix)]
                prefix_samples.sort(key=lambda x: int(x.split('_')[1]))
                sorted_tis_int_samples.extend(prefix_samples)
            tis_int_samples = sorted_tis_int_samples
            
            # Create PCR and cDNA plates
            self.create_continuous_plates(pools, pool_to_samples, tis_int_samples, 'PCR', 'PCR_Tables')
            self.create_continuous_plates(pools, pool_to_samples, tis_int_samples, 'cDNA', 'cDNA_Tables')
            
        except Exception as e:
            processing_progress['error'] = f'Error creating PCR/cDNA plates: {str(e)}'

    def create_continuous_plates(self, pools, pool_to_samples, tis_int_samples, plate_type, base_filename):
        """Create continuous plates with pools first, then Tissue/Intestine samples"""
        try:
            os.makedirs(f'Extraction/{plate_type}', exist_ok=True)
            
            max_rows = 8
            num_columns = 11
            items_per_plate = max_rows * num_columns
            
            file_path = f'Extraction/{plate_type}/{base_filename}.xlsx'
            next_index = 1 if not os.path.exists(file_path) else len(pd.ExcelFile(file_path).sheet_names) + 1
            
            if os.path.exists(file_path):
                mode = 'a'
                if_sheet_exists = 'overlay'
            else:
                mode = 'w'
                if_sheet_exists = None
                
            # Only create if we have data to process
            if not pools and not tis_int_samples:
                processing_progress['message'] = f'No data available for {plate_type} plates'
                return
            
            with pd.ExcelWriter(file_path, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
                # Process pools
                pool_items = pools.copy()
                if pool_items:
                    processing_progress['message'] = f'Creating {plate_type} plates for {len(pool_items)} Swab pools'
                    
                    for plate_start in range(0, len(pool_items), items_per_plate):
                        plate_samples = pool_items[plate_start:plate_start+items_per_plate]
                        
                        # Use the same plate key as Extraction for consistency
                        plate_key = f"Extraction_Plate_{next_index}"
                        
                        # Create table with H2O positions properly integrated (11 columns for PCR/cDNA)
                        modified_table = self.create_table_with_h2o(plate_samples, max_rows, num_columns, plate_key, format_for_pcr_cdna=True)
                        
                        # Add empty column M
                        for row in modified_table:
                            row.append('')
                        
                        # Create DataFrame
                        df_plate = pd.DataFrame(modified_table, 
                                            columns=[i+1 for i in range(12)], 
                                            index=[chr(65+i) for i in range(max_rows)])
                        
                        sheet_name = f'Plate_{next_index:03d}'
                        
                        # Get sample range for headers
                        plate_sample_ids = [item for item in plate_samples if item != '']
                        if any(isinstance(sid, str) and sid.startswith('Pool_') for sid in plate_sample_ids):
                            original_pool_samples = []
                            for item in plate_sample_ids:
                                if isinstance(item, str) and item.startswith('Pool_'):
                                    if item in pool_to_samples:
                                        original_pool_samples.extend(pool_to_samples[item])
                            sample_range = self.get_sample_range(original_pool_samples, use_original=False, force_original=True)
                        else:
                            sample_range = self.get_sample_range(plate_sample_ids)
                        
                        if plate_type == 'PCR':
                            # First PCR table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            
                            # Second PCR table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=18)
                            
                            worksheet = writer.sheets[sheet_name]
                            
                            # Don't clear M5 and M19 - keep the "12" value but will be reformatted
                            
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=0, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='PCR')
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=14, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='PCR')
                        else:
                            # Single cDNA table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            worksheet = writer.sheets[sheet_name]
                            
                            # Don't clear M5 - keep the "12" value but will be reformatted
                            
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=0, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='cDNA')
                        
                        next_index += 1
                
                # Process Tissue/Intestine samples
                if tis_int_samples:
                    processing_progress['message'] = f'Creating {plate_type} plates for {len(tis_int_samples)} Tissue/Intestine samples'
                    
                    # Start Tissue/Intestine plates after pool plates to avoid key conflicts
                    tis_int_start_index = next_index
                    
                    for plate_start in range(0, len(tis_int_samples), items_per_plate):
                        plate_samples = tis_int_samples[plate_start:plate_start+items_per_plate]
                        
                        # Use unique plate key for Tissue/Intestine plates
                        tis_int_plate_num = tis_int_start_index + (plate_start // items_per_plate)
                        plate_key = f"Extraction_Plate_{tis_int_plate_num}_TISINT"
                        
                        # Create table with H2O positions properly integrated (11 columns for PCR/cDNA)
                        modified_table = self.create_table_with_h2o(plate_samples, max_rows, num_columns, plate_key, format_for_pcr_cdna=True)
                        
                        # Add empty column M
                        for row in modified_table:
                            row.append('')
                        
                        # Create DataFrame
                        df_plate = pd.DataFrame(modified_table, 
                                            columns=[i+1 for i in range(12)], 
                                            index=[chr(65+i) for i in range(max_rows)])
                        
                        sheet_name = f'Plate_{tis_int_plate_num}'
                        
                        # Get sample range
                        plate_sample_ids = [item for item in plate_samples if item != '']
                        sample_range = self.get_sample_range(plate_sample_ids)
                        
                        if plate_type == 'PCR':
                            # First PCR table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            
                            # Second PCR table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=18)
                            
                            worksheet = writer.sheets[sheet_name]
                            
                            # Don't clear M5 and M19 - keep the "12" value but will be reformatted
                            
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=0, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='PCR')
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=14, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='PCR')
                        else:
                            # Single cDNA table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            worksheet = writer.sheets[sheet_name]
                            
                            # Don't clear M5 - keep the "12" value but will be reformatted
                            
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=0, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='cDNA')
                        
                        next_index += 1
            
            processing_progress['message'] = f'Created continuous {plate_type} plates'
            
        except Exception as e:
            processing_progress['error'] = f'Error creating continuous {plate_type} plates: {str(e)}'
            
    def create_dynamic_pools(self, samples, h2o_positions, max_rows=8, num_columns=12):
        """Create pools dynamically based on H2O positions in each column"""
        if not self.samples_per_pool:
            return []
        
        pools = []
        pool_count = 1
        
        # Group samples by column capacity
        column_capacities = []
        for col in range(num_columns):
            # Check if this column has H2O
            has_h2o = any(pos[1] == col for pos in h2o_positions)
            capacity = max_rows - 1 if has_h2o else max_rows  # 7 if H2O, 8 if no H2O
            column_capacities.append(capacity)
        
        print(f"DEBUG: Column capacities: {column_capacities}")
        
        # Distribute samples across columns based on capacity
        remaining_samples = samples.copy()
        current_pool_samples = []
        
        for col, capacity in enumerate(column_capacities):
            # Take samples for this column based on its capacity
            samples_for_column = remaining_samples[:capacity]
            remaining_samples = remaining_samples[capacity:]
            
            # Add these samples to current pool
            current_pool_samples.extend(samples_for_column)
            
            # Check if we have enough for a complete pool
            if len(current_pool_samples) >= self.samples_per_pool:
                # Create pools from accumulated samples
                for i in range(0, len(current_pool_samples), self.samples_per_pool):
                    pool_samples = current_pool_samples[i:i + self.samples_per_pool]
                    if len(pool_samples) == self.samples_per_pool:
                        pool_name = f"Pool_{pool_count:03}"
                        pools.append({
                            'name': pool_name,
                            'samples': pool_samples,
                            'column': col
                        })
                        pool_count += 1
                
                # Keep any remaining samples for next columns
                current_pool_samples = current_pool_samples[len(current_pool_samples) % self.samples_per_pool:]
        
        print(f"DEBUG: Created {len(pools)} dynamic pools")
        return pools

    def get_sample_range(self, sample_ids, use_original=False, force_original=False):
        """Get sample range string grouped by prefix type"""
        print(f"DEBUG: get_sample_range called with {len(sample_ids)} samples, use_original={use_original}")
        print(f"DEBUG: Sample IDs received: {sample_ids[:10] if len(sample_ids) > 10 else sample_ids}...")  # Show first 10
        
        if not sample_ids:
            return "No samples"
        
        # Check if these are pool samples (start with "Pool_")
        is_pool_samples = any(isinstance(sid, str) and sid.startswith('Pool_') for sid in sample_ids)
        print(f"DEBUG: is_pool_samples = {is_pool_samples}")
        
        if is_pool_samples:
            # Extract pool numbers and map to sample IDs using prefix mapping
            pool_numbers = []
            for sid in sample_ids:
                if isinstance(sid, str) and sid.startswith('Pool_'):
                    try:
                        pool_num = sid.split('_')[1]
                        pool_numbers.append(int(pool_num))
                    except (IndexError, ValueError):
                        continue
            
            if not pool_numbers:
                return "No samples"
            
            pool_numbers.sort()
            
            # Find the prefix to use
            if force_original or use_original:
                # Use original prefix for check lists
                pool_prefix = None
                for old_prefix, new_prefix in self.prefix_mapping.items():
                    if any(old_prefix in ps for ps in self.pool_samples):
                        pool_prefix = old_prefix
                        break
            else:
                # Use new prefix for plates
                pool_prefix = None
                for old_prefix, new_prefix in self.prefix_mapping.items():
                    if any(old_prefix in ps for ps in self.pool_samples):
                        pool_prefix = new_prefix
                        break
            
            if not pool_prefix:
                # Use the first configured pool sample prefix from settings
                if self.pool_samples:
                    first_pool_prefix = self.pool_samples[0]
                    if force_original or use_original:
                        pool_prefix = first_pool_prefix
                    else:
                        pool_prefix = self.prefix_mapping.get(first_pool_prefix, first_pool_prefix)
                else:
                    pool_prefix = "UNKNOWN"  # Last resort if no pool samples configured
            
            # Create sample range with proper formatting
            if len(pool_numbers) == 1:
                return f"Sample Id: {pool_prefix}_{pool_numbers[0]:03d}"
            else:
                return f"Sample Id: {pool_prefix}_{pool_numbers[0]:03d} to {pool_prefix}_{pool_numbers[-1]:03d}"
        else:
            # Regular sample IDs (non-pool) - group by prefix
            clean_ids = []
            for sid in sample_ids:
                if isinstance(sid, str):
                    clean_id = sid.replace('\n', '_')
                    clean_ids.append(clean_id)
            
            if not clean_ids:
                return "No samples"
            
            # Group by prefix
            prefix_groups = {}
            for sid in clean_ids:
                # Extract prefix (everything before the last underscore and number)
                if '_' in sid:
                    parts = sid.split('_')
                    if len(parts) >= 3:
                        prefix = '_'.join(parts[:-1])  # Everything except the number
                        number = int(parts[-1])
                        if prefix not in prefix_groups:
                            prefix_groups[prefix] = []
                        prefix_groups[prefix].append(number)
            
            if not prefix_groups:
                # Fallback to original behavior if no prefix grouping
                clean_ids.sort()
                if len(clean_ids) == 1:
                    return f"Sample Id: {clean_ids[0]}"
                else:
                    return f"Sample Id: {clean_ids[0]} to {clean_ids[-1]}"
            
            # Create ranges for each prefix group
            ranges = []
            for prefix in sorted(prefix_groups.keys()):
                numbers = sorted(prefix_groups[prefix])
                if len(numbers) == 1:
                    ranges.append(f"{prefix}_{numbers[0]:03d}")
                else:
                    ranges.append(f"{prefix}_{numbers[0]:03d} to {numbers[-1]:03d}")
            
            # Format as type groups with concise numbering
            if len(ranges) == 1:
                return f"Sample Id: {ranges[0]}"
            elif len(ranges) == 2:
                # Extract prefixes and number ranges
                range1_parts = ranges[0].split('_')
                range2_parts = ranges[1].split('_')
                
                # Reconstruct with concise format
                if len(range1_parts) >= 3 and len(range2_parts) >= 3:
                    prefix1 = '_'.join(range1_parts[:-1])
                    prefix2 = '_'.join(range2_parts[:-1])
                    return f"Sample Id: {prefix1}_{range1_parts[-1]} and {prefix2}_{range2_parts[-1]}"
            
            return f"Sample Id: {', '.join(ranges)}"

    def format_plate_worksheet(self, worksheet, max_rows, num_columns, start_row=0, sample_range=None, plate_number=None, table_type='PCR'):
        """Format a plate worksheet with borders and fonts"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_font = Font(name='Arial', size=12, bold=False)
        cell_font = Font(name='Arial', size=14, bold=False)
        bold_font = Font(name='Arial', size=16, bold=True)
        row_col_font = Font(name='Arial', size=22, bold=True)
        
        # Determine header positions based on table type and start_row
        if table_type == 'PCR':
            if start_row == 0:
                # First PCR table
                header_rows = {
                    'perform_by': 2,
                    'date': 2,
                    'extraction_plate': 3,
                    'plate_number': 3,
                    'cdna_date': 3,
                    'sample_id': 4,
                    'data_start': 5
                }
            else:
                # Second PCR table
                header_rows = {
                    'perform_by': 16,
                    'date': 16,
                    'extraction_plate': 17,
                    'plate_number': 17,
                    'cdna_date': 17,
                    'sample_id': 18,
                    'data_start': 19
                }
        else:  # cDNA
            header_rows = {
                'perform_by': 2,
                'date': 2,
                'extraction_plate': 3,
                'plate_number': 3,
                'cdna_date': 3,
                'sample_id': 4,
                'data_start': 5
            }
        
        # Add header labels
        # Row with "Perform by:" and "Date:"
        perform_by_cell = worksheet.cell(row=header_rows['perform_by'], column=2)  # B
        perform_by_cell.value = "Perform by:"
        perform_by_cell.font = header_font
        perform_by_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        date_cell = worksheet.cell(row=header_rows['date'], column=9)  # I
        if table_type == 'PCR':
            date_cell.value = "PCR Date:"
        else:  # cDNA
            date_cell.value = "cDNA Date:"
        date_cell.font = header_font
        date_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Row with "Extraction Plate:", plate number, and date label
        extraction_label_cell = worksheet.cell(row=header_rows['extraction_plate'], column=2)  # B
        extraction_label_cell.value = "Extraction Plate:"
        extraction_label_cell.font = header_font
        extraction_label_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        plate_num_cell = worksheet.cell(row=header_rows['plate_number'], column=4)  # D
        plate_num_cell.value = plate_number if plate_number else ""
        plate_num_cell.font = header_font
        plate_num_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Only add cDNA Date for PCR tables (as requested to remove from cDNA)
        if table_type == 'PCR':
            cdna_date_cell = worksheet.cell(row=header_rows['cdna_date'], column=9)  # I
            cdna_date_cell.value = "cDNA Date:"
            cdna_date_cell.font = header_font
            cdna_date_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Row with Sample ID range
        if sample_range:
            sample_id_cell = worksheet.cell(row=header_rows['sample_id'], column=2)  # B
            sample_id_cell.value = sample_range
            sample_id_cell.font = bold_font
            sample_id_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Format the data table (column and row headers + data cells)
        data_start_row = header_rows['data_start']
        
        # Column headers (1-11) plus column M (12)
        for col in range(num_columns + 1):  # +1 to include column M
            cell = worksheet.cell(row=data_start_row, column=col + 2)  # Start at column B
            if col < num_columns:
                cell.value = str(col + 1)  # Numbers 1-11
            else:
                # Column M - keep existing value (12) or set it
                if cell.value != '12':
                    cell.value = '12'
            cell.font = row_col_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Row headers (A-H) and data cells
        for row in range(max_rows):
            # Row header
            row_header_cell = worksheet.cell(row=data_start_row + row + 1, column=1)  # Column A
            row_header_cell.value = chr(65 + row)  # A, B, C, etc.
            row_header_cell.font = row_col_font
            row_header_cell.alignment = Alignment(horizontal='center', vertical='center')
            row_header_cell.border = thin_border
            
            # Data cells (columns B-L) and empty column M
            for col in range(num_columns + 1):  # +1 for column M (empty)
                cell = worksheet.cell(row=data_start_row + row + 1, column=col + 2)
                if cell.value is None:
                    cell.value = ''
                cell.border = thin_border
                
                # Special formatting for column M (column 13) - last row of each table
                if col == num_columns and row == max_rows - 1:  # Column M, last row (H)
                    cell.font = row_col_font  # Size 22, bold
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    # Check if this is an H2O cell and apply special formatting
                    if cell.value == 'H2O':
                        cell.font = Font(name='Arial', size=14, bold=True, color="FF0000")  # Red bold for H2O
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red background
                    else:
                        cell.font = cell_font
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                
                # Set row height for wrapped text
                worksheet.row_dimensions[data_start_row + row + 1].height = 35
        
        # Set column widths
        for col in range(1, num_columns + 3):  # A through M
            worksheet.column_dimensions[get_column_letter(col)].width = 10

    def process_files(self, source_file, destination_file, sheets_info):
        global processing_progress
        try:
            processing_progress['status'] = 'processing'
            processing_progress['progress'] = 0
            processing_progress['message'] = 'Starting processing...'
            processing_progress['error'] = ''

            output_data = pd.DataFrame(columns=['Host_Id', 'Sample_Id', 'Status'])
            existing_columns = set()
            
            if os.path.exists(destination_file):
                try:
                    existing_data = pd.read_excel(destination_file, sheet_name="Extraction")
                    if 'Status' not in existing_data.columns:
                        existing_data['Status'] = 'unknown'
                    existing_sample_ids = existing_data['Sample_Id'].dropna().tolist()
                    existing_columns = set(self.transform_sample_id(sid) for sid in existing_sample_ids)
                    processing_progress['message'] = f'Loaded {len(existing_columns)} existing Sample_Ids'
                except Exception as e:
                    processing_progress['message'] = f'Warning: Could not load existing data: {str(e)}'

            total_steps = len(sheets_info) + 6
            current_step = 0

            # Process sheets
            for sheet_name, columns in sheets_info.items():
                current_step += 1
                processing_progress['progress'] = int((current_step / total_steps) * 100)
                processing_progress['message'] = f'Processing {sheet_name}...'
                
                try:
                    source_data = pd.read_excel(source_file, sheet_name=sheet_name)
                    if source_data.empty:
                        continue
                    selected_columns = source_data.iloc[:, columns]
                    
                    for index, row in selected_columns.iterrows():
                        host_value = row.iloc[0]
                        for i in range(1, len(row)):
                            col_value = row.iloc[i]
                            
                            # Filter by custom sample type settings only
                            if pd.notna(col_value) and col_value not in existing_columns:
                                # Check if this sample matches any of your custom sample type settings
                                sample_matches_custom_type = False
                                
                                # Check ANA samples
                                if hasattr(self, 'ana_samples') and self.ana_samples:
                                    for ana_prefix in self.ana_samples:
                                        if ana_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Check SAL samples
                                if not sample_matches_custom_type and hasattr(self, 'sal_samples') and self.sal_samples:
                                    for sal_prefix in self.sal_samples:
                                        if sal_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Check PT samples
                                if not sample_matches_custom_type and hasattr(self, 'pt_samples') and self.pt_samples:
                                    for pt_prefix in self.pt_samples:
                                        if pt_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Check TIS samples
                                if not sample_matches_custom_type and hasattr(self, 'tis_samples') and self.tis_samples:
                                    for tis_prefix in self.tis_samples:
                                        if tis_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Check INT samples
                                if not sample_matches_custom_type and hasattr(self, 'int_samples') and self.int_samples:
                                    for int_prefix in self.int_samples:
                                        if int_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Only include samples that match your custom sample type settings
                                if sample_matches_custom_type:
                                    new_row = pd.DataFrame({
                                        'Host_Id': [host_value], 
                                        'Sample_Id': [self.transform_sample_id(col_value)], 
                                        'Status': ['in process']
                                    })
                                    output_data = pd.concat([output_data, new_row], ignore_index=True)
                except Exception as e:
                    processing_progress['message'] = f'Error processing {sheet_name}: {str(e)}'

            # Save output data
            if not output_data.empty:
                try:
                    if not os.path.exists(destination_file):
                        with pd.ExcelWriter(destination_file, engine='openpyxl') as writer:
                            if 'Sample_Id' in output_data.columns:
                                output_data['Sample_Id'] = output_data['Sample_Id'].apply(self.transform_sample_id)
                            output_data.to_excel(writer, sheet_name="Extraction", index=False)
                    else:
                        with pd.ExcelWriter(destination_file, engine='openpyxl', mode='a', 
                                          if_sheet_exists='overlay') as writer:
                            existing_data = pd.read_excel(destination_file, sheet_name="Extraction")
                            if 'Status' not in existing_data.columns:
                                existing_data['Status'] = 'unknown'
                            combined_data = pd.concat([existing_data, output_data], ignore_index=True)
                            combined_data.drop_duplicates(subset=['Sample_Id'], keep='first', inplace=True)
                            if 'Sample_Id' in combined_data.columns:
                                combined_data['Sample_Id'] = combined_data['Sample_Id'].apply(self.transform_sample_id)
                            combined_data.to_excel(writer, sheet_name="Extraction", index=False)
                    processing_progress['message'] = f'Added {len(output_data)} new rows'
                except Exception as e:
                    processing_progress['error'] = f'Error saving data: {str(e)}'
                    return

            # Process ANA/SAL/PT samples
            current_step += 2
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            self.process_ana_sal_samples(destination_file, current_step, total_steps)

            # Process TIS/INT samples
            current_step += 2
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            self.process_tis_int_samples(destination_file, current_step, total_steps)

            # Create combined check list with all samples
            current_step += 1
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            self.create_combined_check_list(self.swab_check_list, self.tis_int_check_list)

            # Update status to "Done"
            current_step += 1
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            processing_progress['message'] = 'Updating sample status...'
            
            try:
                df = pd.read_excel(destination_file, sheet_name='Extraction')
                processed_samples = set()
                transformed_prefixes = list(self.prefix_mapping.values())
                prefix_pattern = '|'.join(transformed_prefixes)
                processed_samples.update(df[df['Sample_Id'].str.contains(prefix_pattern, na=False) & (df['Status'] == 'in process')]['Sample_Id'].tolist())
                df.loc[df['Sample_Id'].isin(processed_samples), 'Status'] = 'Done'
                with pd.ExcelWriter(destination_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name="Extraction", index=False)
                processing_progress['message'] = f'Updated status to "Done" for {len(processed_samples)} samples'
            except Exception as e:
                processing_progress['error'] = f'Error updating status: {str(e)}'

            # Create PCR and cDNA plates if new samples were added
            current_step += 1
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            if len(output_data) > 0:
                new_sample_ids = output_data['Sample_Id'].tolist()
                self.create_pcr_cdna_plates(destination_file, new_sample_ids)
            else:
                processing_progress['message'] = 'No new samples added - skipping PCR/cDNA plate creation'

            processing_progress['status'] = 'completed'
            processing_progress['progress'] = 100
            processing_progress['message'] = 'Processing completed successfully! All tables created.'

        except Exception as e:
            processing_progress['status'] = 'error'
            processing_progress['error'] = f'Processing failed: {str(e)}'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/settings')
def settings_page():
    return render_template('settings.html')

@app.route('/api/settings', methods=['GET'])
def get_settings():
    """Return current settings as JSON"""
    try:
        # Load from file if exists, otherwise return empty settings
        settings_file = 'sample_settings.json'
        if os.path.exists(settings_file):
            with open(settings_file, 'r') as f:
                return json.load(f)
        else:
            # Return empty settings for UI to start fresh
            return {
                'prefix_mapping': {},
                'pool_samples': [],
                'no_pool_samples': [],
                'samples_per_pool': None,
                'sample_id_pattern': None,
                'ana_samples': [],
                'sal_samples': [],
                'pt_samples': [],
                'tis_samples': [],
                'int_samples': [],
                'enable_h2o_random': False,
                'h2o_count': 2,
                'h2o_position_preference': 'random'
            }
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/settings', methods=['POST'])
def save_settings():
    """Save settings to file"""
    try:
        settings = request.get_json()
        settings_file = 'sample_settings.json'
        
        with open(settings_file, 'w') as f:
            json.dump(settings, f, indent=2)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'source_file' not in request.files:
        return jsonify({'error': 'No source file uploaded'}), 400
    
    source_file = request.files['source_file']
    if source_file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(source_file.filename):
        return jsonify({'error': 'Invalid file type. Please upload Excel files only'}), 400

    # Save source file
    source_filename = secure_filename(source_file.filename)
    source_path = os.path.join(app.config['UPLOAD_FOLDER'], source_filename)
    source_file.save(source_path)

    # Handle destination file - check for custom output or existing default
    custom_output = request.form.get('custom_output', '').strip()
    dest_path = None
    
    if custom_output:
        # Use custom output file if specified
        if not custom_output.endswith('.xlsx'):
            custom_output += '.xlsx'
        dest_path = os.path.join(app.config['OUTPUT_FOLDER'], custom_output)
        processing_progress['message'] = f"Using custom output file: {custom_output}"
    else:
        # Check for existing default output file (most recent)
        try:
            existing_files = [f for f in os.listdir(app.config['OUTPUT_FOLDER']) 
                           if f.startswith('Extraction_Output_') and f.endswith('.xlsx')]
            if existing_files:
                # Sort by modification time to get the most recent
                existing_files.sort(key=lambda x: os.path.getmtime(os.path.join(app.config['OUTPUT_FOLDER'], x)), reverse=True)
                dest_path = os.path.join(app.config['OUTPUT_FOLDER'], existing_files[0])
                processing_progress['message'] = f"Using existing output file: {existing_files[0]}"
        except Exception as e:
            print(f"DEBUG: Error checking existing output files: {e}")
    
    # If no existing file found, create new one
    if dest_path is None or not os.path.exists(dest_path):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        dest_filename = f"Extraction_Output_{timestamp}.xlsx"
        dest_path = os.path.join(app.config['OUTPUT_FOLDER'], dest_filename)
        processing_progress['message'] = f"Creating new output file: {dest_filename}"
        
        # Initialize new destination file
        try:
            with pd.ExcelWriter(dest_path, engine='openpyxl') as writer:
                pd.DataFrame(columns=['Host_Id', 'Sample_Id', 'Status']).to_excel(
                    writer, sheet_name='Extraction', index=False
                )
        except Exception as e:
            return jsonify({'error': f'Error initializing destination file: {str(e)}'}), 500
    else:
        # Verify existing file has required structure
        try:
            df = pd.read_excel(dest_path, sheet_name='Extraction')
            if 'Sample_Id' not in df.columns or 'Status' not in df.columns:
                # Add missing columns if needed
                if 'Sample_Id' not in df.columns:
                    df['Sample_Id'] = ''
                if 'Status' not in df.columns:
                    df['Status'] = 'unknown'
                
                # Save back with corrected structure
                with pd.ExcelWriter(dest_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name='Extraction', index=False)
                processing_progress['message'] = f"Updated existing output file structure"
        except Exception as e:
            processing_progress['message'] = f"Warning: Could not verify existing file structure: {str(e)}"

    # Read Excel sheets
    try:
        xl = pd.ExcelFile(source_path)
        sheets_data = {}
        for sheet in xl.sheet_names:
            df = pd.read_excel(source_path, sheet_name=sheet)
            sheets_data[sheet] = {
                'columns': df.columns.tolist(),
                'preview': df.head().to_html(classes='table table-striped table-sm', index=False)
            }
        
        return jsonify({
            'source_path': source_path,
            'dest_path': dest_path,
            'sheets': sheets_data
        })
    except Exception as e:
        return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 500

@app.route('/process', methods=['POST'])
def process_files():
    data = request.get_json()
    source_path = data.get('source_path')
    dest_path = data.get('dest_path')
    sheets_info = data.get('sheets_info', {})

    if not source_path or not dest_path or not sheets_info:
        return jsonify({'error': 'Missing required parameters'}), 400

    # Reset progress
    global processing_progress
    processing_progress = {'status': 'idle', 'progress': 0, 'message': '', 'error': ''}

    # Start processing in background thread
    processor = ExtractionProcessor()
    thread = threading.Thread(target=processor.process_files, args=(source_path, dest_path, sheets_info))
    thread.daemon = True
    thread.start()

    return jsonify({'message': 'Processing started'})

@app.route('/progress')
def get_progress():
    return jsonify(processing_progress)

@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'error': f'Error downloading file: {str(e)}'}), 500

@app.route('/outputs')
def list_outputs():
    try:
        files = []
        for filename in os.listdir(app.config['OUTPUT_FOLDER']):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
                stat = os.stat(file_path)
                files.append({
                    'name': filename,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
        files.sort(key=lambda x: x['modified'], reverse=True)
        return jsonify({'files': files})
    except Exception as e:
        return jsonify({'error': f'Error listing files: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
