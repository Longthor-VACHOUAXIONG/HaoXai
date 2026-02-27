"""
Extraction Table Generator Blueprint
Full implementation from Ext-table standalone app
"""
import os
import sys
import pandas as pd
import re
import json
import random
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, colors
from openpyxl.utils import get_column_letter
from flask import Blueprint, render_template, request, jsonify, send_file, flash, redirect, url_for, current_app
from werkzeug.utils import secure_filename
import threading
import time
from datetime import datetime

extraction_bp = Blueprint('extraction', __name__)

# Get folder paths relative to app directory
def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def format_sample_number(number_str):
    """Format sample number to 3 digits if 1-2 digits, leave as 4 digits if already 4+ digits"""
    try:
        # Extract only digits from the string
        digits = ''.join(c for c in str(number_str) if c.isdigit())
        if not digits:
            return number_str
        
        # Convert to integer to remove leading zeros, then format
        num = int(digits)
        
        # Format to 3 digits if less than 1000, otherwise keep as is
        if num < 1000:
            return f"{num:03d}"
        else:
            return str(num)
    except:
        return number_str

def natural_sort_key(s):
    """Helper for natural sorting (e.g., 10 comes after 2)"""
    import re
    s = str(s)
    # Category 0: Purely numeric (e.g., "21949")
    # Category 1: Mixed/Alpha (e.g., "19-10012", "BT1")
    category = 0 if s.isdigit() else 1
    # Extract numbers and non-numeric parts
    parts = [int(text) if text.isdigit() else text.lower()
            for text in re.split('([0-9]+)', s) if text]
    return (category, parts)

def get_upload_folder():
    folder = os.path.join(get_base_path(), 'uploads', 'extraction')
    os.makedirs(folder, exist_ok=True)
    return folder

def get_output_folder():
    folder = os.path.join(get_base_path(), 'outputs', 'extraction')
    os.makedirs(folder, exist_ok=True)
    return folder

def get_settings_file():
    return os.path.join(get_base_path(), 'sample_settings.json')

# Global variables for progress tracking
processing_progress = {'status': 'idle', 'progress': 0, 'message': '', 'error': ''}

class ExtractionProcessor:
    def __init__(self):
        print("="*50)
        print("EXTRACTION PROCESSOR INITIALIZED - LOADING SETTINGS")
        print("="*50)
        self.load_settings()
        self.extraction_plate_snapshots = []  # Record every plate created for extraction
        print("="*50)
        print("SETTINGS LOADED COMPLETELY")
        print("="*50)
        self.swab_check_list = None
        self.tis_int_check_list = None
    
    def load_settings(self):
        """Load settings from JSON file if exists, otherwise start with empty settings"""
        try:
            settings_file = get_settings_file()
            # If writable settings don't exist, try to load from internal/default location
            if not os.path.exists(settings_file):
                internal_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                internal_settings = os.path.join(internal_path, 'sample_settings.json')
                if os.path.exists(internal_settings):
                    print(f"DEBUG: Using default settings from {internal_settings}")
                    settings_file = internal_settings
            
            if os.path.exists(settings_file):
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
                    self.prefix_mapping = settings.get('prefix_mapping', {})
                    self.pool_samples = settings.get('pool_samples', [])
                    self.no_pool_samples = settings.get('no_pool_samples', [])
                    self.samples_per_pool = settings.get('samples_per_pool', None)
                    self.sample_id_pattern = settings.get('sample_id_pattern', None)
                    # Custom sample type classification
                    self.ana_samples = settings.get('ana_samples', [])
                    self.sal_samples = settings.get('sal_samples', [])
                    self.pt_samples = settings.get('pt_samples', [])
                    self.tis_samples = settings.get('tis_samples', [])
                    self.int_samples = settings.get('int_samples', [])
                    # H2O settings
                    self.enable_h2o_random = settings.get('enable_h2o_random', False)
                    self.h2o_count = settings.get('h2o_count', 2)
                    self.h2o_position_preference = settings.get('h2o_position_preference', 'random')
                    self.enable_sample_sorting = settings.get('enable_sample_sorting', True)
                    self.generic_sample_category = settings.get('generic_sample_category', 'swab')
                    
                # Debug: Print loaded settings
                print(f"DEBUG: Loaded settings:")
                print(f"  prefix_mapping: {self.prefix_mapping}")
                print(f"  pool_samples: {self.pool_samples}")
                print(f"  no_pool_samples: {self.no_pool_samples}")
                print(f"  samples_per_pool: {self.samples_per_pool}")
                print(f"  sample_id_pattern: {self.sample_id_pattern}")
                print(f"  ana_samples: {self.ana_samples}")
                print(f"  sal_samples: {self.sal_samples}")
                print(f"  pt_samples: {self.pt_samples}")
                print(f"  tis_samples: {self.tis_samples}")
                print(f"  int_samples: {self.int_samples}")
                print(f"  enable_h2o_random: {self.enable_h2o_random}")
                print(f"  h2o_count: {self.h2o_count}")
                print(f"  h2o_position_preference: {self.h2o_position_preference}")
                    
                # Validate that required settings are present if file exists
                if not self.prefix_mapping:
                    raise ValueError("Prefix mapping is required")
                if not self.pool_samples and not self.no_pool_samples:
                    raise ValueError("Pool assignments are required")
                if not self.samples_per_pool:
                    raise ValueError("Samples per pool setting is required")
                if not self.sample_id_pattern:
                    raise ValueError("Sample ID pattern setting is required")
            else:
                # Settings file doesn't exist - start with empty settings
                print(f"DEBUG: Settings file not found, using empty settings")
                self.prefix_mapping = {}
                self.pool_samples = []
                self.no_pool_samples = []
                self.samples_per_pool = None
                self.sample_id_pattern = None
                self.ana_samples = []
                self.sal_samples = []
                self.pt_samples = []
                self.tis_samples = []
                self.int_samples = []
                # Default H2O settings
                self.enable_h2o_random = False
                self.h2o_count = 2
                self.h2o_position_preference = 'random'
                self.enable_sample_sorting = True
                self.generic_sample_category = 'swab'
                
        except Exception as e:
            print(f"Error loading settings: {e}")
            # Start with empty settings on error
            self.prefix_mapping = {}
            self.pool_samples = []
            self.no_pool_samples = []
            self.samples_per_pool = None
            self.sample_id_pattern = None
            self.ana_samples = []
            self.sal_samples = []
            self.pt_samples = []
            self.tis_samples = []
            self.int_samples = []
            # Default H2O settings
            self.enable_h2o_random = False
            self.h2o_count = 2
            self.h2o_position_preference = 'random'
            self.enable_sample_sorting = True
            self.generic_sample_category = 'swab'

    def transform_sample_id(self, sample_id):
        if not isinstance(sample_id, str):
            return sample_id
        for old_prefix, new_prefix in self.prefix_mapping.items():
            if old_prefix in sample_id:
                return sample_id.replace(old_prefix, new_prefix)
        return sample_id

    def reverse_transform_sample_id(self, sample_id):
        """Transform sample ID back to original form"""
        if not isinstance(sample_id, str):
            return str(sample_id)
        
        # Reverse the prefix mapping
        for new_prefix, old_prefix in [(v, k) for k, v in self.prefix_mapping.items()]:
            if sample_id.startswith(new_prefix):
                return sample_id.replace(new_prefix, old_prefix)
        return sample_id

    def split_sample_id(self, sample_id):
        if not isinstance(sample_id, str):
            return str(sample_id)
        
        # Remove underscore and split into prefix and number
        if '_' in sample_id:
            parts = sample_id.split('_')
            if len(parts) == 2:
                prefix = parts[0]
                number = parts[1]
                return f"{prefix}\n{number}"  # Two lines without underscore
        
        # Helper for numeric check
        if sample_id.isdigit():
            return f"\n{sample_id}"
            
        # Fallback for other formats
        sample_id = ''.join(c for c in sample_id if c.isalnum() or c in ['_', '-'])
        try:
            match = re.match(self.sample_id_pattern, sample_id)
            if match:
                prefix = match.group(1) + match.group(2)
                number = match.group(3)
                return f"{prefix}\n{number}"
            else:
                return sample_id
        except:
            return sample_id

    def format_cell_for_sample_id(self, cell, worksheet):
        worksheet.row_dimensions[cell.row].height = 35  # Increase height for two lines
        column_letter = get_column_letter(cell.column)
        worksheet.column_dimensions[column_letter].width = 12  # Adjust width for better fit
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable wrap_text for two lines
        cell.number_format = '@'

    def get_sample_extraction_id(self, sample_ids):
        if not sample_ids:
            return "No samples"
            
        sample_groups = {}

        sample_groups = {}
        for sid in sample_ids:
            if not isinstance(sid, str):
                continue
            try:
                # Group by standard format prefixes if matching
                match = re.match(r'([A-Z]+)(\d{2})_(\d+)', sid)
                if match:
                    prefix = match.group(1) + match.group(2)
                    if prefix not in sample_groups:
                        sample_groups[prefix] = []
                    sample_groups[prefix].append(sid)
            except Exception:
                continue
        
        result = []
        reverse_mapping = {v: k for k, v in self.prefix_mapping.items()}
        
        # Process standard prefix groups
        for prefix in sorted(sample_groups.keys()):
            sids = sample_groups[prefix]
            original_prefix = reverse_mapping.get(prefix, prefix)
            try:
                # Find numerical min/max within the prefix group
                numbers = []
                for sid in sids:
                    parts = sid.split('_')
                    if len(parts) > 1:
                        num_part = ''.join(c for c in parts[-1] if c.isdigit())
                        if num_part:
                            numbers.append(int(num_part))
                
                if numbers:
                    min_num = min(numbers)
                    max_num = max(numbers)
                    if min_num == max_num:
                        result.append(f"{original_prefix}_{min_num:03d}")
                    else:
                        # INCLUDE PREFIX ON BOTH ENDS FOR CLARITY
                        result.append(f"{original_prefix}_{min_num:03d} to {original_prefix}_{max_num:03d}")
                else:
                    # Fallback for prefix-only or non-numeric
                    sorted_sids = sorted(sids, key=natural_sort_key)
                    if len(sorted_sids) > 1:
                        result.append(f"{sorted_sids[0]} - {sorted_sids[-1]}")
                    else:
                        result.append(sorted_sids[0])
            except Exception:
                continue
        
        # Handle "Other/Generic" samples (those not matching the standard prefix regex)
        standard_ids = set()
        for sids in sample_groups.values():
            standard_ids.update(sids)
            
        other_samples = [s for s in sample_ids if isinstance(s, str) and s.strip() and s not in standard_ids and s != 'H2O']
        
        if other_samples:
            # Group other samples by their "prefix" (all characters before the final numeric sequence)
            other_groups = {}
            for sid in other_samples:
                if not sid.strip():
                    continue
                    
                # ROBUST GROUPING for IDs like "21-18141"
                # Logic: Find the last dash or underscore followed by digits, or just the last digits.
                match = re.search(r'^(.*?)([-_])?(\d+)$', sid)
                if match:
                    # Prefix is everything before the final numeric part
                    group_prefix = match.group(1) + (match.group(2) if match.group(2) else '')
                else:
                    group_prefix = sid 
                
                if group_prefix not in other_groups:
                    other_groups[group_prefix] = []
                other_groups[group_prefix].append(sid)
            
            # For each group, create a range
            for g_prefix in sorted(other_groups.keys(), key=natural_sort_key):
                g_sids = sorted(other_groups[g_prefix], key=natural_sort_key)
                g_sids = [s for s in g_sids if s.strip()]
                if not g_sids:
                    continue
                    
                if len(g_sids) > 1:
                    result.append(f"{g_sids[0]} - {g_sids[-1]}")
                else:
                    result.append(g_sids[0])
                
        return " and ".join(result) if result else "No samples in range"

    def generate_h2o_positions(self, max_rows, num_columns, plate_key=None, sample_count=None):
        """Generate random H2O positions, restricted to active columns if sample_count is provided"""
        if plate_key:
            # Normalize seat to ensure "Plate_001" and "Continuous_Plate_001" match
            seed_val = str(plate_key)
            if 'Plate_' in seed_val:
                try:
                    # Extract numeric part: "Plate_001" or "Continuous_Plate_005" -> "001" or "005"
                    import re
                    match = re.search(r'Plate_(\d+)', seed_val)
                    if match:
                        seed_val = match.group(1)
                except:
                    pass
            random.seed(seed_val)
        
        def get_distance(p1, p2):
            return abs(p1[0] - p2[0]) + abs(p1[1] - p2[1])

        if sample_count is not None:
             # Calculate total occupied wells (samples + h2o)
             total_occupied_wells = sample_count + self.h2o_count
             available_linear_indices = list(range(total_occupied_wells))
             
             h2o_positions = []
             occupied_columns = set()
             
             for i in range(self.h2o_count):
                 best_pos = None
                 max_min_dist = -1
                 
                 # Attempt to pick the best spaced position
                 for attempt in range(150):
                     idx = random.choice(available_linear_indices)
                     r, c = idx % max_rows, idx // max_rows
                     
                     # RULE: MAXIMUM 1 H2O PER COLUMN
                     if c in occupied_columns:
                         continue
                     
                     if not h2o_positions:
                         best_pos = (r, c)
                         break
                         
                     min_dist = min(get_distance((r, c), existing) for existing in h2o_positions)
                     if min_dist > max_min_dist:
                         max_min_dist = min_dist
                         best_pos = (r, c)
                     
                     if min_dist >= 3:
                         break
                 
                 # Fallback if no valid unique-column position found (rare, but for high density)
                 if best_pos is None:
                     # Relax column constraint
                     idx = random.choice(available_linear_indices)
                     best_pos = (idx % max_rows, idx // max_rows)

                 h2o_positions.append(best_pos)
                 occupied_columns.add(best_pos[1])
                 
                 # Remove this index from available pool
                 lin_idx = best_pos[1] * max_rows + best_pos[0]
                 if lin_idx in available_linear_indices:
                     available_linear_indices.remove(lin_idx)
             
             print(f"DEBUG: Generated spaced linear H2O positions for {plate_key} (Max 1/col): {h2o_positions}")
             return h2o_positions

        # Fallback for extraction/standard plates (non-linear randomization across the whole grid)
        available_columns = list(range(num_columns))
        if self.h2o_position_preference == 'edges':
            available_rows = [0, max_rows - 1]
        elif self.h2o_position_preference == 'center':
            available_rows = list(range(1, max_rows - 1))
        else:
            available_rows = list(range(max_rows))

        h2o_positions = []
        occupied_columns = set()
        
        for i in range(self.h2o_count):
            best_pos = None
            max_min_dist = -1
            
            for attempt in range(150):
                c = random.choice(available_columns)
                r = random.choice(available_rows)
                
                if (r, c) in h2o_positions or c in occupied_columns:
                    continue
                    
                if not h2o_positions:
                    best_pos = (r, c)
                    break
                    
                min_dist = min(get_distance((r, c), existing) for existing in h2o_positions)
                if min_dist > max_min_dist:
                    max_min_dist = min_dist
                    best_pos = (r, c)
                
                if min_dist >= 3:
                    break
            
            if best_pos is None:
                # Fallback to pure random if distance/column constraints impossible
                c = random.choice(available_columns)
                r = random.choice(available_rows)
                best_pos = (r, c)

            h2o_positions.append(best_pos)
            occupied_columns.add(best_pos[1])
        
        print(f"DEBUG: Generated spaced standard H2O positions for {plate_key} (Max 1/col): {h2o_positions}")
        return h2o_positions

    def create_table_with_h2o(self, samples, max_rows=8, num_columns=12, plate_key=None, format_for_pcr_cdna=False):
        """Create a table layout with H2O positions properly integrated"""
        if not self.enable_h2o_random:
            # Create standard table without H2O
            return self.create_standard_table(samples, max_rows, num_columns, format_for_pcr_cdna)
        
        # Generate H2O positions first
        sample_count = len(samples) if samples else 0
        h2o_positions = self.generate_h2o_positions(max_rows, num_columns, plate_key, sample_count=sample_count)
        
        # Create empty table
        table = [['' for _ in range(num_columns)] for _ in range(max_rows)]
        
        # Mark H2O positions
        h2o_set = set(h2o_positions)
        for row, col in h2o_positions:
            table[row][col] = 'H2O'
        
        # Create pools column by column based on actual H2O positions
        pool_assignments = {}
        
        # Check if samples are already pool names (for PCR/cDNA) - don't create new pools
        is_pool_layout = format_for_pcr_cdna and samples and any(
            isinstance(s, str) and s.startswith('Pool_') for s in samples
        )
        
        # Use a global pool counter that persists across plates (only for extraction, not PCR/cDNA)
        if not hasattr(self, 'global_pool_count'):
            self.global_pool_count = 1
        
        # Track which columns actually got samples
        active_columns = set()
        
        sample_index = 0
        
        print(f"DEBUG: Creating table for {len(samples)} items, is_pool_layout={is_pool_layout}, starting from Pool_{self.global_pool_count:03}")
        
        for col in range(num_columns):
            # Check if this column has H2O
            has_h2o = any(pos[1] == col for pos in h2o_positions)
            column_capacity = max_rows - 1 if has_h2o else max_rows
            
            # Place samples in this column
            if sample_index < len(samples):
                active_columns.add(col)
                if is_pool_layout:
                    # For PCR/cDNA - just place the pool names without creating new pools
                    for row in range(max_rows):
                        if table[row][col] == '' and sample_index < len(samples):
                            sample_name = samples[sample_index]
                            if '_' in str(sample_name):
                                sample_name = str(sample_name).replace('_', '\n')
                            
                            table[row][col] = sample_name
                            sample_index += 1
                else:
                    # For Extraction - create new pool names
                    if self.samples_per_pool:
                        pool_name = f"Pool_{self.global_pool_count:03}"
                        
                        # Place samples in this column
                        placed_in_column_count = 0
                        for row in range(max_rows):
                            if table[row][col] == '' and sample_index < len(samples):
                                sample_name = samples[sample_index]
                                if format_for_pcr_cdna and '_' in sample_name:
                                    sample_name = sample_name.replace('_', '\n')
                                
                                table[row][col] = sample_name
                                pool_assignments[samples[sample_index]] = pool_name
                                
                                sample_index += 1
                                placed_in_column_count += 1
                        
                        # Only increment pool count if we placed samples (for extraction only)
                        if placed_in_column_count > 0:
                            self.global_pool_count += 1
        
        # REMOVE H2O from empty columns (columns with only H2O but no samples)
        for col in range(num_columns):
            if col not in active_columns:
                for row in range(max_rows):
                    if table[row][col] == 'H2O':
                        table[row][col] = ''
        
        print(f"DEBUG: Placed {sample_index} items, is_pool_layout={is_pool_layout}")
        
        # Store pool assignments for later use
        if not hasattr(self, 'column_pool_assignments'):
            self.column_pool_assignments = {}
        self.column_pool_assignments[plate_key] = pool_assignments
        
        return table

    def create_standard_table(self, samples, max_rows=8, num_columns=12, format_for_pcr_cdna=False):
        """Create standard table without H2O"""
        table = []
        for col in range(num_columns):
            column = []
            for row in range(max_rows):
                index = col * max_rows + row
                if index < len(samples):
                    sample_name = samples[index]
                    if format_for_pcr_cdna and '_' in sample_name:
                        sample_name = sample_name.replace('_', '\n')
                    column.append(sample_name)
                else:
                    column.append('')
            table.append(column)
        
        # Transpose to get rows
        table = list(zip(*table))
        table = [list(row) for row in table]
        return table

    def get_existing_done_samples(self, destination_file):
        """Get list of sample IDs that are already marked as Done in the output file"""
        try:
            if not os.path.exists(destination_file):
                return set()
            
            # Try to read the Excel file
            try:
                df = pd.read_excel(destination_file, sheet_name='Extraction', engine='openpyxl')
            except:
                # Try without sheet specification
                df = pd.read_excel(destination_file)
            
            if df.empty or 'Sample_Id' not in df.columns or 'Status' not in df.columns:
                return set()
            
            # Get samples with Done status
            done_samples = df[df['Status'] == 'Done']['Sample_Id'].tolist()
            print(f"DEBUG: Found {len(done_samples)} samples already marked as Done")
            return set(done_samples)
            
        except Exception as e:
            print(f"DEBUG: Error reading existing Done samples: {e}")
            return set()

    def pre_save_formatting(self, worksheet):
        for row in worksheet.iter_rows(min_row=8, max_row=15, min_col=2, max_col=13):
            for cell in row:
                if cell.value and '\n' in str(cell.value):
                    parts = str(cell.value).split('\n')
                    cell.value = parts[0] + chr(10) + parts[1]
                    self.format_cell_for_sample_id(cell, worksheet)
        for row in worksheet.iter_rows(min_row=26, max_row=33, min_col=2, max_col=13):
            for cell in row:
                if cell.value and '\n' in str(cell.value):
                    parts = str(cell.value).split('\n')
                    cell.value = parts[0] + chr(10) + parts[1]
                    self.format_cell_for_sample_id(cell, worksheet)
        for row in worksheet.iter_rows(min_row=44, max_row=51, min_col=2, max_col=13):
            for cell in row:
                if cell.value and '\n' in str(cell.value):
                    parts = str(cell.value).split('\n')
                    cell.value = parts[0] + chr(10) + parts[1]
                    self.format_cell_for_sample_id(cell, worksheet)

    def process_ana_sal_samples(self, destination_file, current_step, total_steps):
        try:
            processing_progress['message'] = f"Loading settings: pool_samples={len(self.pool_samples)}, no_pool_samples={len(self.no_pool_samples)}"
            
            # Get existing Done samples to skip them
            done_samples = self.get_existing_done_samples(destination_file)
            processing_progress['message'] = f"Found {len(done_samples)} samples already completed, will skip them"
            
            # Try to read Excel file with comprehensive error handling
            try:
                # First, check if file exists and is readable
                if not os.path.exists(destination_file):
                    raise FileNotFoundError(f"Destination file not found: {destination_file}")
                
                # Check file extension and try to handle different formats
                file_ext = os.path.splitext(destination_file)[1].lower()
                processing_progress['message'] = f"Reading Excel file (format: {file_ext})..."
                
                # Try different approaches to read the Excel file
                df = None
                error_messages = []
                
                # Method 1: Try with openpyxl (most reliable for .xlsx)
                try:
                    df = pd.read_excel(destination_file, sheet_name='Extraction', engine='openpyxl')
                    processing_progress['message'] = f"Successfully read Excel with openpyxl engine"
                except Exception as e:
                    error_messages.append(f"openpyxl failed: {str(e)}")
                
                # Method 2: Try with default engine
                if df is None:
                    try:
                        df = pd.read_excel(destination_file, sheet_name='Extraction')
                        processing_progress['message'] = f"Successfully read Excel with default engine"
                    except Exception as e:
                        error_messages.append(f"Default engine failed: {str(e)}")
                
                # Method 3: Try reading without specifying sheet name first
                if df is None:
                    try:
                        # Try a more direct approach without ExcelFile
                        try:
                            df = pd.read_excel(destination_file, sheet_name=0)  # Try first sheet by index
                            processing_progress['message'] = f"Successfully read first sheet by index"
                        except Exception as index_error:
                            error_messages.append(f"Index-based reading failed: {str(index_error)}")
                            
                            # Try with different parameters
                            try:
                                df = pd.read_excel(destination_file, header=None)  # Read without header assumptions
                                processing_progress['message'] = f"Successfully read Excel without header processing"
                            except Exception as header_error:
                                error_messages.append(f"No-header reading failed: {str(header_error)}")
                                
                                # Last resort - try with all available engines
                                engines_to_try = ['openpyxl', 'xlrd', None]
                                for engine in engines_to_try:
                                    try:
                                        if engine:
                                            df = pd.read_excel(destination_file, engine=engine)
                                        else:
                                            df = pd.read_excel(destination_file)
                                        processing_progress['message'] = f"Successfully read with engine: {engine or 'default'}"
                                        break
                                    except Exception as engine_error:
                                        error_messages.append(f"Engine {engine or 'default'} failed: {str(engine_error)}")
                                        continue
                    
                    except Exception as e:
                        error_messages.append(f"Direct reading method failed: {str(e)}")
                
                # Method 4: Last resort - try to read any Excel data
                if df is None:
                    try:
                        # Try reading without sheet specification
                        df = pd.read_excel(destination_file)
                        processing_progress['message'] = f"Read Excel without sheet specification"
                    except Exception as e:
                        error_messages.append(f"Last resort method failed: {str(e)}")
                
                # If all methods failed, provide detailed error and skip processing
                if df is None:
                    processing_progress['error'] = f"All Excel reading methods failed. File: {destination_file}. Errors: {'; '.join(error_messages)}"
                    processing_progress['message'] = "Skipping ANA/SAL processing due to Excel read error"
                    return  # Skip ANA/SAL processing but don't crash
                
                # Validate dataframe
                if df.empty:
                    processing_progress['message'] = "Excel file is empty - skipping ANA/SAL processing"
                    return
                    
                processing_progress['message'] = f"Excel loaded successfully. Shape: {df.shape}"
                    
            except Exception as e:
                processing_progress['error'] = f'Excel reading error: {str(e)}'
                processing_progress['message'] = "Skipping ANA/SAL processing due to Excel read error"
                return  # Skip processing but continue with other steps
            
            # Ensure Sample_Id is string to handle numeric IDs correctly
            if 'Sample_Id' not in df.columns:
                processing_progress['error'] = f"Error processing ANA/SAL samples: 'Sample_Id' column not found in the 'Extraction' sheet. Available columns: {df.columns.tolist()}"
                processing_progress['message'] = "Aborting ANA/SAL processing due to missing Sample_Id column"
                return

            df['Sample_Id'] = df['Sample_Id'].astype(str)
            # Remove "nan" strings that might result from converting actual NaNs
            df = df[df['Sample_Id'] != 'nan']
                    
            processing_progress['message'] = f"Processing ANA/SAL samples..."
            
            # Use custom sample type classification instead of automatic detection
            if hasattr(self, 'ana_samples') and self.ana_samples:
                ana_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.ana_samples]
            else:
                ana_prefixes = []
            
            if hasattr(self, 'sal_samples') and self.sal_samples:
                sal_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.sal_samples]
            else:
                sal_prefixes = []
                
            if hasattr(self, 'pt_samples') and self.pt_samples:
                pt_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.pt_samples]
            else:
                pt_prefixes = []
            print(f"DEBUG: Prefix breakdown:")
            print(f"  ANA prefixes: {ana_prefixes}")
            print(f"  SAL prefixes: {sal_prefixes}")
            print(f"  PT prefixes: {pt_prefixes}")
            swab_prefixes = ana_prefixes + sal_prefixes + pt_prefixes
            swab_pattern = '|'.join(swab_prefixes) if swab_prefixes else "IMPOSSIBLE_PATTERN_XYZ"
            
            # Find generic/other samples (numeric or non-matching) if routed to swab
            generic_samples_df = df.iloc[0:0].copy()
            if self.generic_sample_category == 'swab':
                # Known pattern includes ALL possible prefixes (ANA, SAL, PT, TIS, INT)
                all_known_prefixes = swab_prefixes
                if hasattr(self, 'tis_samples') and self.tis_samples:
                    all_known_prefixes += [self.prefix_mapping.get(prefix, prefix) for prefix in self.tis_samples]
                if hasattr(self, 'int_samples') and self.int_samples:
                    all_known_prefixes += [self.prefix_mapping.get(prefix, prefix) for prefix in self.int_samples]
                
                all_known_pattern = '|'.join(all_known_prefixes) if all_known_prefixes else "IMPOSSIBLE_PATTERN_XYZ"
                
                # Get valid samples that are NOT "Done", NOT matching any known prefix, and look numeric
                generic_samples_df = df[
                    ~df['Sample_Id'].isin(done_samples) & 
                    (df['Status'] == 'in process') & 
                    ~df['Sample_Id'].str.contains(all_known_pattern, na=False) &
                    df['Sample_Id'].str.contains(r'\d', na=False)
                ]
                print(f"DEBUG: Generic/Numeric samples routed to Swab phase: {len(generic_samples_df)}")

            # Combined dataframe for all samples to be processed in this Swab phase
            # Filter out samples that are already Done
            matching_prefixed_samples = df[df['Sample_Id'].str.contains(swab_pattern, na=False) & (df['Status'] == 'in process') & ~df['Sample_Id'].isin(done_samples)]
            ana_sal_pt_samples = pd.concat([matching_prefixed_samples, generic_samples_df]).drop_duplicates(subset=['Sample_Id'])
            
            # Filter ANA samples (transformed prefixes from original ANA prefixes)
            if ana_prefixes:
                ana_sample_ids = matching_prefixed_samples[matching_prefixed_samples['Sample_Id'].str.contains('|'.join(ana_prefixes), na=False)]['Sample_Id'].tolist()
            else:
                ana_sample_ids = []
            if ana_prefixes and getattr(self, 'enable_sample_sorting', True):
                ana_sample_ids.sort(key=natural_sort_key)
            processing_progress['message'] = f"ANA samples found: {len(ana_sample_ids)}"
            
            # Filter SAL samples (transformed prefixes from original SAL prefixes)
            if sal_prefixes:
                sal_sample_ids = matching_prefixed_samples[matching_prefixed_samples['Sample_Id'].str.contains('|'.join(sal_prefixes), na=False)]['Sample_Id'].tolist()
            else:
                sal_sample_ids = []
            if sal_prefixes and getattr(self, 'enable_sample_sorting', True):
                sal_sample_ids.sort(key=natural_sort_key)
            processing_progress['message'] = f"SAL samples found: {len(sal_sample_ids)}"
            
            # Filter PT samples (transformed prefixes from original CANA_PT prefixes)
            if pt_prefixes:
                pt_sample_ids = matching_prefixed_samples[matching_prefixed_samples['Sample_Id'].str.contains('|'.join(pt_prefixes), na=False)]['Sample_Id'].tolist()
            else:
                pt_sample_ids = []  # No PT prefixes, so no PT samples
            if pt_prefixes and getattr(self, 'enable_sample_sorting', True):
                pt_sample_ids.sort(key=natural_sort_key)
            print(f"DEBUG: PT sample_ids: {len(pt_sample_ids)}")
            processing_progress['message'] = f"PT samples found: {len(pt_sample_ids)}"
            
            # Generic samples list for ordering
            generic_samples = generic_samples_df['Sample_Id'].tolist()
            if generic_samples and getattr(self, 'enable_sample_sorting', True):
                generic_samples.sort(key=natural_sort_key)
            
            ordered_sample_ids = []
            if getattr(self, 'enable_sample_sorting', True):
                # Order by original source based on transformed prefixes with sorting
                canb_prefixes = [new_prefix for old_prefix, new_prefix in self.prefix_mapping.items() if 'CANB_' in old_prefix]
                canb_samples = [sid for sid in ana_sample_ids + sal_sample_ids if any(p in sid for p in canb_prefixes)]
                ordered_sample_ids.extend(sorted(canb_samples, key=natural_sort_key))
                
                canr_prefixes = [new_prefix for old_prefix, new_prefix in self.prefix_mapping.items() if 'CANR_' in old_prefix]
                canr_samples = [sid for sid in ana_sample_ids + sal_sample_ids if any(p in sid for p in canr_prefixes)]
                ordered_sample_ids.extend(sorted(canr_samples, key=natural_sort_key))
                
                iplnahl_prefixes = [new_prefix for old_prefix, new_prefix in self.prefix_mapping.items() if 'IPLNAHL_' in old_prefix]
                iplnahl_samples = [sid for sid in ana_sample_ids + sal_sample_ids if any(p in sid for p in iplnahl_prefixes)]
                ordered_sample_ids.extend(sorted(iplnahl_samples, key=natural_sort_key))
                
                ordered_sample_ids.extend(sorted(pt_sample_ids, key=natural_sort_key))
                
                # Add generic samples at the end
                ordered_sample_ids.extend(sorted(generic_samples, key=natural_sort_key))
            else:
                # ABSOLUTE FILE ORDER: find where these IDs appear in the original df
                # We use all IDs identified for this phase
                all_candidate_ids = set(ana_sample_ids + sal_sample_ids + pt_sample_ids + generic_samples)
                # Filter 'df' (the original full sheet) to get the exact order
                ordered_sample_ids = [sid for sid in df['Sample_Id'] if sid in all_candidate_ids]
                # Unique IDs only, preserving order
                seen = set()
                ordered_sample_ids = [x for x in ordered_sample_ids if not (x in seen or seen.add(x))]
            # Remove duplicates from ordered_sample_ids to prevent multiple assignments
            print(f"DEBUG: Before deduplication: {len(ordered_sample_ids)} samples")
            ordered_sample_ids = list(dict.fromkeys(ordered_sample_ids))
            print(f"DEBUG: After deduplication: {len(ordered_sample_ids)} samples")
            print(f"DEBUG: Sample IDs: {ordered_sample_ids[:10]}...")  # Show first 10
            transformed_sample_ids = [self.transform_sample_id(sid) for sid in ordered_sample_ids]
            pool_assignments = {}
            
            # Apply custom pool assignments from environment variables
            # First, handle explicitly pooled samples
            pooled_samples = []
            not_pooled_samples = []
            auto_pooled_samples = []
            
            for sample_id in ordered_sample_ids:
                # Transform pool prefixes to match transformed sample IDs
                transformed_pool_prefixes = [self.transform_sample_id(prefix) for prefix in self.pool_samples]
                transformed_no_pool_prefixes = [self.transform_sample_id(prefix) for prefix in self.no_pool_samples]
                
                # Debug: Print what we're checking
                print(f"DEBUG: Checking sample {sample_id} against pool prefixes {transformed_pool_prefixes}")
                
                # Check if sample ID starts with any transformed pooled prefix
                is_pooled = any(sample_id.startswith(pool_prefix) for pool_prefix in transformed_pool_prefixes)
                is_not_pooled = any(sample_id.startswith(no_pool_prefix) for no_pool_prefix in transformed_no_pool_prefixes)
                
                print(f"DEBUG: Sample {sample_id} - is_pooled: {is_pooled}, is_not_pooled: {is_not_pooled}")
                
                if is_pooled:
                    pooled_samples.append(sample_id)
                elif is_not_pooled:
                    not_pooled_samples.append(sample_id)
                else:
                    auto_pooled_samples.append(sample_id)
            
            # Assign pools to explicitly pooled samples (8 samples per pool)
            print(f"DEBUG: Pool assignment summary:")
            print(f"  Pooled samples: {len(pooled_samples)}")
            print(f"  Not pooled samples: {len(not_pooled_samples)}")
            print(f"  Auto pooled samples: {len(auto_pooled_samples)}")
            
            pool_count = 1
            # Group pooled samples column by column based on H2O positions
            if self.enable_h2o_random and self.h2o_count > 0:
                print(f"DEBUG: Column-by-column pooling - H2O count: {self.h2o_count}")
                # When H2O is enabled, defer pool creation to table creation phase
                # This allows pools to be created based on actual column capacity
                print(f"DEBUG: Deferring pool creation for {len(pooled_samples)} samples to table creation")
                # Create a placeholder pool assignment that will be updated during table creation
                for sample_id in pooled_samples:
                    df.loc[df['Sample_Id'] == sample_id, 'Pool'] = "TO_BE_ASSIGNED"
                    transformed_sample = self.transform_sample_id(sample_id)
                    pool_assignments[transformed_sample] = "TO_BE_ASSIGNED"
            else:
                # Standard pooling when H2O is disabled
                pool_count = 1
                for i in range(0, len(pooled_samples), self.samples_per_pool):
                    pool_name = f"Pool_{pool_count:03}"
                    pool_sample_ids = pooled_samples[i:i + self.samples_per_pool]
                    for sample_id in pool_sample_ids:
                        print(f"DEBUG: Assigning {pool_name} to sample {sample_id}")
                        df.loc[df['Sample_Id'] == sample_id, 'Pool'] = pool_name
                        transformed_sample = self.transform_sample_id(sample_id)
                        pool_assignments[transformed_sample] = pool_name
                    pool_count += 1
            
            # Assign "Not Pool" to samples that should not be pooled
            for sample_id in not_pooled_samples:
                df.loc[df['Sample_Id'] == sample_id, 'Pool'] = "Not Pool"
                transformed_sample = self.transform_sample_id(sample_id)
                pool_assignments[transformed_sample] = "Not Pool"
            
            # No automatic pooling - only samples explicitly in pool_samples get pooled
            # All other samples that weren't explicitly assigned get "Not Pool"
            remaining_samples = set(ordered_sample_ids) - set(pooled_samples) - set(not_pooled_samples)
            for sample_id in remaining_samples:
                df.loc[df['Sample_Id'] == sample_id, 'Pool'] = "Not Pool"
                transformed_sample = self.transform_sample_id(sample_id)
                pool_assignments[transformed_sample] = "Not Pool"
            
            # Create a mapping from original to transformed sample IDs
            sample_id_mapping = {original: transformed for original, transformed in zip(ordered_sample_ids, transformed_sample_ids)}
            # Update the dataframe with transformed sample IDs
            print(f"DEBUG: Before transformation - DataFrame has {len(df)} rows")
            print(f"DEBUG: Sample IDs to transform: {len(ordered_sample_ids)}")
            for original, transformed in zip(ordered_sample_ids, transformed_sample_ids):
                df.loc[df['Sample_Id'] == original, 'Sample_Id'] = transformed
            print(f"DEBUG: After transformation - DataFrame has {len(df)} rows")
            
            os.makedirs('Extraction', exist_ok=True)
            ana_sal_pt_samples_copy = ana_sal_pt_samples.copy()
            original_pool_mapping = {}
            
            # Rebuild pool mapping using column-by-column logic
            pool_count = 1
            original_pool_mapping = {}
            
            # Handle pooled samples
            if self.enable_h2o_random and self.h2o_count > 0:
                print(f"DEBUG: Deferring pool creation for check list - will be handled in table creation")
                # For pooled samples, create placeholder assignments
                for sample_id in pooled_samples:
                    original_pool_mapping[sample_id] = "TO_BE_ASSIGNED"
            else:
                # Standard pooling when H2O is disabled
                for sample_id in pooled_samples:
                    pool_name = f"Pool_{pool_count:03}"
                    original_pool_mapping[sample_id] = pool_name
                    pool_count += 1
            
            # Not pooled samples
            for sample_id in not_pooled_samples:
                original_pool_mapping[sample_id] = "Not Pool"
            
            # Auto-pooled samples
            if self.enable_h2o_random and self.h2o_count > 0:
                print(f"DEBUG: Deferring auto-pool creation for check list - will be handled in table creation")
                # For auto-pooled samples, create placeholder assignments
                for sample_id in auto_pooled_samples:
                    original_pool_mapping[sample_id] = "TO_BE_ASSIGNED"
            else:
                # Standard auto-pooling when H2O is disabled
                for i in range(0, len(auto_pooled_samples), self.samples_per_pool):
                    pool_name = f"Pool_{pool_count:03}"
                    pool_sample_ids = auto_pooled_samples[i:i + self.samples_per_pool]
                    for sid in pool_sample_ids:
                        original_pool_mapping[sid] = pool_name
                    pool_count += 1
            
            # Map transformed IDs to pools directly
            transformed_pool_mapping = {sample_id_mapping.get(sid, sid): pool for sid, pool in original_pool_mapping.items()}
            # Update sample IDs in the check list dataframe
            ana_sal_pt_samples_copy['Sample_Id'] = ana_sal_pt_samples_copy['Sample_Id'].map(lambda x: sample_id_mapping.get(x, x))
            # Assign pool values using the transformed IDs
            ana_sal_pt_samples_copy['Pool'] = ana_sal_pt_samples_copy['Sample_Id'].map(transformed_pool_mapping)
            
            def get_sort_key(sample_id):
                if 'CANB_' in sample_id:
                    return (0, sample_id)
                elif 'CANR_' in sample_id:
                    return (1, sample_id)
                elif 'IPLNAHL_' in sample_id:
                    return (2, sample_id)
                else:
                    return (3, sample_id)
            
            if getattr(self, 'enable_sample_sorting', True):
                ana_sal_pt_samples_copy['sort_key'] = ana_sal_pt_samples_copy['Sample_Id'].apply(get_sort_key)
                ana_sal_pt_samples_copy = ana_sal_pt_samples_copy.sort_values('sort_key')
                ana_sal_pt_samples_copy = ana_sal_pt_samples_copy.drop('sort_key', axis=1)
            else:
                # Maintain file order for the check list as well
                # The data in 'original_check_list_data' (later) will follow 'ordered_sample_ids'
                pass
            
            # Track plate assignments for each sample
            sample_plate_assignments = {}
            
            extraction_tables = []
            num_columns = 12  # Changed from 11 to 12 as requested
            max_rows_per_column = 8
            items_per_plate = num_columns * max_rows_per_column  # Restored full capacity
            
            # Get the starting plate number from existing Extraction_Tables.xlsx
            file_path = 'Extraction/Extraction_Tables.xlsx'
            file_is_valid = False
            
            if not os.path.exists(file_path):
                next_index = 1  # Start with Plate_001 when file doesn't exist
            else:
                # Check if file is valid (not empty and readable)
                if os.path.getsize(file_path) == 0:
                    print(f"DEBUG: File {file_path} exists but is empty (0 bytes)")
                    next_index = 1
                    # File is invalid, we will overwrite it
                else:
                    try:
                        existing_sheets = pd.ExcelFile(file_path).sheet_names
                        # Extract existing plate numbers and find the highest
                        existing_numbers = []
                        for sheet in existing_sheets:
                            if sheet.startswith('Plate_'):
                                try:
                                    num = int(sheet.replace('Plate_', ''))
                                    existing_numbers.append(num)
                                except ValueError:
                                    continue
                        next_index = max(existing_numbers) + 1 if existing_numbers else 1
                        file_is_valid = True
                    except Exception as e:
                        print(f"DEBUG: Error reading existing sheets (file likely corrupt): {e}")
                        next_index = 1
                        # File is invalid/corrupt, we will overwrite it
            
            # Create tables with H2O integration
            print(f"DEBUG: Creating {len(transformed_sample_ids)} extraction samples with H2O enabled={self.enable_h2o_random}")
            
            # Calculate how many samples can fit per plate accounting for H2O
            total_positions = num_columns * max_rows_per_column  # 96 positions (12x8)
            samples_per_plate = total_positions - self.h2o_count if self.enable_h2o_random else total_positions
            print(f"DEBUG: Total positions per plate: {total_positions}, H2O count: {self.h2o_count}, Samples per plate: {samples_per_plate}")
            
            plate_index = 0
            for plate_start in range(0, len(transformed_sample_ids), samples_per_plate):
                plate_samples = transformed_sample_ids[plate_start:plate_start + samples_per_plate]
                plate_key = f"Plate_{next_index + plate_index:03d}"
                print(f"DEBUG: Creating Extraction table {plate_key} with {len(plate_samples)} samples (should be {samples_per_plate})")
                
                # Create table with H2O positions properly integrated
                current_table = self.create_table_with_h2o(plate_samples, max_rows_per_column, num_columns, plate_key)
                print(f"DEBUG: Created table with {sum(1 for row in current_table for cell in row if cell == 'H2O')} H2O positions")
                
                # Update pool assignments from table creation if H2O is enabled
                if self.enable_h2o_random and hasattr(self, 'column_pool_assignments') and plate_key in self.column_pool_assignments:
                    pool_assignments_from_table = self.column_pool_assignments[plate_key]
                    print(f"DEBUG: Updating {len(pool_assignments_from_table)} pool assignments from table creation")
                    
                    # Update the main pool_assignments dictionary
                    for sample_id, pool_name in pool_assignments_from_table.items():
                        pool_assignments[sample_id] = pool_name
                    
                    # Update the dataframe as well
                    for sample_id, pool_name in pool_assignments_from_table.items():
                        df.loc[df['Sample_Id'] == sample_id, 'Pool'] = pool_name
                
                # Assign plate numbers to samples (skip H2O positions)
                for row_idx, row in enumerate(current_table):
                    for col_idx, cell in enumerate(row):
                        if cell and cell != 'H2O':
                            sample_plate_assignments[cell] = f'Plate_{next_index + plate_index:03d}'
                
                extraction_tables.append(current_table)
                plate_index += 1
            
            # After all plates are created, update the check list with final pool assignments
            print(f"DEBUG: Final pool assignments: {len(pool_assignments)} entries")
            
            # Add plate information to check list
            ana_sal_pt_samples_copy['Plate'] = ana_sal_pt_samples_copy['Sample_Id'].map(sample_plate_assignments)
            
            # Update pool assignments in check list dataframe
            for transformed_id, pool_name in pool_assignments.items():
                ana_sal_pt_samples_copy.loc[ana_sal_pt_samples_copy['Sample_Id'] == transformed_id, 'Pool'] = pool_name
            
            # Reorder columns to include Plate
            if ana_sal_pt_samples_copy.empty:
                print(f"DEBUG: ANA/SAL dataframe is empty, skipping column reordering")
                self.swab_check_list = pd.DataFrame(columns=['Sample_Id', 'Pool', 'Plate', 'Status'])
                return
            
            columns = ana_sal_pt_samples_copy.columns.tolist()
            print(f"DEBUG: Available columns: {columns}")
            
            # Check if columns exist before trying to remove them
            for col in ['Pool', 'Status', 'Plate']:
                if col in columns:
                    columns.remove(col)
                    print(f"DEBUG: Removed column {col}")
                else:
                    print(f"DEBUG: Column {col} not found in dataframe")
            
            columns.extend(['Pool', 'Plate', 'Status'])
            ana_sal_pt_samples_copy = ana_sal_pt_samples_copy[columns]
            
            # Add logging to indicate what's being saved
            processing_progress['message'] = f"Writing {len(ana_sal_pt_samples_copy)} samples to Check_list_Swab-sample-table.xlsx"
            
            # Debug output first few rows of data to verify sample IDs
            if not ana_sal_pt_samples_copy.empty:
                sample_examples = ana_sal_pt_samples_copy['Sample_Id'].head(5).tolist()
                processing_progress['message'] = f"Sample ID examples: {sample_examples}"
            
            # Create check list with INDIVIDUAL sample entries (matching extraction table)
            original_check_list_data = []
            
            # Create reverse mapping from transformed back to original
            reverse_sample_id_mapping = {transformed: original for original, transformed in zip(ordered_sample_ids, transformed_sample_ids)}
            
            # Map plate assignments from transformed IDs back to original IDs
            original_plate_assignments = {}
            for transformed_id, plate in sample_plate_assignments.items():
                original_id = reverse_sample_id_mapping.get(transformed_id, transformed_id)
                original_plate_assignments[original_id] = plate
            
            # Map pool assignments from transformed IDs back to original IDs
            original_pool_assignments = {}
            for transformed_id, pool in pool_assignments.items():
                original_id = reverse_sample_id_mapping.get(transformed_id, transformed_id)
                original_pool_assignments[original_id] = pool
            
            # Create individual entries for each sample
            for original_id in ordered_sample_ids:
                check_list_entry = {
                    'Sample_Id': original_id,
                    'Plate': original_plate_assignments.get(original_id, ''),
                    'Pool': original_pool_assignments.get(original_id, 'Not Pool'),
                    'Status': 'in process'
                }
                original_check_list_data.append(check_list_entry)
            
            original_check_list = pd.DataFrame(original_check_list_data)
            
            # Reorder columns for original check list
            orig_columns = original_check_list.columns.tolist()
            for col in ['Pool', 'Plate', 'Status']:
                if col in orig_columns:
                    orig_columns.remove(col)
            orig_columns.extend(['Pool', 'Plate', 'Status'])
            original_check_list = original_check_list[orig_columns]
            
            # Store original check list data for later combination
            self.swab_check_list = original_check_list
            
            # Use the existing file_path and next_index from above
            if os.path.exists(file_path) and file_is_valid:
                mode = 'a'
                if_sheet_exists = 'overlay'
            else:
                print(f"DEBUG: Creating new/overwriting Extraction_Tables.xlsx (Valid: {file_is_valid})")
                mode = 'w'
                if_sheet_exists = None
                
            with pd.ExcelWriter(file_path, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
                for table_index, table in enumerate(extraction_tables):
                    # Table already has H2O positions integrated
                    modified_table = table
                    
                    extraction_df = pd.DataFrame(modified_table)
                    # Count only actual samples (not H2O) for sample range calculation
                    table_sample_ids = []
                    for row in modified_table:
                        for cell in row:
                            if cell and cell != 'H2O':
                                table_sample_ids.append(cell)
                    sample_extraction_id = self.get_sample_extraction_id(table_sample_ids)
                    
                    # Snapshot this plate for 100% synchronization with cDNA/PCR
                    plate_number = next_index + table_index
                    self.extraction_plate_snapshots.append({
                        'plate_number': plate_number,
                        'range_header': sample_extraction_id,
                        'table': table,
                        'type': 'swab',
                        'pool_assignments': pool_assignments.copy() # Store specific mapping for this phase
                    })
                    column_pools = [None] * num_columns
                    for col in range(num_columns):
                        for row in range(len(modified_table)):
                            sample_id = modified_table[row][col]
                            if sample_id and sample_id != 'H2O' and sample_id in pool_assignments:
                                column_pools[col] = pool_assignments[sample_id]
                                break
                    sheet_name = f'Plate_{plate_number:03d}'
                    workbook = writer.book
                    worksheet = workbook.create_sheet(sheet_name) if sheet_name not in workbook.sheetnames else workbook[sheet_name]
                    
                    # Define formatting styles
                    sample_id_font = Font(name='Arial', size=14, bold=False)
                    pool_font = Font(name='Arial', size=14, bold=True)
                    col_row_font = Font(name='Arial', size=22, bold=True)
                    header_font = Font(name='Arial', size=14, bold=True)
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    fills = [
                        PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
                        PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),  # Sky Blue
                        PatternFill(start_color="808080", end_color="808080", fill_type="solid")   # Gray
                    ]
                    
                    # Write three tables vertically stacked
                    for variant in range(3):
                        row_offset = variant * 18  # 18 rows between each table (including spacing)
                        
                        # Headers and metadata with font size 14
                        worksheet[f'I{3 + row_offset}'] = "Date:"
                        worksheet[f'I{3 + row_offset}'].font = header_font
                        worksheet[f'D{3 + row_offset}'] = "Inactivated by:"
                        worksheet[f'D{3 + row_offset}'].font = header_font
                        worksheet[f'C{1 + row_offset}'] = "Protocol: "
                        worksheet[f'C{1 + row_offset}'].font = header_font
                        worksheet[f'F{1 + row_offset}'] = "Kitt#: "
                        worksheet[f'F{1 + row_offset}'].font = header_font
                        worksheet[f'I{1 + row_offset}'] = "Lot#:"
                        worksheet[f'I{1 + row_offset}'].font = header_font
                        worksheet[f'L{1 + row_offset}'] = "Expiry:"
                        worksheet[f'L{1 + row_offset}'].font = header_font
                        worksheet[f'C{6 + row_offset}'] = "Sample Id"
                        worksheet[f'C{6 + row_offset}'].font = header_font
                        worksheet[f'E{6 + row_offset}'] = sample_extraction_id
                        worksheet[f'E{6 + row_offset}'].font = header_font
                        worksheet[f'I{5 + row_offset}'] = "Date:"
                        worksheet[f'I{5 + row_offset}'].font = header_font
                        worksheet[f'D{5 + row_offset}'] = "Extracted by:"
                        worksheet[f'D{5 + row_offset}'].font = header_font
                        
                        # Column headers
                        for col in range(num_columns):
                            cell = worksheet.cell(row=7 + row_offset, column=col + 2)
                            cell.value = str(col + 1)
                            cell.font = col_row_font
                            if fills[variant]:
                                cell.fill = fills[variant]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Row headers
                        for row in range(max_rows_per_column):
                            cell = worksheet.cell(row=row + 8 + row_offset, column=1)
                            cell.value = chr(65 + row)
                            cell.font = col_row_font
                            if fills[variant]:
                                cell.fill = fills[variant]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Sample IDs
                        for row in range(len(extraction_df)):
                            for col in range(len(extraction_df.columns)):
                                sample_id = extraction_df.iloc[row, col]
                                if sample_id:
                                    cell = worksheet.cell(row=row + 8 + row_offset, column=col + 2)
                                    cell.value = self.split_sample_id(sample_id) if sample_id != 'H2O' else 'H2O'
                                    if sample_id == 'H2O':
                                        cell.font = Font(name='Arial', size=14, bold=True, color="FF0000")  # Red bold for H2O
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red background
                                    else:
                                        cell.font = sample_id_font
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable wrap_text for two lines
                                        self.format_cell_for_sample_id(cell, worksheet)
                                    cell.border = thin_border
                        
                        # Pool row (no fill applied)
                        for col in range(num_columns):
                            if column_pools[col]:
                                cell = worksheet.cell(row=16 + row_offset, column=col + 2)
                                cell.value = column_pools[col]
                                cell.font = pool_font
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Plate number
                        worksheet[f'N{8 + row_offset}'] = str(plate_number)
                        worksheet.merge_cells(f'N{8 + row_offset}:N{15 + row_offset}')
                        merged_cell = worksheet[f'N{8 + row_offset}']
                        merged_cell.font = Font(name='Arial', size=22, bold=True)
                        if fills[variant]:
                            merged_cell.fill = fills[variant]
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        merged_cell.border = Border(left=Side(style='thin'))
                        
                        # Set dimensions
                        for row in range(8 + row_offset, 16 + row_offset):
                            worksheet.row_dimensions[row].height = 35
                        for col in range(2, num_columns + 2):
                            worksheet.column_dimensions[get_column_letter(col)].width = 10
                    self.pre_save_formatting(worksheet)
            
            processing_progress['message'] = f"Created {len(extraction_tables)} Swab sample tables (3 tables per sheet)"
            
        except Exception as e:
            processing_progress['error'] = f'Error processing ANA/SAL samples: {str(e)}'

    def process_tis_int_samples(self, destination_file, current_step, total_steps):
        try:
            # Get existing Done samples to skip them
            done_samples = self.get_existing_done_samples(destination_file)
            processing_progress['message'] = f"Found {len(done_samples)} samples already completed, will skip them"
            
            # Try to read Excel file with better error handling
            try:
                df = pd.read_excel(destination_file, sheet_name='Extraction', engine='openpyxl')
            except Exception as e:
                # If openpyxl fails, try with different engine
                try:
                    df = pd.read_excel(destination_file, sheet_name='Extraction', engine='xlrd')
                except Exception as e2:
                    # If all engines fail, try basic read
                    df = pd.read_excel(destination_file, sheet_name='Extraction')
                    
            # Ensure Sample_Id is string to prevent errors with numeric IDs
            if 'Sample_Id' not in df.columns:
                processing_progress['error'] = f"Error processing TIS/INT samples: 'Sample_Id' column not found in the 'Extraction' sheet. Available columns: {df.columns.tolist()}"
                processing_progress['message'] = "Aborting TIS/INT processing due to missing Sample_Id column"
                return

            df['Sample_Id'] = df['Sample_Id'].astype(str)
            df = df[df['Sample_Id'] != 'nan']
                    
            processing_progress['message'] = f"Processing TIS/INT samples..."
            
            # Custom TIS sample type classification
            if hasattr(self, 'tis_samples') and self.tis_samples:
                tis_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.tis_samples]
            else:
                tis_prefixes = []
            
            # Custom INT sample type classification
            if hasattr(self, 'int_samples') and self.int_samples:
                int_prefixes = [self.prefix_mapping.get(prefix, prefix) for prefix in self.int_samples]
            else:
                int_prefixes = []
            tis_int_prefixes = tis_prefixes + int_prefixes
            # Fallback for prefixes if settings are empty
            if not tis_int_prefixes:
                tis_int_prefixes = [v for k, v in self.prefix_mapping.items() if 'TIS' in k or 'INT' in k]
                tis_int_pattern = '|'.join(tis_int_prefixes) if tis_int_prefixes else "IMPOSSIBLE_PATTERN_XYZ"
            else:
                tis_int_pattern = '|'.join(tis_int_prefixes)

            # Filter out samples that already Done
            available = df[~df['Sample_Id'].isin(done_samples) & (df['Status'] == 'in process')]
            
            # Helper for natural/smart sorting
            def natural_sort_key(s):
                import re
                return [int(text) if text.isdigit() else text.lower()
                        for text in re.split('([0-9]+)', str(s))]

            # 1. Identify explicit TIS/INT samples
            tis_int_df = available[available['Sample_Id'].str.contains(tis_int_pattern, na=False)]
            
            # Filter TIS samples
            curr_tis_ids = []
            if tis_prefixes:
                curr_tis_ids = tis_int_df[tis_int_df['Sample_Id'].str.contains('|'.join(tis_prefixes), na=False)]['Sample_Id'].tolist()
            
            # Filter INT samples
            curr_int_ids = []
            if int_prefixes:
                curr_int_ids = tis_int_df[tis_int_df['Sample_Id'].str.contains('|'.join(int_prefixes), na=False)]['Sample_Id'].tolist()
            
            # 2. Identify generic samples if category is 'tissue'
            generic_sample_ids = []
            if getattr(self, 'generic_sample_category', 'swab') == 'tissue':
                # Known prefixes include everything defined in settings + mapping values
                all_known_prefixes = list(self.prefix_mapping.values())
                for attr in ['ana_samples', 'sal_samples', 'pt_samples', 'tis_samples', 'int_samples']:
                    if hasattr(self, attr):
                        all_known_prefixes.extend([self.prefix_mapping.get(p, p) for p in getattr(self, attr)])
                
                # Deduplicate and create pattern
                all_known_prefixes = list(set(all_known_prefixes))
                known_pattern = '|'.join(all_known_prefixes) if all_known_prefixes else "IMPOSSIBLE_PATTERN_XYZ"
                
                generic_df = available[~available['Sample_Id'].str.contains(known_pattern, na=False)]
                generic_sample_ids = generic_df['Sample_Id'].tolist()
                # Filter for likely valid IDs (alphanumeric, at least 1 digit)
                generic_sample_ids = [s for s in generic_sample_ids if isinstance(s, str) and any(c.isdigit() for c in s)]

            # 3. Sorting logic
            if getattr(self, 'enable_sample_sorting', True):
                # Apply natural sorting to each group
                curr_tis_ids.sort(key=natural_sort_key)
                curr_int_ids.sort(key=natural_sort_key)
                generic_sample_ids.sort(key=natural_sort_key)
                all_sample_ids = curr_tis_ids + curr_int_ids + generic_sample_ids
            else:
                # Absolute File Order: find where these IDs appear in the original df
                all_candidate_ids = set(curr_tis_ids + curr_int_ids + generic_sample_ids)
                all_sample_ids = [sid for sid in df['Sample_Id'] if sid in all_candidate_ids]
                # Unique IDs only
                seen = set()
                all_sample_ids = [x for x in all_sample_ids if not (x in seen or seen.add(x))]

            # Remove duplicates or handle empty case
            if not all_sample_ids:
                processing_progress['message'] = "No samples matching Tissue/Intestine criteria - skipping phase"
                return

            transformed_sample_ids = [self.transform_sample_id(sid) for sid in all_sample_ids]
            
            # SAVE the final sorted transformed list for downstream use in PCR/cDNA plates
            self.final_sorted_tis_int_transformed = transformed_sample_ids
            
            # Create a combined dataframe for the rest of processing
            tis_int_samples = available[available['Sample_Id'].isin(all_sample_ids)]
            
            processing_progress['message'] = f"Success! Found TIS: {len(curr_tis_ids)}, INT: {len(curr_int_ids)}, Generic: {len(generic_sample_ids)}"
            
            # Create mapping from original to transformed sample IDs
            sample_id_mapping = {original: transformed for original, transformed in zip(all_sample_ids, transformed_sample_ids)}
            
            # Update the main dataframe
            for original, transformed in zip(all_sample_ids, transformed_sample_ids):
                df.loc[df['Sample_Id'] == original, 'Sample_Id'] = transformed
                
            os.makedirs('Extraction', exist_ok=True)
            
            # Create a copy for the check list
            tis_int_samples_copy = tis_int_samples.copy()
            
            # Update the Sample_Id column in the check list dataframe with transformed IDs
            tis_int_samples_copy['Sample_Id'] = tis_int_samples_copy['Sample_Id'].map(lambda x: sample_id_mapping.get(x, x))
            
            def get_sort_key(sample_id):
                if 'CANB_' in sample_id:
                    return (0, sample_id)
                else:
                    return (1, sample_id)
            
            tis_int_samples_copy['sort_key'] = tis_int_samples_copy['Sample_Id'].apply(get_sort_key)
            tis_int_samples_copy = tis_int_samples_copy.sort_values('sort_key')
            tis_int_samples_copy = tis_int_samples_copy.drop('sort_key', axis=1)
            
            # Track plate assignments for each sample
            tis_int_plate_assignments = {}
            
            extraction_tables = []
            num_columns = 11  # Tissue/Individual uses 11 active columns
            max_rows_per_column = 8
            items_per_plate = num_columns * max_rows_per_column
            
            # Get the starting plate number from existing Extraction_Tables.xlsx
            file_path = 'Extraction/Extraction_Tables.xlsx'
            file_is_valid = False
            
            if not os.path.exists(file_path):
                next_index = 1  # Start with Plate_001 when file doesn't exist
            else:
                # Check if file is valid (not empty and readable)
                if os.path.getsize(file_path) == 0:
                    print(f"DEBUG: File {file_path} exists but is empty (0 bytes)")
                    next_index = 1
                    # File is invalid, we will overwrite it
                else:
                    try:
                        existing_sheets = pd.ExcelFile(file_path).sheet_names
                        # Extract existing plate numbers and find the highest
                        existing_numbers = []
                        for sheet in existing_sheets:
                            if sheet.startswith('Plate_'):
                                try:
                                    num = int(sheet.replace('Plate_', ''))
                                    existing_numbers.append(num)
                                except ValueError:
                                    continue
                        next_index = max(existing_numbers) + 1 if existing_numbers else 1
                        file_is_valid = True
                    except Exception as e:
                        print(f"DEBUG: Error reading existing sheets (file likely corrupt): {e}")
                        next_index = 1
                        # File is invalid/corrupt, we will overwrite it
            
            # Create TIS/INT tables with H2O integration
            print(f"DEBUG: Creating {len(transformed_sample_ids)} TIS/INT extraction samples with H2O enabled={self.enable_h2o_random}")
            
            # Calculate how many samples can fit per plate accounting for H2O
            total_positions = num_columns * max_rows_per_column  # 88 positions (11x8)
            samples_per_plate = total_positions - self.h2o_count if self.enable_h2o_random else total_positions
            print(f"DEBUG: TIS/INT - Total positions per plate: {total_positions}, H2O count: {self.h2o_count}, Samples per plate: {samples_per_plate}")
            
            plate_index = 0
            for plate_start in range(0, len(transformed_sample_ids), samples_per_plate):
                plate_samples = transformed_sample_ids[plate_start:plate_start + samples_per_plate]
                plate_key = f"Plate_{next_index + plate_index:03d}"
                print(f"DEBUG: Creating TIS/INT Extraction table {plate_key} with {len(plate_samples)} samples")
                
                # Create table with H2O positions properly integrated (11 columns for TIS/INT)
                current_table = self.create_table_with_h2o(plate_samples, max_rows_per_column, num_columns, plate_key)
                print(f"DEBUG: Created TIS/INT table with {sum(1 for row in current_table for cell in row if cell == 'H2O')} H2O positions")
                
                # Assign plate numbers to samples (skip H2O positions)
                for row_idx, row in enumerate(current_table):
                    for col_idx, cell in enumerate(row):
                        if cell and cell != 'H2O':
                            tis_int_plate_assignments[cell] = f'Plate_{next_index + plate_index:03d}'
                
                extraction_tables.append(current_table)
                plate_index += 1
            
            # Add plate information to check list
            tis_int_samples_copy['Plate'] = tis_int_samples_copy['Sample_Id'].map(tis_int_plate_assignments)
            
            # Reorder columns to include Plate
            if tis_int_samples_copy.empty:
                print(f"DEBUG: TIS/INT dataframe is empty, skipping column reordering")
                self.tis_int_check_list = pd.DataFrame(columns=['Sample_Id', 'Pool', 'Plate', 'Status'])
                return
            
            columns = tis_int_samples_copy.columns.tolist()
            # Check if columns exist before trying to remove them
            for col in ['Status', 'Plate']:
                if col in columns:
                    columns.remove(col)
                    print(f"DEBUG: Removed column {col}")
                else:
                    print(f"DEBUG: Column {col} not found in dataframe")
            columns.extend(['Plate', 'Status'])
            tis_int_samples_copy = tis_int_samples_copy[columns]
            
            # Add logging to indicate what's being saved
            processing_progress['message'] = f"Writing {len(tis_int_samples_copy)} samples to Check_list_Tissue-Intestine.xlsx"
            
            # Debug output first few rows of data to verify sample IDs
            if not tis_int_samples_copy.empty:
                sample_examples = tis_int_samples_copy['Sample_Id'].head(5).tolist()
                processing_progress['message'] = f"Sample ID examples: {sample_examples}"
                
            # Create check list with ORIGINAL sample IDs (not transformed)
            original_tis_int_check_list = tis_int_samples.copy()
            # Create reverse mapping from transformed back to original
            reverse_sample_id_mapping = {transformed: original for original, transformed in zip(all_sample_ids, transformed_sample_ids)}
            
            # Map plate assignments from transformed IDs back to original IDs
            original_plate_assignments = {}
            for transformed_id, plate in tis_int_plate_assignments.items():
                original_id = reverse_sample_id_mapping.get(transformed_id, transformed_id)
                original_plate_assignments[original_id] = plate
            
            # Assign plate to original check list using original sample IDs
            original_tis_int_check_list['Plate'] = original_tis_int_check_list['Sample_Id'].map(original_plate_assignments)
            original_tis_int_check_list['Status'] = 'in process'
            
            # Reorder columns for original check list
            orig_columns = original_tis_int_check_list.columns.tolist()
            for col in ['Plate', 'Status']:
                if col in orig_columns:
                    orig_columns.remove(col)
            orig_columns.extend(['Plate', 'Status'])
            original_tis_int_check_list = original_tis_int_check_list[orig_columns]
            
            # Store original check list data for later combination
            self.tis_int_check_list = original_tis_int_check_list
            
            # Use the existing file_path and next_index from above
            # Only append if the file exists AND is valid (readable Excel file)
            if os.path.exists(file_path) and file_is_valid:
                mode = 'a'
                if_sheet_exists = 'overlay'
            else:
                print(f"DEBUG: Creating new/overwriting Extraction_Tables.xlsx (Valid: {file_is_valid})")
                mode = 'w'
                if_sheet_exists = None
                
            with pd.ExcelWriter(file_path, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
                for table_index, table in enumerate(extraction_tables):
                    extraction_df = pd.DataFrame(table)
                    table_sample_ids = [sid for row in table for sid in row if sid is not None and sid != 'H2O']
                    sample_extraction_id = self.get_sample_extraction_id(table_sample_ids)
                    
                    # Snapshot this plate for 100% synchronization with cDNA/PCR
                    plate_number = next_index + table_index
                    self.extraction_plate_snapshots.append({
                        'plate_number': plate_number,
                        'range_header': sample_extraction_id,
                        'table': table,
                        'type': 'tissue',
                        'pool_assignments': {} # Tissue doesn't use pooling
                    })
                    
                    sheet_name = f'Plate_{plate_number:03d}'
                    workbook = writer.book
                    worksheet = workbook.create_sheet(sheet_name) if sheet_name not in workbook.sheetnames else workbook[sheet_name]
                    
                    # Define formatting styles
                    sample_id_font = Font(name='Arial', size=14, bold=False)
                    col_row_font = Font(name='Arial', size=22, bold=True)
                    header_font = Font(name='Arial', size=14, bold=True)
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    fills = [
                        PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
                        PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),  # Sky Blue
                        PatternFill(start_color="808080", end_color="808080", fill_type="solid")   # Gray
                    ]
                    
                    # Write three tables vertically stacked
                    for variant in range(3):
                        row_offset = variant * 18  # 18 rows between each table (including spacing)
                        
                        # Headers and metadata with font size 14
                        worksheet[f'I{3 + row_offset}'] = "Date:"
                        worksheet[f'I{3 + row_offset}'].font = header_font
                        worksheet[f'D{3 + row_offset}'] = "Inactivated by:"
                        worksheet[f'D{3 + row_offset}'].font = header_font
                        worksheet[f'C{1 + row_offset}'] = "Protocol: "
                        worksheet[f'C{1 + row_offset}'].font = header_font
                        worksheet[f'F{1 + row_offset}'] = "Kitt#: "
                        worksheet[f'F{1 + row_offset}'].font = header_font
                        worksheet[f'I{1 + row_offset}'] = "Lot#:"
                        worksheet[f'I{1 + row_offset}'].font = header_font
                        worksheet[f'L{1 + row_offset}'] = "Expiry:"
                        worksheet[f'L{1 + row_offset}'].font = header_font
                        worksheet[f'C{6 + row_offset}'] = "Sample Id"
                        worksheet[f'C{6 + row_offset}'].font = header_font
                        worksheet[f'E{6 + row_offset}'] = sample_extraction_id
                        worksheet[f'E{6 + row_offset}'].font = header_font
                        worksheet[f'I{5 + row_offset}'] = "Date:"
                        worksheet[f'I{5 + row_offset}'].font = header_font
                        worksheet[f'D{5 + row_offset}'] = "Extracted by:"
                        worksheet[f'D{5 + row_offset}'].font = header_font
                        
                        # Column headers
                        for col in range(num_columns):
                            cell = worksheet.cell(row=7 + row_offset, column=col + 2)
                            if col < 11:
                                cell.value = str(col + 1)
                            else:
                                cell.value = "" # Leave column 12 blank for manual controls
                            cell.font = col_row_font
                            if fills[variant]:
                                cell.fill = fills[variant]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Row headers
                        for row in range(max_rows_per_column):
                            cell = worksheet.cell(row=row + 8 + row_offset, column=1)
                            cell.value = chr(65 + row)
                            cell.font = col_row_font
                            if fills[variant]:
                                cell.fill = fills[variant]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Sample IDs
                        for row in range(len(extraction_df)):
                            for col in range(len(extraction_df.columns)):
                                sample_id = extraction_df.iloc[row, col]
                                if sample_id:
                                    cell = worksheet.cell(row=row + 8 + row_offset, column=col + 2)
                                    cell.value = self.split_sample_id(sample_id) if sample_id != 'H2O' else 'H2O'
                                    if sample_id == 'H2O':
                                        cell.font = Font(name='Arial', size=14, bold=True, color="FF0000")  # Red bold for H2O
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red background
                                    else:
                                        cell.font = sample_id_font
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable wrap_text for two lines
                                        self.format_cell_for_sample_id(cell, worksheet)
                                    cell.border = thin_border
                                    # ALWAYS EXPLICITLY BLANK COLUMN 12 (Excel M/13) to clear old data
                                    worksheet.cell(row=row + 8 + row_offset, column=13).value = ""
                        
                        # Plate number (Shifted to N to reserve M/12)
                        worksheet[f'N{8 + row_offset}'] = str(plate_number)
                        worksheet.merge_cells(f'N{8 + row_offset}:N{15 + row_offset}')
                        merged_cell = worksheet[f'N{8 + row_offset}']
                        merged_cell.font = Font(name='Arial', size=22, bold=True)
                        if fills[variant]:
                            merged_cell.fill = fills[variant]
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        merged_cell.border = Border(left=Side(style='thin'))
                        
                        # Set dimensions
                        for row in range(8 + row_offset, 16 + row_offset):
                            worksheet.row_dimensions[row].height = 35
                        for col in range(2, num_columns + 2):
                            worksheet.column_dimensions[get_column_letter(col)].width = 10
                    self.pre_save_formatting(worksheet)
            
            processing_progress['message'] = f"Created {len(extraction_tables)} Tissue sample tables (3 tables per sheet)"
            
        except Exception as e:
            processing_progress['error'] = f'Error processing TIS/INT samples: {str(e)}'

    def create_combined_check_list(self, swab_check_list, tis_int_check_list):
        """Create a single combined check list with all samples"""
        try:
            if swab_check_list is None and tis_int_check_list is None:
                processing_progress['message'] = 'No samples to create check list'
                return
            
            # Combine both check lists
            combined_check_list = pd.DataFrame()
            
            if swab_check_list is not None and not swab_check_list.empty:
                # Ensure Pool column exists for swab samples
                if 'Pool' not in swab_check_list.columns:
                    swab_check_list['Pool'] = 'Not Pool'
                combined_check_list = pd.concat([combined_check_list, swab_check_list], ignore_index=True)
            
            if tis_int_check_list is not None and not tis_int_check_list.empty:
                # Add Pool column for tissue samples (they don't have pools)
                if 'Pool' not in tis_int_check_list.columns:
                    tis_int_check_list['Pool'] = 'Not Pool'
                combined_check_list = pd.concat([combined_check_list, tis_int_check_list], ignore_index=True)
            
            if combined_check_list.empty:
                processing_progress['message'] = 'No valid samples for check list'
                return
            
            # Sort by sample type and ID for better organization
            def get_sort_priority(sample_id):
                if isinstance(sample_id, str):
                    if 'CANB_' in sample_id:
                        return (0, sample_id)
                    elif 'CANR_' in sample_id:
                        return (1, sample_id)
                    elif 'IPLNAHL_' in sample_id:
                        return (2, sample_id)
                    elif 'CANA_' in sample_id:
                        return (3, sample_id)
                    else:
                        return (4, sample_id)
                return (5, str(sample_id))
            
            combined_check_list['sort_priority'] = combined_check_list['Sample_Id'].apply(get_sort_priority)
            combined_check_list = combined_check_list.sort_values('sort_priority')
            combined_check_list = combined_check_list.drop('sort_priority', axis=1)
            
            # Ensure these columns exist in the dataframe
            for col in ['Pool', 'Plate', 'Status']:
                if col not in combined_check_list.columns:
                    combined_check_list[col] = '' if col != 'Pool' else 'Not Pool'
            
            # Reorder columns: keep all existing columns but move/add Pool, Plate, Status to end
            existing_cols = [c for c in combined_check_list.columns if c not in ['Pool', 'Plate', 'Status']]
            columns = existing_cols + ['Pool', 'Plate', 'Status']
            
            combined_check_list = combined_check_list[columns]
            
            # Save the combined check list
            combined_check_list.to_excel('Extraction/Check_list_All_Samples.xlsx', sheet_name='Check_List', index=False)
            processing_progress['message'] = f"Created combined check list with {len(combined_check_list)} samples"
            
        except Exception as e:
            processing_progress['error'] = f'Error creating combined check list: {str(e)}'

    def create_pcr_cdna_plates(self, destination_file, new_sample_ids):
        """Create PCR and cDNA plates based on captured Extraction plate snapshots"""
        try:
            if not hasattr(self, 'extraction_plate_snapshots') or not self.extraction_plate_snapshots:
                processing_progress['message'] = "No extraction plates were generated for this run. Skipping PCR/cDNA."
                return

            snapshots = self.extraction_plate_snapshots
            processing_progress['message'] = f"Synchronizing PCR/cDNA tables for {len(snapshots)} plates..."
            
            # Create continuous plates for both PCR and cDNA
            self.create_snapshot_plates(snapshots, 'PCR', 'PCR_Tables')
            self.create_snapshot_plates(snapshots, 'cDNA', 'cDNA_Tables')
            
        except Exception as e:
            processing_progress['error'] = f'Error creating PCR/cDNA plates: {str(e)}'

    def create_snapshot_plates(self, snapshots, plate_type, base_filename):
        """Redistribute snapshot columns into 11-column plates for PCR/cDNA with continuous packing"""
        try:
            os.makedirs(f'Extraction/{plate_type}', exist_ok=True)
            file_path = f'Extraction/{plate_type}/{base_filename}.xlsx'
            
            next_index = 1
            if os.path.exists(file_path):
                try:
                    next_index = len(pd.ExcelFile(file_path).sheet_names) + 1
                    mode = 'a'
                    if_sheet_exists = 'overlay'
                except:
                    mode = 'w'
                    if_sheet_exists = None
            else:
                mode = 'w'
                if_sheet_exists = None

            # 1. Group snapshots by type to avoid mixing
            swab_snapshots = [s for s in snapshots if s.get('type') == 'swab']
            tissue_snapshots = [s for s in snapshots if s.get('type') != 'swab']

            with pd.ExcelWriter(file_path, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
                # --- PROCESS SWAB SNAPSHOTS (Continuous Packing) ---
                if swab_snapshots:
                    # 1. Gather all unique Pool IDs and track their Samples + Source Extraction Plates
                    all_swab_pools = []
                    pool_to_samples = {}
                    pool_to_ext_plate = {} # Map Pool_ID -> Extraction Plate Number
                    
                    for snap in swab_snapshots:
                        table = snap['table']
                        mapping = snap.get('pool_assignments', {})
                        ext_plate_num = snap.get('plate_number', 0)
                        num_cols = len(table[0])
                        for c in range(num_cols):
                            col_samples = [table[r][c] for r in range(8) if table[r][c] and table[r][c] != 'H2O']
                            if col_samples:
                                sample_id = col_samples[0]
                                pool_name = mapping.get(sample_id, sample_id)
                                
                                if pool_name not in all_swab_pools:
                                    all_swab_pools.append(pool_name)
                                    pool_to_samples[pool_name] = []
                                    pool_to_ext_plate[pool_name] = ext_plate_num
                                
                                for s in col_samples:
                                    if s not in pool_to_samples[pool_name]:
                                        pool_to_samples[pool_name].append(s)
                    
                    pool_idx = 0
                    items_per_plate = 88 - self.h2o_count if self.enable_h2o_random else 88
                    num_plates = (len(all_swab_pools) + items_per_plate - 1) // items_per_plate
                    
                    for p in range(num_plates):
                        sheet_name = f'Plate_{next_index:03d}'
                        plate_pool_count = min(len(all_swab_pools) - pool_idx, items_per_plate)
                        h2o_positions = self.generate_h2o_positions(8, 11, f"Redistributed_PCR_{next_index}", sample_count=plate_pool_count)
                        h2o_set = set(h2o_positions)
                        
                        modified_table = [['' for _ in range(12)] for _ in range(8)]
                        current_plate_samples = []
                        current_plate_ext_plates = set() # Track source extraction plates for this PCR plate
                        
                        # Fill well-by-well
                        for c in range(11):
                            for r in range(8):
                                if (r, c) in h2o_set:
                                    modified_table[r][c] = 'H2O'
                                elif pool_idx < len(all_swab_pools):
                                    pool_name = all_swab_pools[pool_idx]
                                    modified_table[r][c] = str(pool_name).replace('_', '\n')
                                    if pool_name in pool_to_samples:
                                        current_plate_samples.extend(pool_to_samples[pool_name])
                                    if pool_name in pool_to_ext_plate:
                                        current_plate_ext_plates.add(pool_to_ext_plate[pool_name])
                                    pool_idx += 1
                        
                        df_plate = pd.DataFrame(modified_table, columns=[i+1 for i in range(12)], index=[chr(65+i) for i in range(8)])
                        range_header = self.get_sample_extraction_id(current_plate_samples)
                        
                        # Format extraction plate string (e.g., "1 - 3")
                        sorted_ext_plates = sorted(list(current_plate_ext_plates))
                        if not sorted_ext_plates:
                            ext_plate_str = "Unknown"
                        elif len(sorted_ext_plates) == 1:
                            ext_plate_str = str(sorted_ext_plates[0])
                        else:
                            ext_plate_str = f"{sorted_ext_plates[0]} - {sorted_ext_plates[-1]}"
                        
                        if plate_type == 'PCR':
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=18)
                            worksheet = writer.sheets[sheet_name]
                            self.format_plate_worksheet(worksheet, 8, 11, start_row=0, sample_range=range_header, plate_number=ext_plate_str, table_type='PCR')
                            self.format_plate_worksheet(worksheet, 8, 11, start_row=14, sample_range=range_header, plate_number=ext_plate_str, table_type='PCR')
                        else:
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            worksheet = writer.sheets[sheet_name]
                            self.format_plate_worksheet(worksheet, 8, 11, start_row=0, sample_range=range_header, plate_number=ext_plate_str, table_type='cDNA')
                        
                        next_index += 1

                # --- PROCESS TISSUE SNAPSHOTS (1:1 Mirrored Name Matching) ---
                for snap in tissue_snapshots:
                    table = snap['table']
                    # Use extraction plate number directly for the worksheet name
                    ext_plate_num = snap.get('plate_number', '?')
                    rows = 8
                    cols = 11
                    modified_table = []
                    all_samples = []
                    for r in range(rows):
                        new_row = []
                        for c in range(11):
                            cell = table[r][c]
                            if cell == 'H2O':
                                new_row.append('H2O')
                            elif cell and cell != '':
                                val = str(cell).replace('_', '\n')
                                new_row.append(val)
                                all_samples.append(cell)
                            else:
                                new_row.append('')
                        new_row.append('') # Col 12 blank
                        modified_table.append(new_row)
                    
                    df_plate = pd.DataFrame(modified_table, columns=[i+1 for i in range(12)], index=[chr(65+i) for i in range(rows)])
                    
                    # MATCH SHEET NAME TO EXTRACTION PLATE NUMBER
                    sheet_name = f'Plate_{int(ext_plate_num) if str(ext_plate_num).isdigit() else ext_plate_num:03d}'
                    range_header = self.get_sample_extraction_id(all_samples)
                    ext_plate_str = str(ext_plate_num)
                    
                    if plate_type == 'PCR':
                        df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                        df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=18)
                        worksheet = writer.sheets[sheet_name]
                        self.format_plate_worksheet(worksheet, 8, 11, start_row=0, sample_range=range_header, plate_number=ext_plate_str, table_type='PCR')
                        self.format_plate_worksheet(worksheet, 8, 11, start_row=14, sample_range=range_header, plate_number=ext_plate_str, table_type='PCR')
                    else:
                        df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                        worksheet = writer.sheets[sheet_name]
                        self.format_plate_worksheet(worksheet, 8, 11, start_row=0, sample_range=range_header, plate_number=ext_plate_str, table_type='cDNA')

            print(f"DEBUG: Successfully created redistributed {plate_type} plates.")
        except Exception as e:
            print(f"ERROR: Failed to create {plate_type} plates: {e}")
            raise

    def create_continuous_plates(self, pools, pool_to_samples, tis_int_samples, plate_type, base_filename):
        """Create continuous plates with pools first, then Tissue/Intestine samples"""
        try:
            os.makedirs(f'Extraction/{plate_type}', exist_ok=True)
            
            max_rows = 8
            num_columns = 12
            # Account for H2O positions when calculating items per plate
            total_positions = max_rows * num_columns  # 96 positions
            items_per_plate = total_positions - self.h2o_count if self.enable_h2o_random else total_positions
            print(f"DEBUG: PCR/cDNA items_per_plate = {items_per_plate} (H2O enabled: {self.enable_h2o_random}, count: {self.h2o_count})")
            
            file_path = f'Extraction/{plate_type}/{base_filename}.xlsx'
            next_index = 1 if not os.path.exists(file_path) else len(pd.ExcelFile(file_path).sheet_names) + 1
            
            if os.path.exists(file_path):
                mode = 'a'
                if_sheet_exists = 'overlay'
            else:
                mode = 'w'
                if_sheet_exists = None
                
            # Only create if we have data to process
            if not pools and not tis_int_samples:
                processing_progress['message'] = f'No data available for {plate_type} plates'
                return
            
            with pd.ExcelWriter(file_path, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
                # Process pools
                pool_items = pools.copy()
                if pool_items:
                    processing_progress['message'] = f'Creating {plate_type} plates for {len(pool_items)} Swab pools'
                    
                    for plate_start in range(0, len(pool_items), items_per_plate):
                        plate_samples = pool_items[plate_start:plate_start+items_per_plate]
                        
                        # Use the same plate key as Extraction for consistency
                        plate_key = f"Extraction_Plate_{next_index}"
                        
                        # Create table with H2O positions properly integrated (11 columns for PCR/cDNA)
                        modified_table = self.create_table_with_h2o(plate_samples, max_rows, num_columns, plate_key, format_for_pcr_cdna=True)
                        
                        # Add empty column M
                        for row in modified_table:
                            row.append('')
                        
                        # Create DataFrame
                        df_plate = pd.DataFrame(modified_table, 
                                            columns=[i+1 for i in range(12)], 
                                            index=[chr(65+i) for i in range(max_rows)])
                        
                        sheet_name = f'Plate_{next_index:03d}'
                        
                        # Get sample range for headers
                        plate_sample_ids = [item for item in plate_samples if item != '']
                        if any(isinstance(sid, str) and sid.startswith('Pool_') for sid in plate_sample_ids):
                            original_pool_samples = []
                            for item in plate_sample_ids:
                                if isinstance(item, str) and item.startswith('Pool_'):
                                    if item in pool_to_samples:
                                        original_pool_samples.extend(pool_to_samples[item])
                            sample_range = self.get_sample_range(original_pool_samples, use_original=False, force_original=True)
                        else:
                            sample_range = self.get_sample_range(plate_sample_ids)
                        
                        if plate_type == 'PCR':
                            # First PCR table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            
                            # Second PCR table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=18)
                            
                            worksheet = writer.sheets[sheet_name]
                            
                            # Don't clear M5 and M19 - keep the "12" value but will be reformatted
                            
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=0, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='PCR')
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=14, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='PCR')
                        else:
                            # Single cDNA table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            worksheet = writer.sheets[sheet_name]
                            
                            # Don't clear M5 - keep the "12" value but will be reformatted
                            
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=0, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='cDNA')
                        
                        next_index += 1
                
                # Process Tissue/Intestine samples
                if tis_int_samples:
                    processing_progress['message'] = f'Creating {plate_type} plates for {len(tis_int_samples)} Tissue/Intestine samples'
                    
                    # Start Tissue/Intestine plates after pool plates to avoid key conflicts
                    tis_int_start_index = next_index
                    
                    for plate_start in range(0, len(tis_int_samples), items_per_plate):
                        plate_samples = tis_int_samples[plate_start:plate_start+items_per_plate]
                        
                        # Use unique plate key for Tissue/Intestine plates
                        tis_int_plate_num = tis_int_start_index + (plate_start // items_per_plate)
                        plate_key = f"Extraction_Plate_{tis_int_plate_num}_TISINT"
                        
                        # Create table with H2O positions properly integrated (11 columns for PCR/cDNA)
                        modified_table = self.create_table_with_h2o(plate_samples, max_rows, num_columns, plate_key, format_for_pcr_cdna=True)
                        
                        # Add empty column M
                        for row in modified_table:
                            row.append('')
                        
                        # Create DataFrame
                        df_plate = pd.DataFrame(modified_table, 
                                            columns=[i+1 for i in range(12)], 
                                            index=[chr(65+i) for i in range(max_rows)])
                        
                        sheet_name = f'Plate_{tis_int_plate_num}'
                        
                        # Get sample range
                        plate_sample_ids = [item for item in plate_samples if item != '']
                        sample_range = self.get_sample_range(plate_sample_ids)
                        
                        if plate_type == 'PCR':
                            # First PCR table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            
                            # Second PCR table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=18)
                            
                            worksheet = writer.sheets[sheet_name]
                            
                            # Don't clear M5 and M19 - keep the "12" value but will be reformatted
                            
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=0, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='PCR')
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=14, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='PCR')
                        else:
                            # Single cDNA table
                            df_plate.to_excel(writer, sheet_name=sheet_name, index=True, index_label='', startrow=4)
                            worksheet = writer.sheets[sheet_name]
                            
                            # Don't clear M5 - keep the "12" value but will be reformatted
                            
                            self.format_plate_worksheet(worksheet, max_rows, num_columns, start_row=0, 
                                                    sample_range=sample_range, plate_number=sheet_name, table_type='cDNA')
                        
                        next_index += 1
            
            processing_progress['message'] = f'Created continuous {plate_type} plates'
            
        except Exception as e:
            processing_progress['error'] = f'Error creating continuous {plate_type} plates: {str(e)}'
            
    def create_dynamic_pools(self, samples, h2o_positions, max_rows=8, num_columns=12):
        """Create pools dynamically based on H2O positions in each column"""
        if not self.samples_per_pool:
            return []
        
        pools = []
        pool_count = 1
        
        # Group samples by column capacity
        column_capacities = []
        for col in range(num_columns):
            # Check if this column has H2O
            has_h2o = any(pos[1] == col for pos in h2o_positions)
            capacity = max_rows - 1 if has_h2o else max_rows  # 7 if H2O, 8 if no H2O
            column_capacities.append(capacity)
        
        print(f"DEBUG: Column capacities: {column_capacities}")
        
        # Distribute samples across columns based on capacity
        remaining_samples = samples.copy()
        current_pool_samples = []
        
        for col, capacity in enumerate(column_capacities):
            # Take samples for this column based on its capacity
            samples_for_column = remaining_samples[:capacity]
            remaining_samples = remaining_samples[capacity:]
            
            # Add these samples to current pool
            current_pool_samples.extend(samples_for_column)
            
            # Check if we have enough for a complete pool
            if len(current_pool_samples) >= self.samples_per_pool:
                # Create pools from accumulated samples
                for i in range(0, len(current_pool_samples), self.samples_per_pool):
                    pool_samples = current_pool_samples[i:i + self.samples_per_pool]
                    if len(pool_samples) == self.samples_per_pool:
                        pool_name = f"Pool_{pool_count:03}"
                        pools.append({
                            'name': pool_name,
                            'samples': pool_samples,
                            'column': col
                        })
                        pool_count += 1
                
                # Keep any remaining samples for next columns
                current_pool_samples = current_pool_samples[len(current_pool_samples) % self.samples_per_pool:]
        
        print(f"DEBUG: Created {len(pools)} dynamic pools")
        return pools

    def get_sample_range(self, sample_ids, use_original=False, force_original=False):
        """Get sample range string grouped by prefix type"""
        print(f"DEBUG: get_sample_range called with {len(sample_ids)} samples, use_original={use_original}")
        print(f"DEBUG: Sample IDs received: {sample_ids[:10] if len(sample_ids) > 10 else sample_ids}...")  # Show first 10
        
        if not sample_ids:
            return "No samples"
        
        # Check if these are pool samples (start with "Pool_")
        is_pool_samples = any(isinstance(sid, str) and sid.startswith('Pool_') for sid in sample_ids)
        print(f"DEBUG: is_pool_samples = {is_pool_samples}")
        
        if is_pool_samples:
            # Extract pool numbers and map to sample IDs using prefix mapping
            pool_numbers = []
            for sid in sample_ids:
                if isinstance(sid, str) and sid.startswith('Pool_'):
                    try:
                        pool_num = sid.split('_')[1]
                        pool_numbers.append(int(pool_num))
                    except (IndexError, ValueError):
                        continue
            
            if not pool_numbers:
                return "No samples"
            
            pool_numbers.sort()
            
            # Find the prefix to use
            if force_original or use_original:
                # Use original prefix for check lists
                pool_prefix = None
                for old_prefix, new_prefix in self.prefix_mapping.items():
                    if any(old_prefix in ps for ps in self.pool_samples):
                        pool_prefix = old_prefix
                        break
            else:
                # Use new prefix for plates
                pool_prefix = None
                for old_prefix, new_prefix in self.prefix_mapping.items():
                    if any(old_prefix in ps for ps in self.pool_samples):
                        pool_prefix = new_prefix
                        break
            
            if not pool_prefix:
                # Use the first configured pool sample prefix from settings
                if self.pool_samples:
                    first_pool_prefix = self.pool_samples[0]
                    if force_original or use_original:
                        pool_prefix = first_pool_prefix
                    else:
                        pool_prefix = self.prefix_mapping.get(first_pool_prefix, first_pool_prefix)
                else:
                    pool_prefix = "UNKNOWN"  # Last resort if no pool samples configured
            
            # Create sample range with proper formatting
            if len(pool_numbers) == 1:
                return f"Sample Id: {pool_prefix}_{pool_numbers[0]:03d}"
            else:
                return f"Sample Id: {pool_prefix}_{pool_numbers[0]:03d} to {pool_prefix}_{pool_numbers[-1]:03d}"
        else:
            # Regular sample IDs (non-pool) - group by prefix
            clean_ids = []
            for sid in sample_ids:
                if isinstance(sid, str):
                    clean_id = sid.replace('\n', '_')
                    clean_ids.append(clean_id)
            
            if not clean_ids:
                return "No samples"
            
            # Group by prefix
            prefix_groups = {}
            for sid in clean_ids:
                # Extract prefix (everything before the last underscore and number)
                if '_' in sid:
                    parts = sid.split('_')
                    if len(parts) >= 3:
                        prefix = '_'.join(parts[:-1])  # Everything except the number
                        number = int(parts[-1])
                        if prefix not in prefix_groups:
                            prefix_groups[prefix] = []
                        prefix_groups[prefix].append(number)
            
            if not prefix_groups:
                # Fallback to original behavior if no prefix grouping
                clean_ids.sort()
                if len(clean_ids) == 1:
                    return f"Sample Id: {clean_ids[0]}"
                else:
                    return f"Sample Id: {clean_ids[0]} to {clean_ids[-1]}"
            
            # Create ranges for each prefix group
            ranges = []
            for prefix in sorted(prefix_groups.keys()):
                numbers = sorted(prefix_groups[prefix])
                if len(numbers) == 1:
                    ranges.append(f"{prefix}_{numbers[0]:03d}")
                else:
                    ranges.append(f"{prefix}_{numbers[0]:03d} to {numbers[-1]:03d}")
            
            # Format as type groups with concise numbering
            if len(ranges) == 1:
                return f"Sample Id: {ranges[0]}"
            elif len(ranges) == 2:
                # Extract prefixes and number ranges
                range1_parts = ranges[0].split('_')
                range2_parts = ranges[1].split('_')
                
                # Reconstruct with concise format
                if len(range1_parts) >= 3 and len(range2_parts) >= 3:
                    prefix1 = '_'.join(range1_parts[:-1])
                    prefix2 = '_'.join(range2_parts[:-1])
                    return f"Sample Id: {prefix1}_{range1_parts[-1]} and {prefix2}_{range2_parts[-1]}"
            
            return f"Sample Id: {', '.join(ranges)}"

    def format_plate_worksheet(self, worksheet, max_rows, num_columns, start_row=0, sample_range=None, plate_number=None, table_type='PCR'):
        """Format a plate worksheet with borders and fonts"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_font = Font(name='Arial', size=12, bold=False)
        cell_font = Font(name='Arial', size=14, bold=False)
        bold_font = Font(name='Arial', size=16, bold=True)
        row_col_font = Font(name='Arial', size=22, bold=True)
        
        # Determine header positions based on table type and start_row
        if table_type == 'PCR':
            if start_row == 0:
                # First PCR table
                header_rows = {
                    'perform_by': 2,
                    'date': 2,
                    'extraction_plate': 3,
                    'plate_number': 3,
                    'cdna_date': 3,
                    'sample_id': 4,
                    'data_start': 5
                }
            else:
                # Second PCR table
                header_rows = {
                    'perform_by': 16,
                    'date': 16,
                    'extraction_plate': 17,
                    'plate_number': 17,
                    'cdna_date': 17,
                    'sample_id': 18,
                    'data_start': 19
                }
        else:  # cDNA
            header_rows = {
                'perform_by': 2,
                'date': 2,
                'extraction_plate': 3,
                'plate_number': 3,
                'cdna_date': 3,
                'sample_id': 4,
                'data_start': 5
            }
        
        # Add header labels
        # Row with "Perform by:" and "Date:"
        perform_by_cell = worksheet.cell(row=header_rows['perform_by'], column=2)  # B
        perform_by_cell.value = "Perform by:"
        perform_by_cell.font = header_font
        perform_by_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        date_cell = worksheet.cell(row=header_rows['date'], column=9)  # I
        if table_type == 'PCR':
            date_cell.value = "PCR Date:"
        else:  # cDNA
            date_cell.value = "cDNA Date:"
        date_cell.font = header_font
        date_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Row with "Extraction Plate:", plate number, and date label
        extraction_label_cell = worksheet.cell(row=header_rows['extraction_plate'], column=2)  # B
        extraction_label_cell.value = "Extraction Plate:"
        extraction_label_cell.font = header_font
        extraction_label_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        plate_num_cell = worksheet.cell(row=header_rows['plate_number'], column=4)  # D
        plate_num_cell.value = plate_number if plate_number else ""
        plate_num_cell.font = header_font
        plate_num_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Only add cDNA Date for PCR tables (as requested to remove from cDNA)
        if table_type == 'PCR':
            cdna_date_cell = worksheet.cell(row=header_rows['cdna_date'], column=9)  # I
            cdna_date_cell.value = "cDNA Date:"
            cdna_date_cell.font = header_font
            cdna_date_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Row with Sample ID range
        if sample_range:
            sample_id_cell = worksheet.cell(row=header_rows['sample_id'], column=2)  # B
            sample_id_cell.value = sample_range
            sample_id_cell.font = bold_font
            sample_id_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Format the data table (column and row headers + data cells)
        data_start_row = header_rows['data_start']
        
        # Column headers (1-11) plus column M (12)
        for col in range(num_columns + 1):  # +1 to include column M
            cell = worksheet.cell(row=data_start_row, column=col + 2)  # Start at column B
            if col < 11:
                cell.value = str(col + 1)  # Numbers 1-11
            else:
                # Column 12 or M - leave blank as requested
                cell.value = ""
            cell.font = row_col_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            # EXPLICITLY BLANK COLUMN 12 (M/13) header
            worksheet.cell(row=data_start_row, column=13).value = ""
        
        # Row headers (A-H) and data cells
        for row in range(max_rows):
            # Row header
            row_header_cell = worksheet.cell(row=data_start_row + row + 1, column=1)  # Column A
            row_header_cell.value = chr(65 + row)  # A, B, C, etc.
            row_header_cell.font = row_col_font
            row_header_cell.alignment = Alignment(horizontal='center', vertical='center')
            row_header_cell.border = thin_border
            
            # Data cells (columns B-L) and empty column M
            for col in range(num_columns + 1):  # +1 for column M (empty)
                cell = worksheet.cell(row=data_start_row + row + 1, column=col + 2)
                if cell.value is None:
                    cell.value = ''
                cell.border = thin_border
                
                # Special formatting for column M (column 13) - last row of each table
                if col == num_columns and row == max_rows - 1:  # Column M, last row (H)
                    cell.font = row_col_font  # Size 22, bold
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    # Check if this is an H2O cell and apply special formatting
                    if cell.value == 'H2O':
                        cell.font = Font(name='Arial', size=14, bold=True, color="FF0000")  # Red bold for H2O
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red background
                    else:
                        cell.font = cell_font
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    # EXPLICITLY BLANK COLUMN 12 (M/13) data
                    worksheet.cell(row=data_start_row + row + 1, column=13).value = ""
                
                # Set row height for wrapped text
                worksheet.row_dimensions[data_start_row + row + 1].height = 35
        
        # Set column widths
        for col in range(1, num_columns + 3):  # A through M
            worksheet.column_dimensions[get_column_letter(col)].width = 10

    def process_files(self, source_file, destination_file, sheets_info):
        global processing_progress
        try:
            processing_progress['status'] = 'processing'
            processing_progress['progress'] = 0
            processing_progress['message'] = 'Starting processing...'
            processing_progress['error'] = ''
            
            # CRITICAL: Clear snapshots at the beginning of every run to avoid data leakage
            self.extraction_plate_snapshots = []

            output_data = pd.DataFrame(columns=['Host_Id', 'Sample_Id', 'Status'])
            re_enable_samples = []
            existing_columns = set()
            
            # Load existing Done samples from output file (destination_file) as it is the source of truth
            # This ensures we don't re-process samples that are already marked as Done
            self.previous_done_samples = set()
            if os.path.exists(destination_file):
                try:
                    # Using openpyxl explicitly for reliability
                    df_check = pd.read_excel(destination_file, sheet_name='Extraction', engine='openpyxl')
                    if 'Status' in df_check.columns:
                        done_rows = df_check[df_check['Status'] == 'Done']
                        self.previous_done_samples = set(done_rows['Sample_Id'].astype(str).tolist())
                        print(f"DEBUG: Pre-loaded {len(self.previous_done_samples)} Done samples from output file")
                except Exception as e:
                    print(f"DEBUG: Could not pre-load Done samples from output file: {e}")
            
            if os.path.exists(destination_file):
                try:
                    existing_data = pd.read_excel(destination_file, sheet_name="Extraction")
                    if 'Status' not in existing_data.columns:
                        existing_data['Status'] = 'unknown'
                    existing_sample_ids = existing_data['Sample_Id'].dropna().tolist()
                    existing_columns = set(self.transform_sample_id(sid) for sid in existing_sample_ids)
                    processing_progress['message'] = f'Loaded {len(existing_columns)} existing Sample_Ids'
                except Exception as e:
                    processing_progress['message'] = f'Warning: Could not load existing data: {str(e)}'

            total_steps = len(sheets_info) + 6
            current_step = 0

            # Process sheets
            for sheet_name, columns in sheets_info.items():
                current_step += 1
                processing_progress['progress'] = int((current_step / total_steps) * 100)
                processing_progress['message'] = f'Processing {sheet_name}...'
                
                try:
                    source_data = pd.read_excel(source_file, sheet_name=sheet_name)
                    if source_data.empty:
                        continue
                    selected_columns = source_data.iloc[:, columns]
                    
                    for index, row in selected_columns.iterrows():
                        host_value = row.iloc[0]
                        for i in range(1, len(row)):
                            col_value = row.iloc[i]
                            
                            # Filter by custom sample type settings and exclude already Done samples
                            transformed_col_value = self.transform_sample_id(col_value)
                            if pd.notna(col_value) and transformed_col_value not in existing_columns and transformed_col_value not in self.previous_done_samples:
                                # Check if this sample matches any of your custom sample type settings
                                sample_matches_custom_type = False
                                
                                # Check ANA samples
                                if hasattr(self, 'ana_samples') and self.ana_samples:
                                    for ana_prefix in self.ana_samples:
                                        if ana_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Check SAL samples
                                if not sample_matches_custom_type and hasattr(self, 'sal_samples') and self.sal_samples:
                                    for sal_prefix in self.sal_samples:
                                        if sal_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Check PT samples
                                if not sample_matches_custom_type and hasattr(self, 'pt_samples') and self.pt_samples:
                                    for pt_prefix in self.pt_samples:
                                        if pt_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Check TIS samples
                                if not sample_matches_custom_type and hasattr(self, 'tis_samples') and self.tis_samples:
                                    for tis_prefix in self.tis_samples:
                                        if tis_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Check INT samples
                                if not sample_matches_custom_type and hasattr(self, 'int_samples') and self.int_samples:
                                    for int_prefix in self.int_samples:
                                        if int_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                    for int_prefix in self.int_samples:
                                        if int_prefix in str(col_value):
                                            sample_matches_custom_type = True
                                            break
                                
                                # Check for generic/numeric samples if enabled
                                if not sample_matches_custom_type:
                                    # Accept numeric strings or strings with digits
                                    str_val = str(col_value)
                                    if str_val.isdigit() or (any(c.isdigit() for c in str_val) and len(str_val) >= 3):
                                        sample_matches_custom_type = True
                                
                                # Only include samples that match your custom sample type settings
                                if sample_matches_custom_type:
                                    if transformed_col_value in existing_columns or transformed_col_value in self.previous_done_samples:
                                        # Sample already in database, mark it for status reset to ensure it's re-processed
                                        # with potentially new category settings
                                        re_enable_samples.append(transformed_col_value)
                                    else:
                                        # New sample
                                        new_row = pd.DataFrame({
                                            'Host_Id': [host_value], 
                                            'Sample_Id': [transformed_col_value], 
                                            'Status': ['in process']
                                        })
                                        output_data = pd.concat([output_data, new_row], ignore_index=True)
                except Exception as e:
                    processing_progress['message'] = f'Error processing {sheet_name}: {str(e)}'

            # Save output data
            if not output_data.empty:
                try:
                    if not os.path.exists(destination_file):
                        with pd.ExcelWriter(destination_file, engine='openpyxl') as writer:
                            if 'Sample_Id' in output_data.columns:
                                output_data['Sample_Id'] = output_data['Sample_Id'].apply(self.transform_sample_id)
                            output_data.to_excel(writer, sheet_name="Extraction", index=False)
                    else:
                        with pd.ExcelWriter(destination_file, engine='openpyxl', mode='a', 
                                          if_sheet_exists='overlay') as writer:
                            existing_data = pd.read_excel(destination_file, sheet_name="Extraction")
                            if 'Status' not in existing_data.columns:
                                existing_data['Status'] = 'unknown'
                            
                            # Combine new data
                            if not output_data.empty:
                                combined_data = pd.concat([existing_data, output_data], ignore_index=True)
                                combined_data.drop_duplicates(subset=['Sample_Id'], keep='first', inplace=True)
                            else:
                                combined_data = existing_data
                            
                            # RESET status for samples that are in the current source file
                            # This allows re-categorization and re-processing
                            if re_enable_samples:
                                combined_data.loc[combined_data['Sample_Id'].isin(re_enable_samples), 'Status'] = 'in process'
                                processing_progress['message'] = f'Reset status for {len(re_enable_samples)} samples to "in process" for re-processing'
                            
                            if 'Sample_Id' in combined_data.columns:
                                combined_data['Sample_Id'] = combined_data['Sample_Id'].apply(self.transform_sample_id)
                            combined_data.to_excel(writer, sheet_name="Extraction", index=False)
                    
                    if not output_data.empty:
                        processing_progress['message'] = f'Added {len(output_data)} new rows'
                except Exception as e:
                    processing_progress['error'] = f'Error saving data: {str(e)}'
                    return

            # Process ANA/SAL/PT samples
            current_step += 2
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            self.process_ana_sal_samples(destination_file, current_step, total_steps)
            if processing_progress.get('error'):
                return

            # Process TIS/INT samples
            current_step += 2
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            self.process_tis_int_samples(destination_file, current_step, total_steps)
            if processing_progress.get('error'):
                return

            # Create combined check list with all samples
            current_step += 1
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            self.create_combined_check_list(self.swab_check_list, self.tis_int_check_list)

            # Create PCR and cDNA plates if any plates were generated (Snapshot logic)
            current_step += 1
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            if (hasattr(self, 'extraction_plate_snapshots') and self.extraction_plate_snapshots) or len(output_data) > 0:
                new_sample_ids = output_data['Sample_Id'].tolist() if not output_data.empty else []
                self.create_pcr_cdna_plates(destination_file, new_sample_ids)
            else:
                processing_progress['message'] = 'No plates generated - skipping PCR/cDNA'

            # Update status to "Done" (AFTER PCR/cDNA plates are created)
            current_step += 1
            processing_progress['progress'] = int((current_step / total_steps) * 100)
            processing_progress['message'] = 'Updating sample status...'
            
            try:
                df = pd.read_excel(destination_file, sheet_name='Extraction')
                # Mark ALL samples that were processed in this run as Done
                # We identify them by the 'in process' status currently in the file
                processed_ids = df[df['Status'] == 'in process']['Sample_Id'].tolist()
                df.loc[df['Sample_Id'].isin(processed_ids), 'Status'] = 'Done'
                
                with pd.ExcelWriter(destination_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name="Extraction", index=False)
                processing_progress['message'] = f'Updated status to "Done" for {len(processed_ids)} samples'
            except Exception as e:
                processing_progress['error'] = f'Error updating status: {str(e)}'

            processing_progress['status'] = 'completed'
            processing_progress['progress'] = 100
            processing_progress['message'] = 'Processing completed successfully! All tables created.'

        except Exception as e:
            processing_progress['status'] = 'error'
            processing_progress['error'] = f'Processing failed: {str(e)}'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

@extraction_bp.route('/')
def index():
    return render_template('extraction/index.html')

@extraction_bp.route('/settings')
def settings_page():
    return render_template('extraction/settings.html')

@extraction_bp.route('/viewer')
def viewer_page():
    return render_template('extraction/viewer.html')

@extraction_bp.route('/api/settings', methods=['GET'])
def get_settings():
    """Return current settings as JSON"""
    try:
        # Load from file if exists, otherwise return empty settings
        settings_file = get_settings_file()
        if os.path.exists(settings_file):
            with open(settings_file, 'r') as f:
                return json.load(f)
        else:
            # Return empty settings for UI to start fresh
            return {
                'prefix_mapping': {},
                'pool_samples': [],
                'no_pool_samples': [],
                'samples_per_pool': None,
                'sample_id_pattern': None,
                'ana_samples': [],
                'sal_samples': [],
                'pt_samples': [],
                'tis_samples': [],
                'int_samples': [],
                'enable_h2o_random': False,
                'h2o_count': 2,
                'h2o_position_preference': 'random'
            }
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@extraction_bp.route('/api/settings', methods=['POST'])
def save_settings():
    """Save settings to file"""
    try:
        settings = request.get_json()
        settings_file = get_settings_file()
        
        with open(settings_file, 'w') as f:
            json.dump(settings, f, indent=2)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@extraction_bp.route('/upload', methods=['POST'])
def upload_files():
    if 'source_file' not in request.files:
        return jsonify({'error': 'No source file uploaded'}), 400
    
    source_file = request.files['source_file']
    if source_file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(source_file.filename):
        return jsonify({'error': 'Invalid file type. Please upload Excel files only'}), 400

    # Save source file
    source_filename = secure_filename(source_file.filename)
    source_path = os.path.join(get_upload_folder(), source_filename)
    source_file.save(source_path)

    # Handle destination file - check for custom output or existing default
    custom_output = request.form.get('custom_output', '').strip()
    dest_path = None
    
    if custom_output:
        # Use custom output file if specified
        if not custom_output.endswith('.xlsx'):
            custom_output += '.xlsx'
        dest_path = os.path.join(get_output_folder(), custom_output)
        processing_progress['message'] = f"Using custom output file: {custom_output}"
    else:
        # Check for existing default output file (most recent)
        try:
            existing_files = [f for f in os.listdir(get_output_folder()) 
                           if f.startswith('Extraction_Output_') and f.endswith('.xlsx')]
            if existing_files:
                # Sort by modification time to get the most recent
                existing_files.sort(key=lambda x: os.path.getmtime(os.path.join(get_output_folder(), x)), reverse=True)
                dest_path = os.path.join(get_output_folder(), existing_files[0])
                processing_progress['message'] = f"Using existing output file: {existing_files[0]}"
        except Exception as e:
            print(f"DEBUG: Error checking existing output files: {e}")
    
    # If no existing file found, create new one
    if dest_path is None or not os.path.exists(dest_path):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        dest_filename = f"Extraction_Output_{timestamp}.xlsx"
        dest_path = os.path.join(get_output_folder(), dest_filename)
        processing_progress['message'] = f"Creating new output file: {dest_filename}"
        
        # Initialize new destination file
        try:
            with pd.ExcelWriter(dest_path, engine='openpyxl') as writer:
                pd.DataFrame(columns=['Host_Id', 'Sample_Id', 'Status']).to_excel(
                    writer, sheet_name='Extraction', index=False
                )
        except Exception as e:
            return jsonify({'error': f'Error initializing destination file: {str(e)}'}), 500
    else:
        # Verify existing file has required structure
        try:
            df = pd.read_excel(dest_path, sheet_name='Extraction')
            if 'Sample_Id' not in df.columns or 'Status' not in df.columns:
                # Add missing columns if needed
                if 'Sample_Id' not in df.columns:
                    df['Sample_Id'] = ''
                if 'Status' not in df.columns:
                    df['Status'] = 'unknown'
                
                # Save back with corrected structure
                with pd.ExcelWriter(dest_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name='Extraction', index=False)
                processing_progress['message'] = f"Updated existing output file structure"
        except Exception as e:
            processing_progress['message'] = f"Warning: Could not verify existing file structure: {str(e)}"

    # Read Excel sheets
    try:
        xl = pd.ExcelFile(source_path)
        sheets_data = {}
        for sheet in xl.sheet_names:
            df = pd.read_excel(source_path, sheet_name=sheet)
            sheets_data[sheet] = {
                'columns': df.columns.tolist(),
                'preview': df.head().to_html(classes='table table-striped table-sm', index=False)
            }
        
        return jsonify({
            'source_path': source_path,
            'dest_path': dest_path,
            'sheets': sheets_data
        })
    except Exception as e:
        return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 500

@extraction_bp.route('/process', methods=['POST'])
def process_files():
    data = request.get_json()
    source_path = data.get('source_path')
    dest_path = data.get('dest_path')
    sheets_info = data.get('sheets_info', {})

    if not source_path or not dest_path or not sheets_info:
        return jsonify({'error': 'Missing required parameters'}), 400

    # Reset progress
    global processing_progress
    processing_progress = {'status': 'idle', 'progress': 0, 'message': '', 'error': ''}

    # Start processing in background thread
    processor = ExtractionProcessor()
    thread = threading.Thread(target=processor.process_files, args=(source_path, dest_path, sheets_info))
    thread.daemon = True
    thread.start()

    return jsonify({'message': 'Processing started'})

@extraction_bp.route('/progress')
def get_progress():
    return jsonify(processing_progress)

@extraction_bp.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(get_output_folder(), filename)
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'error': f'Error downloading file: {str(e)}'}), 500

@extraction_bp.route('/outputs')
def list_outputs():
    try:
        files = []
        for filename in os.listdir(get_output_folder()):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(get_output_folder(), filename)
                stat = os.stat(file_path)
                files.append({
                    'name': filename,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
        files.sort(key=lambda x: x['modified'], reverse=True)
        return jsonify({'files': files})
    except Exception as e:
        return jsonify({'error': f'Error listing files: {str(e)}'}), 500

@extraction_bp.route('/api/extraction-data')
def get_extraction_data():
    """Get extraction table data showing SourceId, SampleId, Plate, Pool, and Status"""
    try:
        # Find the most recent output file
        output_folder = get_output_folder()
        output_files = [f for f in os.listdir(output_folder) 
                       if f.startswith('Extraction_Output_') and f.endswith('.xlsx')]
        
        if not output_files:
            return jsonify({'error': 'No extraction output file found', 'data': []})
        
        # Get most recent file
        output_files.sort(key=lambda x: os.path.getmtime(os.path.join(output_folder, x)), reverse=True)
        output_file = os.path.join(output_folder, output_files[0])
        
        # Read extraction sheet
        df = pd.read_excel(output_file, sheet_name='Extraction', engine='openpyxl')
        
        # Get sample data with status
        result = []
        for _, row in df.iterrows():
            result.append({
                'host_id': str(row.get('Host_Id', '')) if pd.notna(row.get('Host_Id', '')) else '',
                'sample_id': str(row.get('Sample_Id', '')) if pd.notna(row.get('Sample_Id', '')) else '',
                'pool': str(row.get('Pool', '')) if pd.notna(row.get('Pool', '')) else '',
                'status': str(row.get('Status', '')) if pd.notna(row.get('Status', '')) else ''
            })
        
        return jsonify({
            'success': True,
            'file': output_files[0],
            'total_samples': len(result),
            'done_count': sum(1 for r in result if r['status'] == 'Done'),
            'in_process_count': sum(1 for r in result if r['status'] == 'in process'),
            'data': result
        })
        
    except Exception as e:
        return jsonify({'error': f'Error reading extraction data: {str(e)}', 'data': []})

@extraction_bp.route('/api/plate-data/<plate_type>')
def get_plate_data(plate_type):
    """Get plate data from Extraction, PCR, or cDNA tables showing SampleId and Plate No"""
    try:
        if plate_type.lower() == 'extraction':
            file_path = 'Extraction/Extraction_Tables.xlsx'
        elif plate_type.lower() == 'pcr':
            file_path = 'Extraction/PCR/PCR_Tables.xlsx'
        elif plate_type.lower() == 'cdna':
            file_path = 'Extraction/cDNA/cDNA_Tables.xlsx'
        else:
            return jsonify({'error': f'Unknown plate type: {plate_type}'}), 400
        
        if not os.path.exists(file_path):
            return jsonify({'error': f'{plate_type} tables file not found', 'data': []})
        # Read all sheets (plates)
        xl = pd.ExcelFile(file_path, engine='openpyxl')
        result = []
        
        for sheet_name in xl.sheet_names:
            # Read the plate data
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            
            # Track samples seen in this sheet to avoid duplicates (multiple tables per sheet)
            seen_in_sheet = set()
            
            # Find table header rows (rows containing "1 2 3 4 5 6 7 8 9 10 11" or "1 2 3 ... 12")
            # These are the column number headers, data is in the 8 rows below
            for row_idx, row in df.iterrows():
                # Check if this row is a column header row (contains 1, 2, 3... up to 11 or 12)
                row_values = [str(cell).strip() if pd.notna(cell) else '' for cell in row]
                
                # Look for sequential numbers 1-11 or 1-12 in the row
                is_header_row = False
                num_columns = 0
                
                # Count exact matches for numbers 1-12 (not substrings or decimals)
                exact_numbers = set()
                for v in row_values:
                    try:
                        if v.isdigit():
                            num = int(v)
                            if 1 <= num <= 12:
                                exact_numbers.add(num)
                    except:
                        pass
                
                # Check if we have sequential numbers 1-6, 1-8, 1-10, or 1-11 present (at least 5 out of 6, 7 out of 8, or 10 out of 11/12)
                if len(exact_numbers) >= 5:
                    is_header_row = True
                    if 12 in exact_numbers:
                        num_columns = 12
                    elif 11 in exact_numbers:
                        num_columns = 11
                    elif 10 in exact_numbers:
                        num_columns = 10
                    elif 9 in exact_numbers:
                        num_columns = 9
                    elif 8 in exact_numbers:
                        num_columns = 8
                    elif 7 in exact_numbers:
                        num_columns = 7
                    else:
                        num_columns = 6
                    
                    # Detect if column 0 has row labels (A-H) or is part of data
                    # If "1" is at index 0, data starts at column 0; if at index 1, column 0 has row labels
                    first_col_idx = 0
                    for idx, v in enumerate(row_values):
                        if v == '1':
                            first_col_idx = idx
                            break
                
                if is_header_row:
                    # Check if samples span 2 rows (prefix on row 1, number on row 2)
                    first_data_row_idx = row_idx + 1
                    if first_data_row_idx < len(df):
                        first_row = df.iloc[first_data_row_idx]
                        # Use first_col_idx to check the correct first data cell
                        test_cell = first_row.iloc[first_col_idx] if len(first_row) > first_col_idx else None
                        if pd.notna(test_cell):
                            test_str = str(test_cell).replace('\n', '_').strip()
                            is_two_row_format = '_' not in test_str and '\n' not in str(test_cell) and not test_str.startswith('Pool')
                        else:
                            is_two_row_format = False
                    else:
                        is_two_row_format = False
                    
                    if is_two_row_format:
                        # Read 16 rows (8 samples x 2 rows each), combine pairs
                        for sample_idx in range(8):
                            prefix_row_idx = row_idx + 1 + (sample_idx * 2)
                            number_row_idx = prefix_row_idx + 1
                            
                            if number_row_idx >= len(df):
                                break
                            
                            prefix_row = df.iloc[prefix_row_idx]
                            number_row = df.iloc[number_row_idx]
                            
                            # Use first_col_idx to start reading from correct column
                            for col_idx in range(first_col_idx, min(first_col_idx + num_columns, len(prefix_row))):
                                prefix_cell = prefix_row.iloc[col_idx] if col_idx < len(prefix_row) else None
                                number_cell = number_row.iloc[col_idx] if col_idx < len(number_row) else None
                                
                                if pd.notna(prefix_cell) and pd.notna(number_cell):
                                    prefix_str = str(prefix_cell).strip()
                                    number_str = str(number_cell).strip()
                                    
                                    if prefix_str and number_str and prefix_str != 'H2O' and number_str != 'H2O':
                                        # Format the number to 3 digits (or keep 4+ digits) 
                                        # Only format if it's a sample number (not column header)
                                        if not number_str.isdigit() or len(number_str) > 1 or any(c.isalpha() for c in prefix_str):
                                            formatted_number = format_sample_number(number_str)
                                        else:
                                            formatted_number = number_str
                                        cell_str = f"{prefix_str}_{formatted_number}"
                                        if cell_str not in seen_in_sheet:
                                            seen_in_sheet.add(cell_str)
                                            result.append({
                                                'sample_id': cell_str,
                                                'plate_no': sheet_name
                                            })
                    else:
                        # Single-row format: Read 8 rows
                        for data_row_offset in range(1, 9):
                            data_row_idx = row_idx + data_row_offset
                            if data_row_idx >= len(df):
                                break
                            
                            data_row = df.iloc[data_row_idx]
                            
                            # Use first_col_idx to start reading from correct column
                            for col_idx in range(first_col_idx, min(first_col_idx + num_columns, len(data_row))):
                                cell = data_row.iloc[col_idx] if col_idx < len(data_row) else None
                                
                                if pd.notna(cell) and cell != '' and cell != 'H2O':
                                    cell_str = str(cell).replace('\n', '_').strip()
                                    
                                    if len(cell_str) <= 1 or cell_str.isdigit():
                                        continue
                                    
                                    # Format sample numbers in single-row format too
                                    if '_' in cell_str:
                                        parts = cell_str.split('_')
                                        if len(parts) == 2:
                                            prefix, number = parts
                                            formatted_number = format_sample_number(number)
                                            cell_str = f"{prefix}_{formatted_number}"
                                    
                                    skip_keywords = ['to', 'Sample', 'Plate', 'Date', 'Perform', 'Extraction', 'cDNA', 'PCR', 'by:']
                                    if any(kw.lower() in cell_str.lower() for kw in skip_keywords):
                                        continue
                                    
                                    if '_' in cell_str or cell_str.startswith('Pool'):
                                        if cell_str not in seen_in_sheet:
                                            seen_in_sheet.add(cell_str)
                                            result.append({
                                                'sample_id': cell_str,
                                                'plate_no': sheet_name
                                            })
        
        return jsonify({
            'success': True,
            'plate_type': plate_type,
            'total_samples': len(result),
            'data': result
        })
        
    except Exception as e:
        return jsonify({'error': f'Error reading plate data: {str(e)}', 'plates': []})

@extraction_bp.route('/api/done-samples')
def get_done_samples():
    """Get list of samples that are already marked as Done (to be skipped)"""
    try:
        output_folder = get_output_folder()
        output_files = [f for f in os.listdir(output_folder) 
                       if f.startswith('Extraction_Output_') and f.endswith('.xlsx')]
        
        if not output_files:
            return jsonify({'done_samples': [], 'count': 0})
        
        output_files.sort(key=lambda x: os.path.getmtime(os.path.join(output_folder, x)), reverse=True)
        output_file = os.path.join(output_folder, output_files[0])
        
        processor = ExtractionProcessor()
        done_samples = list(processor.get_existing_done_samples(output_file))
        
        return jsonify({
            'success': True,
            'file': output_files[0],
            'done_samples': done_samples,
            'count': len(done_samples)
        })
        
    except Exception as e:
        return jsonify({'error': str(e), 'done_samples': [], 'count': 0})

@extraction_bp.route('/api/export-plate-data/<plate_type>')
def export_plate_data(plate_type):
    """Export plate data to Excel file with SampleId and Plate No columns"""
    try:
        if plate_type.lower() == 'extraction':
            file_path = 'Extraction/Extraction_Tables.xlsx'
        elif plate_type.lower() == 'pcr':
            file_path = 'Extraction/PCR/PCR_Tables.xlsx'
        elif plate_type.lower() == 'cdna':
            file_path = 'Extraction/cDNA/cDNA_Tables.xlsx'
        else:
            return jsonify({'error': f'Unknown plate type: {plate_type}'}), 400
        
        if not os.path.exists(file_path):
            return jsonify({'error': f'{plate_type} tables file not found'}), 404
        
        # Read all sheets and collect sample data
        xl = pd.ExcelFile(file_path, engine='openpyxl')
        result = []
        
        for sheet_name in xl.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            
            for row_idx, row in df.iterrows():
                for col_idx, cell in enumerate(row):
                    if pd.notna(cell) and cell != '' and cell != 'H2O':
                        cell_str = str(cell).replace('\n', '_')
                        if '_' in cell_str or cell_str.startswith('Pool'):
                            result.append({
                                'Sample_Id': cell_str,
                                'Plate_No': sheet_name
                            })
        
        # Create DataFrame and export to Excel
        export_df = pd.DataFrame(result)
        
        # Create temporary export file
        export_folder = get_output_folder()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_filename = f'{plate_type.upper()}_Sample_List_{timestamp}.xlsx'
        export_path = os.path.join(export_folder, export_filename)
        
        export_df.to_excel(export_path, index=False, sheet_name='Sample_List')
        
        return send_file(export_path, as_attachment=True, download_name=export_filename)
        
    except Exception as e:
        return jsonify({'error': f'Error exporting data: {str(e)}'}), 500

@extraction_bp.route('/api/upload-table', methods=['POST'])
def upload_table():
    """Upload custom extraction table file and extract SampleId and Plate No"""
    try:
        if 'file' not in request.files:
            print("DEBUG upload_table: No 'file' in request.files")
            return jsonify({'error': 'No file uploaded', 'data': []})
        
        file = request.files['file']
        if file.filename == '':
            print("DEBUG upload_table: Empty filename")
            return jsonify({'error': 'No file selected', 'data': []})
        
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        upload_path = os.path.join(get_upload_folder(), filename)
        file.save(upload_path)
        print(f"DEBUG upload_table: Saved file to {upload_path}")
        
        # Also copy to Extraction folder if it's an Extraction_Tables.xlsx file
        if filename == 'Extraction_Tables.xlsx':
            import shutil
            extraction_folder = os.path.join(get_base_path(), 'Extraction')
            os.makedirs(extraction_folder, exist_ok=True)
            extraction_path = os.path.join(extraction_folder, 'Extraction_Tables.xlsx')
            shutil.copy2(upload_path, extraction_path)
        
        # Read all sheets and extract sample data
        xl = pd.ExcelFile(upload_path, engine='openpyxl')
        print(f"DEBUG upload_table: Found sheets: {xl.sheet_names}")
        result = []
        
        for sheet_name in xl.sheet_names:
            print(f"DEBUG upload_table: Processing sheet '{sheet_name}'...")
            df = pd.read_excel(upload_path, sheet_name=sheet_name, header=None, engine='openpyxl')
            
            # Track samples seen in this sheet to avoid duplicates
            seen_in_sheet = set()
            
            # Find table header rows (containing "1 2 3 4 5 6 7 8 9 10 11" or "... 12")
            for row_idx, row in df.iterrows():
                row_values = [str(cell).strip() if pd.notna(cell) else '' for cell in row]
                
                # Check for header row pattern
                is_header_row = False
                num_columns = 0
                
                # Count exact matches for numbers 1-12 (not substrings or decimals)
                exact_numbers = set()
                for v in row_values:
                    try:
                        if v.isdigit():
                            num = int(v)
                            if 1 <= num <= 12:
                                exact_numbers.add(num)
                    except:
                        pass
                
                # Check if we have sequential numbers 1-6, 1-8, 1-10, or 1-11 present (at least 5 out of 6, 7 out of 8, or 10 out of 11/12)
                if len(exact_numbers) >= 5:
                    is_header_row = True
                    print(f"DEBUG upload_table: Found header row at index {row_idx} with numbers {exact_numbers}")
                    if 12 in exact_numbers:
                        num_columns = 12
                    elif 11 in exact_numbers:
                        num_columns = 11
                    elif 10 in exact_numbers:
                        num_columns = 10
                    elif 9 in exact_numbers:
                        num_columns = 9
                    elif 8 in exact_numbers:
                        num_columns = 8
                    elif 7 in exact_numbers:
                        num_columns = 7
                    else:
                        num_columns = 6
                    
                    # Detect if column 0 has row labels or is part of data
                    first_col_idx = 0
                    for idx, v in enumerate(row_values):
                        if v == '1':
                            first_col_idx = idx
                            break
                
                if is_header_row:
                    # Check if samples span 2 rows
                    first_data_row_idx = row_idx + 1
                    if first_data_row_idx < len(df):
                        first_row = df.iloc[first_data_row_idx]
                        test_cell = first_row.iloc[first_col_idx] if len(first_row) > first_col_idx else None
                        if pd.notna(test_cell):
                            test_str = str(test_cell).replace('\n', '_').strip()
                            is_two_row_format = '_' not in test_str and '\n' not in str(test_cell) and not test_str.startswith('Pool')
                        else:
                            is_two_row_format = False
                    else:
                        is_two_row_format = False
                    
                    if is_two_row_format:
                        # Read 16 rows (8 samples x 2 rows each)
                        for sample_idx in range(8):
                            prefix_row_idx = row_idx + 1 + (sample_idx * 2)
                            number_row_idx = prefix_row_idx + 1
                            
                            if number_row_idx >= len(df):
                                break
                            
                            prefix_row = df.iloc[prefix_row_idx]
                            number_row = df.iloc[number_row_idx]
                            
                            for col_idx in range(first_col_idx, min(first_col_idx + num_columns, len(prefix_row))):
                                prefix_cell = prefix_row.iloc[col_idx] if col_idx < len(prefix_row) else None
                                number_cell = number_row.iloc[col_idx] if col_idx < len(number_row) else None
                                
                                if pd.notna(prefix_cell) and pd.notna(number_cell):
                                    prefix_str = str(prefix_cell).strip()
                                    number_str = str(number_cell).strip()
                                    
                                    if prefix_str and number_str and prefix_str != 'H2O' and number_str != 'H2O':
                                        # Format the number to 3 digits (or keep 4+ digits) 
                                        # Only format if it's a sample number (not column header)
                                        if not number_str.isdigit() or len(number_str) > 1 or any(c.isalpha() for c in prefix_str):
                                            formatted_number = format_sample_number(number_str)
                                        else:
                                            formatted_number = number_str
                                        cell_str = f"{prefix_str}_{formatted_number}"
                                        if cell_str not in seen_in_sheet:
                                            seen_in_sheet.add(cell_str)
                                            result.append({
                                                'sample_id': cell_str,
                                                'plate_no': sheet_name
                                            })
                    else:
                        # Single-row format: Read 8 rows
                        for data_row_offset in range(1, 9):
                            data_row_idx = row_idx + data_row_offset
                            if data_row_idx >= len(df):
                                break
                            
                            data_row = df.iloc[data_row_idx]
                            
                            for col_idx in range(first_col_idx, min(first_col_idx + num_columns, len(data_row))):
                                cell = data_row.iloc[col_idx] if col_idx < len(data_row) else None
                                
                                if pd.notna(cell) and cell != '' and cell != 'H2O':
                                    cell_str = str(cell).replace('\n', '_').strip()
                                    
                                    if len(cell_str) <= 1 or cell_str.isdigit():
                                        continue
                                    
                                    # Format sample numbers in single-row format too
                                    if '_' in cell_str:
                                        parts = cell_str.split('_')
                                        if len(parts) == 2:
                                            prefix, number = parts
                                            formatted_number = format_sample_number(number)
                                            cell_str = f"{prefix}_{formatted_number}"
                                    
                                    skip_keywords = ['to', 'Sample', 'Plate', 'Date', 'Perform', 'Extraction', 'cDNA', 'PCR', 'by:']
                                    if any(kw.lower() in cell_str.lower() for kw in skip_keywords):
                                        continue
                                    
                                    if '_' in cell_str or cell_str.startswith('Pool'):
                                        if cell_str not in seen_in_sheet:
                                            seen_in_sheet.add(cell_str)
                                            result.append({
                                                'sample_id': cell_str,
                                                'plate_no': sheet_name
                                            })
        
        print(f"DEBUG upload_table: FINAL RESULT: {len(result)} samples found across {len(xl.sheet_names)} sheets")
        
        # ── FALLBACK: Flat-table format (columns like SampleId, PlateNo) ──
        if len(result) == 0:
            print("DEBUG upload_table: Grid parser found nothing. Trying flat-table fallback...")
            for sheet_name in xl.sheet_names:
                try:
                    df_flat = pd.read_excel(upload_path, sheet_name=sheet_name, engine='openpyxl')
                    cols_lower = {c: c for c in df_flat.columns}
                    cols_map = {c.lower().replace(' ', '').replace('_', ''): c for c in df_flat.columns}
                    
                    # Find sample ID column
                    sample_col = None
                    for key in ['sampleid', 'sample', 'id', 'samplename']:
                        if key in cols_map:
                            sample_col = cols_map[key]
                            break
                    
                    # Find plate column
                    plate_col = None
                    for key in ['plateno', 'plate', 'platenumber', 'sheet', 'group']:
                        if key in cols_map:
                            plate_col = cols_map[key]
                            break
                    
                    if sample_col:
                        print(f"DEBUG upload_table: Flat-table detected! sample_col='{sample_col}', plate_col='{plate_col}'")
                        for _, row in df_flat.iterrows():
                            sid = row.get(sample_col)
                            if pd.notna(sid):
                                sid_str = str(sid).strip()
                                if sid_str and sid_str.lower() != 'nan':
                                    plate_val = str(row.get(plate_col, sheet_name)).strip() if plate_col and pd.notna(row.get(plate_col)) else sheet_name
                                    result.append({
                                        'sample_id': sid_str,
                                        'plate_no': plate_val
                                    })
                        print(f"DEBUG upload_table: Flat-table yielded {len(result)} samples")
                    else:
                        print(f"DEBUG upload_table: No recognizable sample column in sheet '{sheet_name}'. Columns: {list(df_flat.columns)}")
                except Exception as e:
                    print(f"DEBUG upload_table: Flat-table fallback error on sheet '{sheet_name}': {e}")
        
        if len(result) == 0:
            print("DEBUG upload_table: WARNING - No samples found after all parsing attempts!")
        return jsonify({
            'success': True,
            'filename': filename,
            'total_samples': len(result),
            'data': result
        })
        
    except Exception as e:
        print(f"DEBUG upload_table: EXCEPTION: {str(e)}")
        return jsonify({'error': f'Error processing file: {str(e)}', 'data': []})

@extraction_bp.route('/api/export-uploaded-data', methods=['POST'])
def export_uploaded_data():
    """Export uploaded data as Excel file"""
    try:
        data = request.get_json()
        if not data or 'data' not in data:
            return jsonify({'error': 'No data provided'}), 400
        
        items = data['data']
        
        # Create DataFrame
        df = pd.DataFrame([{
            'Sample_Id': item.get('sample_id', ''),
            'Plate_No': item.get('plate_no', '')
        } for item in items])
        
        # Create Excel file in memory
        from io import BytesIO
        output = BytesIO()
        df.to_excel(output, index=False, sheet_name='Sample_List')
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Extraction_Sample_List.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error exporting data: {str(e)}'}), 500
